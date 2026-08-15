from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DB_FILE = os.environ.get(
    "CHATBOT_DB",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "chatbot_company.db"),
)


def now_text() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value))


def normalize_price(value: Any) -> int | None:
    raw = re.sub(r"[^0-9-]+", "", clean_text(value))
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        return None


def normalize_date(value: Any) -> str:
    text = clean_text(value)
    digits = re.sub(r"\D+", "", text)
    if len(digits) >= 8:
        return digits[:8]
    return text


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def canonical_json(record: dict[str, Any]) -> str:
    return json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def row_key(record: dict[str, Any]) -> str:
    parts = [
        record["portal_price_business_type"],
        record["product_classification_no"],
        record["product_classification_name"],
        record["product_identifier_no"],
        record["korean_product_name"],
        record["unit"],
        record["delivery_condition_name"],
        record["supply_region_name"],
        record["posted_date"],
    ]
    return sha256_text("\n".join(parts))


def current_key(record: dict[str, Any]) -> str:
    parts = [
        record["portal_price_business_type"],
        record["product_classification_no"],
        record["product_identifier_no"],
        record["unit"],
        record["delivery_condition_name"],
        record["supply_region_name"],
    ]
    return sha256_text("\n".join(parts))


def ensure_meta_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS etl_job_log (
            job_id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_name TEXT NOT NULL,
            source_name TEXT NOT NULL,
            started_at DATETIME NOT NULL,
            finished_at DATETIME,
            status TEXT NOT NULL,
            input_row_count INTEGER DEFAULT 0,
            inserted_count INTEGER DEFAULT 0,
            updated_count INTEGER DEFAULT 0,
            skipped_count INTEGER DEFAULT 0,
            error_count INTEGER DEFAULT 0,
            error_message TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS source_manifest (
            source_id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_name TEXT UNIQUE NOT NULL,
            source_type TEXT NOT NULL,
            source_url_or_file TEXT,
            source_refreshed_at DATETIME,
            row_count INTEGER DEFAULT 0,
            checksum TEXT,
            status TEXT,
            error_message TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


def ensure_schema(conn: sqlite3.Connection) -> None:
    ensure_meta_schema(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facility_material_price_import_log (
            import_id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file_path TEXT NOT NULL,
            source_file_name TEXT NOT NULL,
            source_file_hash TEXT NOT NULL,
            sheet_name TEXT,
            output_date TEXT,
            started_at DATETIME NOT NULL,
            finished_at DATETIME,
            status TEXT NOT NULL,
            total_row_count INTEGER DEFAULT 0,
            inserted_count INTEGER DEFAULT 0,
            updated_count INTEGER DEFAULT 0,
            unchanged_count INTEGER DEFAULT 0,
            missing_count INTEGER DEFAULT 0,
            current_active_count INTEGER DEFAULT 0,
            dictionary_count INTEGER DEFAULT 0,
            error_message TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facility_material_price_master (
            source_row_key TEXT PRIMARY KEY,
            row_hash TEXT NOT NULL,
            product_classification_no TEXT,
            product_classification_name TEXT,
            product_identifier_no TEXT,
            korean_product_name TEXT,
            unit TEXT,
            price_amount INTEGER,
            delivery_condition_name TEXT,
            supply_region_name TEXT,
            pps_department_name TEXT,
            posted_date TEXT,
            is_deleted INTEGER NOT NULL DEFAULT 0,
            delete_yn TEXT,
            vat_inclusion_name TEXT,
            portal_price_business_type TEXT,
            current_group_key TEXT NOT NULL,
            raw_json TEXT NOT NULL,
            first_import_id INTEGER NOT NULL,
            last_import_id INTEGER NOT NULL,
            first_seen_at DATETIME NOT NULL,
            last_seen_at DATETIME NOT NULL,
            missing_in_latest_import INTEGER NOT NULL DEFAULT 0,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_master_clsfc ON facility_material_price_master(product_classification_no)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_master_idnt ON facility_material_price_master(product_identifier_no)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_master_name ON facility_material_price_master(product_classification_name, korean_product_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_master_group ON facility_material_price_master(current_group_key)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_master_latest ON facility_material_price_master(last_import_id, missing_in_latest_import)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facility_material_price_change_history (
            history_id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL,
            source_row_key TEXT NOT NULL,
            change_type TEXT NOT NULL,
            previous_row_hash TEXT,
            new_row_hash TEXT,
            previous_raw_json TEXT,
            new_raw_json TEXT,
            changed_at DATETIME NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_history_import ON facility_material_price_change_history(import_id)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facility_material_price_current (
            current_group_key TEXT PRIMARY KEY,
            product_classification_no TEXT,
            product_classification_name TEXT,
            product_identifier_no TEXT,
            korean_product_name TEXT,
            unit TEXT,
            latest_price_amount INTEGER,
            delivery_condition_name TEXT,
            supply_region_name TEXT,
            pps_department_name TEXT,
            latest_posted_date TEXT,
            vat_inclusion_name TEXT,
            portal_price_business_type TEXT,
            source_row_key TEXT NOT NULL,
            source_row_hash TEXT NOT NULL,
            generated_at DATETIME NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_current_name ON facility_material_price_current(product_classification_name, korean_product_name)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_current_clsfc ON facility_material_price_current(product_classification_no)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_price_current_idnt ON facility_material_price_current(product_identifier_no)")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS facility_material_item_dictionary (
            dictionary_key TEXT PRIMARY KEY,
            product_classification_no TEXT,
            product_identifier_no TEXT,
            product_classification_name TEXT,
            korean_product_name TEXT,
            portal_price_business_types TEXT,
            units TEXT,
            active_price_count INTEGER NOT NULL DEFAULT 0,
            historical_row_count INTEGER NOT NULL DEFAULT 0,
            latest_posted_date TEXT,
            min_active_price_amount INTEGER,
            max_active_price_amount INTEGER,
            search_text TEXT NOT NULL,
            generated_at DATETIME NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_facility_material_item_dictionary_search ON facility_material_item_dictionary(search_text)")
    conn.commit()


def find_header_row(ws: Any, max_rows: int = 30) -> int:
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True), start=1):
        values = [clean_text(v) for v in row]
        joined = "|".join(values)
        if "물품분류" in joined and "물품식별" in joined and "가격" in joined and "게시일자" in joined:
            return idx
    raise ValueError("header row not found: expected 물품분류/물품식별/가격/게시일자")


def extract_output_date(ws: Any) -> str:
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for value in row:
            text = clean_text(value)
            match = re.search(r"출력일자\s*:\s*(\d{4}-\d{2}-\d{2}|\d{8})", text)
            if match:
                return normalize_date(match.group(1))
    return ""


def iter_records(path: Path) -> tuple[str, str, list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        header_row = find_header_row(ws)
        output_date = extract_output_date(ws)
        sheet_name = ws.title
        records: list[dict[str, Any]] = []
        for excel_row_no, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            values = [clean_text(v) for v in row[:13]]
            if not any(values):
                continue
            while len(values) < 13:
                values.append("")
            delete_yn = values[10].upper()
            record = {
                "excel_row_no": excel_row_no,
                "product_classification_no": values[0],
                "product_classification_name": normalize_space(values[1]),
                "product_identifier_no": values[2],
                "korean_product_name": normalize_space(values[3]),
                "unit": values[4],
                "price_amount": normalize_price(values[5]),
                "delivery_condition_name": normalize_space(values[6]),
                "supply_region_name": normalize_space(values[7]),
                "pps_department_name": normalize_space(values[8]),
                "posted_date": normalize_date(values[9]),
                "is_deleted": 1 if delete_yn == "Y" else 0,
                "delete_yn": delete_yn,
                "vat_inclusion_name": normalize_space(values[11]),
                "portal_price_business_type": normalize_space(values[12]),
            }
            record["source_row_key"] = row_key(record)
            record["current_group_key"] = current_key(record)
            record["row_hash"] = sha256_text(canonical_json({k: v for k, v in record.items() if k not in {"excel_row_no", "row_hash"}}))
            records.append(record)
        return sheet_name, output_date, records
    finally:
        wb.close()


def start_import(conn: sqlite3.Connection, *, path: Path, file_hash: str, sheet_name: str, output_date: str, force: bool) -> int:
    row = conn.execute(
        """
        SELECT import_id FROM facility_material_price_import_log
        WHERE source_file_hash = ? AND status = 'success'
        LIMIT 1
        """,
        (file_hash,),
    ).fetchone()
    if row and not force:
        raise RuntimeError(f"same file already imported successfully: import_id={row[0]} use --force to re-import")
    cur = conn.execute(
        """
        INSERT INTO facility_material_price_import_log (
            source_file_path, source_file_name, source_file_hash, sheet_name,
            output_date, started_at, status
        ) VALUES (?, ?, ?, ?, ?, ?, 'running')
        """,
        (str(path), path.name, file_hash, sheet_name, output_date, now_text()),
    )
    return int(cur.lastrowid)


def upsert_records(conn: sqlite3.Connection, records: list[dict[str, Any]], import_id: int) -> tuple[int, int, int]:
    existing = {
        str(row[0]): (str(row[1]), str(row[2] or ""))
        for row in conn.execute("SELECT source_row_key, row_hash, raw_json FROM facility_material_price_master")
    }
    inserted = updated = unchanged = 0
    seen_keys: set[str] = set()
    changed_at = now_text()

    for record in records:
        source_row_key = record["source_row_key"]
        row_hash = record["row_hash"]
        raw_json = canonical_json(record)
        if source_row_key in seen_keys:
            # UI-ADOSAA-003R can contain exact duplicate historical rows. They
            # do not create a separate current price state, so keep one canonical
            # row and count the duplicate as unchanged for import accounting.
            unchanged += 1
            continue
        seen_keys.add(source_row_key)
        prev = existing.get(source_row_key)
        params = (
            source_row_key,
            row_hash,
            record["product_classification_no"],
            record["product_classification_name"],
            record["product_identifier_no"],
            record["korean_product_name"],
            record["unit"],
            record["price_amount"],
            record["delivery_condition_name"],
            record["supply_region_name"],
            record["pps_department_name"],
            record["posted_date"],
            record["is_deleted"],
            record["delete_yn"],
            record["vat_inclusion_name"],
            record["portal_price_business_type"],
            record["current_group_key"],
            raw_json,
            import_id,
            import_id,
            changed_at,
            changed_at,
            0,
        )
        if prev is None:
            conn.execute(
                """
                INSERT INTO facility_material_price_master (
                    source_row_key, row_hash, product_classification_no,
                    product_classification_name, product_identifier_no,
                    korean_product_name, unit, price_amount, delivery_condition_name,
                    supply_region_name, pps_department_name, posted_date, is_deleted,
                    delete_yn, vat_inclusion_name, portal_price_business_type,
                    current_group_key, raw_json, first_import_id, last_import_id,
                    first_seen_at, last_seen_at, missing_in_latest_import
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                params,
            )
            conn.execute(
                """
                INSERT INTO facility_material_price_change_history (
                    import_id, source_row_key, change_type, previous_row_hash,
                    new_row_hash, previous_raw_json, new_raw_json, changed_at
                ) VALUES (?, ?, 'insert', NULL, ?, NULL, ?, ?)
                """,
                (import_id, source_row_key, row_hash, raw_json, changed_at),
            )
            inserted += 1
        elif prev[0] != row_hash:
            conn.execute(
                """
                UPDATE facility_material_price_master SET
                    row_hash = ?,
                    product_classification_no = ?,
                    product_classification_name = ?,
                    product_identifier_no = ?,
                    korean_product_name = ?,
                    unit = ?,
                    price_amount = ?,
                    delivery_condition_name = ?,
                    supply_region_name = ?,
                    pps_department_name = ?,
                    posted_date = ?,
                    is_deleted = ?,
                    delete_yn = ?,
                    vat_inclusion_name = ?,
                    portal_price_business_type = ?,
                    current_group_key = ?,
                    raw_json = ?,
                    last_import_id = ?,
                    last_seen_at = ?,
                    missing_in_latest_import = 0,
                    updated_at = CURRENT_TIMESTAMP
                WHERE source_row_key = ?
                """,
                (
                    row_hash,
                    record["product_classification_no"],
                    record["product_classification_name"],
                    record["product_identifier_no"],
                    record["korean_product_name"],
                    record["unit"],
                    record["price_amount"],
                    record["delivery_condition_name"],
                    record["supply_region_name"],
                    record["pps_department_name"],
                    record["posted_date"],
                    record["is_deleted"],
                    record["delete_yn"],
                    record["vat_inclusion_name"],
                    record["portal_price_business_type"],
                    record["current_group_key"],
                    raw_json,
                    import_id,
                    changed_at,
                    source_row_key,
                ),
            )
            conn.execute(
                """
                INSERT INTO facility_material_price_change_history (
                    import_id, source_row_key, change_type, previous_row_hash,
                    new_row_hash, previous_raw_json, new_raw_json, changed_at
                ) VALUES (?, ?, 'update', ?, ?, ?, ?, ?)
                """,
                (import_id, source_row_key, prev[0], row_hash, prev[1], raw_json, changed_at),
            )
            updated += 1
        else:
            conn.execute(
                """
                UPDATE facility_material_price_master
                SET last_import_id = ?, last_seen_at = ?, missing_in_latest_import = 0
                WHERE source_row_key = ?
                """,
                (import_id, changed_at, source_row_key),
            )
            unchanged += 1

    conn.execute(
        """
        UPDATE facility_material_price_master
        SET missing_in_latest_import = 1
        WHERE last_import_id <> ?
        """,
        (import_id,),
    )
    missing_count = int(conn.execute("SELECT changes()").fetchone()[0] or 0)
    if missing_count:
        conn.execute(
            """
            INSERT INTO facility_material_price_change_history (
                import_id, source_row_key, change_type, previous_row_hash,
                new_row_hash, previous_raw_json, new_raw_json, changed_at
            )
            SELECT ?, source_row_key, 'missing_in_latest_import', row_hash, NULL, raw_json, NULL, ?
            FROM facility_material_price_master
            WHERE missing_in_latest_import = 1 AND last_import_id <> ?
            """,
            (import_id, changed_at, import_id),
        )
    return inserted, updated, unchanged


def refresh_current(conn: sqlite3.Connection, generated_at: str) -> int:
    conn.execute("DELETE FROM facility_material_price_current")
    conn.execute(
        """
        INSERT OR REPLACE INTO facility_material_price_current (
            current_group_key, product_classification_no, product_classification_name,
            product_identifier_no, korean_product_name, unit, latest_price_amount,
            delivery_condition_name, supply_region_name, pps_department_name,
            latest_posted_date, vat_inclusion_name, portal_price_business_type,
            source_row_key, source_row_hash, generated_at
        )
        SELECT
            m.current_group_key,
            m.product_classification_no,
            m.product_classification_name,
            m.product_identifier_no,
            m.korean_product_name,
            m.unit,
            m.price_amount,
            m.delivery_condition_name,
            m.supply_region_name,
            m.pps_department_name,
            m.posted_date,
            m.vat_inclusion_name,
            m.portal_price_business_type,
            m.source_row_key,
            m.row_hash,
            ?
        FROM facility_material_price_master m
        JOIN (
            SELECT current_group_key, MAX(posted_date || '|' || source_row_key) AS max_selector
            FROM facility_material_price_master
            WHERE missing_in_latest_import = 0
            GROUP BY current_group_key
        ) latest
          ON latest.current_group_key = m.current_group_key
         AND latest.max_selector = (m.posted_date || '|' || m.source_row_key)
        WHERE m.is_deleted = 0
          AND m.missing_in_latest_import = 0
        """,
        (generated_at,),
    )
    return int(conn.execute("SELECT COUNT(*) FROM facility_material_price_current").fetchone()[0] or 0)


def refresh_dictionary(conn: sqlite3.Connection, generated_at: str) -> int:
    conn.execute("DELETE FROM facility_material_item_dictionary")
    conn.execute(
        """
        INSERT OR REPLACE INTO facility_material_item_dictionary (
            dictionary_key, product_classification_no, product_identifier_no,
            product_classification_name, korean_product_name, portal_price_business_types,
            units, active_price_count, historical_row_count, latest_posted_date,
            min_active_price_amount, max_active_price_amount, search_text, generated_at
        )
        SELECT
            lower(hex(randomblob(16))),
            product_classification_no,
            product_identifier_no,
            MAX(product_classification_name),
            MAX(korean_product_name),
            GROUP_CONCAT(DISTINCT portal_price_business_type),
            GROUP_CONCAT(DISTINCT unit),
            SUM(CASE WHEN is_deleted = 0 AND missing_in_latest_import = 0 THEN 1 ELSE 0 END),
            COUNT(*),
            MAX(posted_date),
            MIN(CASE WHEN is_deleted = 0 AND missing_in_latest_import = 0 THEN price_amount ELSE NULL END),
            MAX(CASE WHEN is_deleted = 0 AND missing_in_latest_import = 0 THEN price_amount ELSE NULL END),
            lower(
                IFNULL(product_classification_no, '') || ' ' ||
                IFNULL(product_identifier_no, '') || ' ' ||
                IFNULL(MAX(product_classification_name), '') || ' ' ||
                IFNULL(MAX(korean_product_name), '') || ' ' ||
                IFNULL(GROUP_CONCAT(DISTINCT portal_price_business_type), '') || ' ' ||
                IFNULL(GROUP_CONCAT(DISTINCT unit), '')
            ),
            ?
        FROM facility_material_price_master
        WHERE IFNULL(product_classification_no, '') <> ''
           OR IFNULL(product_identifier_no, '') <> ''
           OR IFNULL(product_classification_name, '') <> ''
           OR IFNULL(korean_product_name, '') <> ''
        GROUP BY product_classification_no, product_identifier_no, product_classification_name, korean_product_name
        """,
        (generated_at,),
    )
    return int(conn.execute("SELECT COUNT(*) FROM facility_material_item_dictionary").fetchone()[0] or 0)


def finish_import(
    conn: sqlite3.Connection,
    *,
    import_id: int,
    source_name: str,
    source_file: Path,
    file_hash: str,
    status: str,
    total: int,
    inserted: int,
    updated: int,
    unchanged: int,
    current_count: int,
    dictionary_count: int,
    error_message: str = "",
) -> None:
    missing_count = int(
        conn.execute("SELECT COUNT(*) FROM facility_material_price_master WHERE missing_in_latest_import = 1").fetchone()[0] or 0
    )
    finished_at = now_text()
    conn.execute(
        """
        UPDATE facility_material_price_import_log
        SET finished_at = ?, status = ?, total_row_count = ?, inserted_count = ?,
            updated_count = ?, unchanged_count = ?, missing_count = ?,
            current_active_count = ?, dictionary_count = ?, error_message = ?
        WHERE import_id = ?
        """,
        (
            finished_at,
            status,
            total,
            inserted,
            updated,
            unchanged,
            missing_count,
            current_count,
            dictionary_count,
            error_message[:1000],
            import_id,
        ),
    )
    conn.execute(
        """
        INSERT INTO etl_job_log (
            job_name, source_name, started_at, finished_at, status,
            input_row_count, inserted_count, updated_count, skipped_count,
            error_count, error_message
        )
        SELECT 'import_facility_material_price_file', ?, started_at, ?, ?, ?, ?, ?, ?, ?, ?
        FROM facility_material_price_import_log
        WHERE import_id = ?
        """,
        (
            source_name,
            finished_at,
            status,
            total,
            inserted,
            updated,
            unchanged,
            1 if status != "success" else 0,
            error_message[:1000],
            import_id,
        ),
    )
    conn.execute(
        """
        INSERT INTO source_manifest (
            source_name, source_type, source_url_or_file, source_refreshed_at,
            row_count, checksum, status, error_message
        ) VALUES (?, 'manual_xlsx_snapshot_incremental', ?, ?, ?, ?, ?, ?)
        ON CONFLICT(source_name) DO UPDATE SET
            source_type=excluded.source_type,
            source_url_or_file=excluded.source_url_or_file,
            source_refreshed_at=excluded.source_refreshed_at,
            row_count=excluded.row_count,
            checksum=excluded.checksum,
            status=excluded.status,
            error_message=excluded.error_message,
            updated_at=CURRENT_TIMESTAMP
        """,
        (source_name, str(source_file), finished_at, current_count, file_hash, status, error_message[:1000]),
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Import UI-ADOSAA-003R facility common material price XLSX.")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--db", default=DB_FILE)
    parser.add_argument("--force", action="store_true")
    args = parser.parse_args()

    source_file = args.xlsx.resolve()
    if not source_file.exists():
        raise SystemExit(f"file not found: {source_file}")

    file_hash = sha256_file(source_file)
    sheet_name, output_date, records = iter_records(source_file)
    source_name = "g2b_ui_adosaa_003r_facility_material_price_file"

    conn = sqlite3.connect(args.db)
    try:
        ensure_schema(conn)
        import_id = start_import(
            conn,
            path=source_file,
            file_hash=file_hash,
            sheet_name=sheet_name,
            output_date=output_date,
            force=args.force,
        )
        try:
            with conn:
                inserted, updated, unchanged = upsert_records(conn, records, import_id)
                generated_at = now_text()
                current_count = refresh_current(conn, generated_at)
                dictionary_count = refresh_dictionary(conn, generated_at)
                finish_import(
                    conn,
                    import_id=import_id,
                    source_name=source_name,
                    source_file=source_file,
                    file_hash=file_hash,
                    status="success",
                    total=len(records),
                    inserted=inserted,
                    updated=updated,
                    unchanged=unchanged,
                    current_count=current_count,
                    dictionary_count=dictionary_count,
                )
        except Exception as exc:
            with conn:
                finish_import(
                    conn,
                    import_id=import_id,
                    source_name=source_name,
                    source_file=source_file,
                    file_hash=file_hash,
                    status="failed",
                    total=len(records),
                    inserted=0,
                    updated=0,
                    unchanged=0,
                    current_count=0,
                    dictionary_count=0,
                    error_message=f"{type(exc).__name__}: {exc}",
                )
            raise

        stats = {
            "db": args.db,
            "file": str(source_file),
            "sheet": sheet_name,
            "output_date": output_date,
            "import_id": import_id,
            "row_count": len(records),
            "inserted": inserted,
            "updated": updated,
            "unchanged": unchanged,
            "current_active_count": current_count,
            "dictionary_count": dictionary_count,
        }
        print(json.dumps(stats, ensure_ascii=False, indent=2))
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
