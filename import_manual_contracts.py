"""
import_manual_contracts.py — 수기 계약 데이터 DB 반영
======================================================
부산시설공단 수의(수기) 계약 53건을 기존 DB에 반영.
- 기본: dry-run (건수/금액 확인만)
- --execute: 실제 DB 반영
- --rollback: 수기 반영 데이터 삭제
"""
import sqlite3, sys, os, re, argparse
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding='utf-8')

# ============ 설정 ============
EXCEL_NAME = '00000.2026년 계약대장현황(시설공단)-노인일자리(수기).xlsx'
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 엑셀 파일 탐색 순서: 1) 스크립트와 같은 폴더 2) 상위 폴더 3) Windows 하드코딩
EXCEL = None
for candidate in [
    os.path.join(SCRIPT_DIR, EXCEL_NAME),
    os.path.join(os.path.dirname(SCRIPT_DIR), EXCEL_NAME),
    r'c:\Users\doors\OneDrive\바탕 화면\사무실 작업' + '\\' + EXCEL_NAME,
]:
    if os.path.exists(candidate):
        EXCEL = candidate
        break
if EXCEL is None:
    print(f'❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_NAME}')
    print(f'   스크립트와 같은 폴더에 넣어주세요: {SCRIPT_DIR}')
    sys.exit(1)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'procurement_contracts.db')
AG_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'busan_agencies_master.db')
AGENCY_CD = 'B552587'  # 부산시설공단
AGENCY_NM = '부산시설공단'
MANUAL_PREFIX = 'MANUAL_BSFC'
TARGET_METHOD = '수의(수기)'

# ============ 엑셀 컬럼 인덱스 ============
COL_NAME = 3        # 계약명
COL_TYPE = 7        # 계약종류 (공사-총액, 용역-단가, 물품-총액 등)
COL_METHOD = 8      # 계약방법
COL_AMT = 9         # 최초계약액
COL_DATE = 10       # 계약체결일
COL_START = 11      # 계약시작일
COL_END = 12        # 계약종료일
COL_TOT_AMT = 22    # 총계약액
COL_CORP = 23       # 주계약업체
COL_BIZNO = 24      # 사업자번호
COL_REPR = 25       # 대표자
COL_ADDR = 26       # 주소
COL_PHONE = 27      # 전화번호


def format_date(val):
    """날짜를 YYYY-MM-DD 형식으로 변환"""
    if val is None:
        return ''
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    return str(val)[:10]


def build_corp_list(corp_name, bizno, addr, phone, repr_name):
    """조달청 corpList 형식으로 변환 (간이)
    형식: [1^대표^업체명^사업자번호^대표자^주소^전화^100]
    """
    bizno = str(bizno or '').strip()
    corp_name = (corp_name or '').strip()
    addr = (addr or '').strip()
    phone = (phone or '').strip()
    repr_name = (repr_name or '').strip()
    return f"[1^대표^{corp_name}^{bizno}^{repr_name}^{addr}^{phone}^100]"


def get_table_for_type(contract_type):
    """계약종류 → DB 테이블 매핑"""
    if '공사' in contract_type:
        return 'cnstwk_cntrct'
    elif '용역' in contract_type:
        return 'servc_cntrct'
    elif '물품' in contract_type:
        return 'thng_cntrct'
    return None


def get_method_name(method):
    """엑셀 계약방법 → DB cntrctCnclsMthdNm 매핑"""
    mapping = {
        '수의(수기)': '수의계약',
        '수의(전자)': '수의계약',
        '소액수의': '수의계약',
        '경쟁입찰': '제한경쟁',
        '중앙조달': '수의계약',
    }
    return mapping.get(method, '수의계약')


def load_existing_keys(conn, agency_codes):
    """기존 DB에서 시설공단 계약의 (계약명, 금액) 키 로드"""
    ph = ','.join([f"'{c}'" for c in agency_codes])
    keys = set()
    for tbl, name_col in [('cnstwk_cntrct', 'cnstwkNm'), 
                           ('servc_cntrct', 'cntrctNm'), 
                           ('thng_cntrct', 'cntrctNm')]:
        rows = conn.execute(f"""
            SELECT {name_col}, thtmCntrctAmt, totCntrctAmt 
            FROM {tbl} WHERE dminsttCd IN ({ph})
        """).fetchall()
        for r in rows:
            name = (r[0] or '').strip()
            amt = int(r[1] or r[2] or 0)
            keys.add((name, amt))
    return keys


def main():
    parser = argparse.ArgumentParser(description='수기 계약 DB 반영')
    parser.add_argument('--execute', action='store_true', help='실제 DB 반영 (기본: dry-run)')
    parser.add_argument('--rollback', action='store_true', help='수기 반영 데이터 삭제')
    args = parser.parse_args()

    conn = sqlite3.connect(DB_PATH)
    conn_ag = sqlite3.connect(AG_DB_PATH)

    # 시설공단 기관코드 로드
    agency_codes = [r[0] for r in conn_ag.execute(
        "SELECT dminsttCd FROM agency_master WHERE compare_unit='부산시설공단'"
    ).fetchall()]

    # ============ ROLLBACK 모드 ============
    if args.rollback:
        print('=== 수기 반영 데이터 삭제 (ROLLBACK) ===')
        for tbl in ['cnstwk_cntrct', 'servc_cntrct', 'thng_cntrct']:
            cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl} WHERE bsnsDivNm='수기입력'").fetchone()[0]
            if cnt > 0:
                print(f'  {tbl}: {cnt}건 삭제')
                conn.execute(f"DELETE FROM {tbl} WHERE bsnsDivNm='수기입력'")
        conn.commit()
        print('✅ 롤백 완료')
        return

    # ============ 엑셀 로드 ============
    print(f'엑셀: {os.path.basename(EXCEL)}')
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    # 기존 키 로드
    existing_keys = load_existing_keys(conn, agency_codes)
    print(f'기존 DB 시설공단 키: {len(existing_keys)}건')

    # 기존 수기 입력 확인
    existing_manual = 0
    for tbl in ['cnstwk_cntrct', 'servc_cntrct', 'thng_cntrct']:
        existing_manual += conn.execute(
            f"SELECT COUNT(*) FROM {tbl} WHERE bsnsDivNm='수기입력'"
        ).fetchone()[0]
    if existing_manual > 0:
        print(f'⚠️ 기존 수기 입력 {existing_manual}건 존재. --rollback 후 재실행하세요.')
        return

    # ============ 엑셀 → 레코드 변환 ============
    records = {'cnstwk_cntrct': [], 'servc_cntrct': [], 'thng_cntrct': []}
    skipped_dup = 0
    skipped_method = 0
    seq = 0

    for r in range(2, ws.max_row + 1):
        method = ws.cell(r, COL_METHOD).value or ''
        if method != TARGET_METHOD:
            skipped_method += 1
            continue

        name = (ws.cell(r, COL_NAME).value or '').strip()
        amt = int(ws.cell(r, COL_AMT).value or 0)
        
        # 중복 체크
        if (name, amt) in existing_keys:
            skipped_dup += 1
            continue

        contract_type = ws.cell(r, COL_TYPE).value or ''
        table = get_table_for_type(contract_type)
        if not table:
            print(f'  ⚠️ Row {r}: 알 수 없는 계약종류 "{contract_type}" → 스킵')
            continue

        seq += 1
        tot_amt = int(ws.cell(r, COL_TOT_AMT).value or amt)
        date_str = format_date(ws.cell(r, COL_DATE).value)
        start_str = format_date(ws.cell(r, COL_START).value)
        end_str = format_date(ws.cell(r, COL_END).value)
        
        corp_list = build_corp_list(
            ws.cell(r, COL_CORP).value,
            ws.cell(r, COL_BIZNO).value,
            ws.cell(r, COL_ADDR).value,
            ws.cell(r, COL_PHONE).value,
            ws.cell(r, COL_REPR).value,
        )

        period = f"{start_str} ~ {end_str}" if start_str and end_str else ''
        unty_no = f"{MANUAL_PREFIX}_{seq:04d}"
        
        # 계약명 컬럼명이 테이블별로 다름
        name_col = 'cnstwkNm' if table == 'cnstwk_cntrct' else 'cntrctNm'
        
        record = {
            'untyCntrctNo': unty_no,
            'bsnsDivNm': '수기입력',
            'dcsnCntrctNo': unty_no + '00',
            name_col: name,
            'cntrctCnclsMthdNm': get_method_name(method),
            'totCntrctAmt': str(tot_amt),
            'thtmCntrctAmt': str(amt),
            'cntrctCnclsDate': date_str,
            'cntrctDate': date_str,
            'cntrctPrd': period,
            'corpList': corp_list,
            'dminsttCd': AGENCY_CD,
            'dminsttNm_req': AGENCY_NM,
            'cntrctInsttCd': AGENCY_CD,
            'cntrctInsttNm': AGENCY_NM,
            'rgstDt': datetime.now().strftime('%Y-%m-%d %H:%M'),
        }
        
        records[table].append(record)

    # ============ 결과 출력 ============
    total_count = sum(len(v) for v in records.values())
    total_amt = sum(int(r.get('thtmCntrctAmt', 0)) for recs in records.values() for r in recs)
    
    print(f'\n{"="*60}')
    print(f'  {"DRY-RUN" if not args.execute else "실행"} 결과')
    print(f'{"="*60}')
    print(f'  엑셀 전체: {ws.max_row - 1}건')
    print(f'  계약방법 필터({TARGET_METHOD} 외): {skipped_method}건 스킵')
    print(f'  중복 (이미 DB 존재): {skipped_dup}건 스킵')
    print(f'  반영 대상: {total_count}건 ({total_amt/1e8:.1f}억)')
    print()
    for tbl, recs in records.items():
        if recs:
            amt = sum(int(r.get('thtmCntrctAmt', 0)) for r in recs)
            print(f'  {tbl}: {len(recs)}건 ({amt/1e8:.1f}억)')
    
    # ============ 실행 ============
    if args.execute:
        print(f'\n  💾 DB 반영 중...')
        for tbl, recs in records.items():
            if not recs:
                continue
            cols = list(recs[0].keys())
            placeholders = ','.join(['?'] * len(cols))
            col_names = ','.join(cols)
            for rec in recs:
                vals = [rec.get(c) for c in cols]
                conn.execute(f"INSERT INTO {tbl} ({col_names}) VALUES ({placeholders})", vals)
            print(f'    {tbl}: {len(recs)}건 INSERT 완료')
        conn.commit()
        
        # 검증
        print(f'\n  ✅ 반영 완료! 검증:')
        for tbl in ['cnstwk_cntrct', 'servc_cntrct', 'thng_cntrct']:
            cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl} WHERE bsnsDivNm='수기입력'").fetchone()[0]
            amt = conn.execute(f"SELECT COALESCE(SUM(CAST(thtmCntrctAmt AS INTEGER)),0) FROM {tbl} WHERE bsnsDivNm='수기입력'").fetchone()[0]
            if cnt > 0:
                print(f'    {tbl}: {cnt}건 ({amt/1e8:.1f}억)')
        print(f'\n  다음 단계: python build_api_cache.py && python test_integrity.py')
    else:
        print(f'\n  ℹ️ dry-run 모드입니다. 실제 반영: python import_manual_contracts.py --execute')

    conn.close()


if __name__ == '__main__':
    main()
