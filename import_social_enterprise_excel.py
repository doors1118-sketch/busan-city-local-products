#!/usr/bin/env python3
"""
Import MOEL social enterprise certification XLSX files into chatbot_company.db.

Default mode is dry-run. Use --apply to mutate the database.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import shutil
import sqlite3
import sys
import uuid
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


DEFAULT_SOURCE_NAME = "moel_social_enterprise_excel"
POLICY_SUBTYPE = "social_enterprise"
POLICY_TYPE = "policy_company"
ISSUER = "고용노동부"
BUSAN_REGION_LABELS = {"부산", "부산광역시"}


class ImportErrorWithContext(RuntimeError):
    pass


def now_text() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def normalize_business_no(value: Any) -> Tuple[str, str]:
    raw = clean_text(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) != 10:
        return raw, ""
    return raw, digits


def parse_date_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""

    digits = re.sub(r"\D", "", text)
    if len(digits) == 8:
        y, m, d = int(digits[:4]), int(digits[4:6]), int(digits[6:])
        try:
            return dt.date(y, m, d).isoformat()
        except ValueError:
            return text
    return text


def infer_source_refreshed_at(source_file: Path) -> str:
    # The MOEL file name often starts with YYMMDD, e.g. 260612.
    match = re.search(r"(?<!\d)(\d{6})(?!\d)", source_file.name)
    if match:
        yymmdd = match.group(1)
        year = 2000 + int(yymmdd[:2])
        month = int(yymmdd[2:4])
        day = int(yymmdd[4:6])
        try:
            return dt.datetime(year, month, day, 0, 0, 0).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    modified = dt.datetime.fromtimestamp(source_file.stat().st_mtime)
    return modified.strftime("%Y-%m-%d %H:%M:%S")


REQUIRED_HEADERS = {
    "지역": "region",
    "인증번호": "certification_no",
    "기관명": "company_name",
    "사업내용": "business_description",
    "사회적목적실현유형": "purpose_type",
    "사회서비스분야": "service_field",
    "사업자등록번호": "business_no",
    "대표자": "representative_name",
    "소재지": "address",
    "홈페이지": "homepage",
    "인증일자": "certification_date",
    "인증회차": "certification_round",
}


def find_header_row(rows: Iterable[Tuple[Any, ...]]) -> Tuple[int, Dict[int, str]]:
    for idx, row in enumerate(rows, start=1):
        headers: Dict[int, str] = {}
        normalized_values = [normalize_header(v) for v in row]
        if "사업자등록번호" not in normalized_values or "기관명" not in normalized_values:
            continue
        for col_idx, raw in enumerate(row, start=1):
            normalized = normalize_header(raw)
            if normalized in REQUIRED_HEADERS:
                headers[col_idx] = REQUIRED_HEADERS[normalized]
        missing = set(REQUIRED_HEADERS.values()) - set(headers.values())
        if missing:
            raise ImportErrorWithContext(f"Header row found at {idx}, but missing columns: {sorted(missing)}")
        return idx, headers
    raise ImportErrorWithContext("Could not find a header row containing 기관명 and 사업자등록번호.")


def read_social_enterprise_xlsx(source_file: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    wb = load_workbook(source_file, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    header_row_idx, col_map = find_header_row(all_rows)

    records: List[Dict[str, Any]] = []
    invalid_rows: List[Dict[str, Any]] = []
    seen_biznos: Dict[str, int] = {}
    duplicate_biznos: List[str] = []

    for row_no, row in enumerate(all_rows[header_row_idx:], start=header_row_idx + 1):
        raw_record: Dict[str, Any] = {}
        for col_idx, field in col_map.items():
            raw_record[field] = clean_text(row[col_idx - 1]) if col_idx - 1 < len(row) else ""

        if not raw_record.get("company_name") and not raw_record.get("business_no"):
            continue

        raw_bizno, canonical_bizno = normalize_business_no(raw_record.get("business_no"))
        if not canonical_bizno:
            invalid_rows.append({
                "source_row_no": row_no,
                "company_name": raw_record.get("company_name", ""),
                "raw_business_no": raw_bizno,
                "reason": "invalid_business_no",
            })
            continue

        if canonical_bizno in seen_biznos:
            duplicate_biznos.append(canonical_bizno)
        seen_biznos[canonical_bizno] = row_no

        raw_record.update({
            "source_row_no": row_no,
            "raw_business_no": raw_bizno,
            "canonical_business_no": canonical_bizno,
            "certification_date": parse_date_text(raw_record.get("certification_date")),
        })
        records.append(raw_record)

    meta = {
        "sheet_name": ws.title,
        "header_row": header_row_idx,
        "record_count": len(records),
        "invalid_rows": invalid_rows[:20],
        "invalid_count": len(invalid_rows),
        "duplicate_business_numbers": sorted(set(duplicate_biznos))[:20],
        "duplicate_count": len(set(duplicate_biznos)),
    }
    return records, meta


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS social_enterprise_master (
            source_name TEXT NOT NULL,
            canonical_business_no TEXT NOT NULL,
            region TEXT,
            certification_no TEXT,
            company_name TEXT,
            business_description TEXT,
            purpose_type TEXT,
            service_field TEXT,
            representative_name TEXT,
            address TEXT,
            homepage TEXT,
            certification_date TEXT,
            certification_round TEXT,
            source_file_name TEXT,
            source_file_sha256 TEXT,
            source_refreshed_at TEXT,
            import_batch_id TEXT,
            matched_company_internal_id INTEGER,
            matched_company_name TEXT,
            matched_is_busan_company INTEGER,
            match_status TEXT NOT NULL DEFAULT 'not_matched',
            record_status TEXT NOT NULL DEFAULT 'active',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (source_name, canonical_business_no)
        );

        CREATE INDEX IF NOT EXISTS idx_social_enterprise_master_bizno
            ON social_enterprise_master(canonical_business_no);
        CREATE INDEX IF NOT EXISTS idx_social_enterprise_master_match
            ON social_enterprise_master(match_status, matched_is_busan_company);

        CREATE TABLE IF NOT EXISTS social_enterprise_import_log (
            import_batch_id TEXT PRIMARY KEY,
            source_name TEXT NOT NULL,
            source_file_name TEXT,
            source_file_sha256 TEXT,
            source_refreshed_at TEXT,
            uploaded_by TEXT,
            started_at TEXT,
            finished_at TEXT,
            status TEXT,
            total_count INTEGER DEFAULT 0,
            valid_business_no_count INTEGER DEFAULT 0,
            invalid_count INTEGER DEFAULT 0,
            duplicate_count INTEGER DEFAULT 0,
            source_busan_count INTEGER DEFAULT 0,
            matched_count INTEGER DEFAULT 0,
            matched_busan_count INTEGER DEFAULT 0,
            unmatched_count INTEGER DEFAULT 0,
            policy_upsert_count INTEGER DEFAULT 0,
            expired_previous_count INTEGER DEFAULT 0,
            error_message TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        """
    )


def load_company_matches(conn: sqlite3.Connection, biznos: List[str]) -> Dict[str, Dict[str, Any]]:
    if not biznos:
        return {}
    matches: Dict[str, Dict[str, Any]] = {}
    chunk_size = 800
    for start in range(0, len(biznos), chunk_size):
        chunk = biznos[start:start + chunk_size]
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"""
            SELECT
                i.canonical_business_no,
                i.company_internal_id,
                m.company_name,
                COALESCE(m.is_busan_company, 0) AS is_busan_company
            FROM company_identity i
            LEFT JOIN company_master m ON m.company_internal_id = i.company_internal_id
            WHERE i.canonical_business_no IN ({placeholders})
            """,
            chunk,
        ).fetchall()
        for row in rows:
            matches[row["canonical_business_no"]] = {
                "company_internal_id": row["company_internal_id"],
                "company_name": row["company_name"],
                "is_busan_company": int(row["is_busan_company"] or 0),
            }
    return matches


def summarize_records(records: List[Dict[str, Any]], matches: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    source_busan_count = sum(1 for r in records if clean_text(r.get("region")) in BUSAN_REGION_LABELS)
    matched_count = 0
    matched_busan_count = 0
    source_busan_matched = 0
    source_busan_unmatched = 0
    for r in records:
        match = matches.get(r["canonical_business_no"])
        if match:
            matched_count += 1
            if int(match.get("is_busan_company") or 0) == 1:
                matched_busan_count += 1
        if clean_text(r.get("region")) in BUSAN_REGION_LABELS:
            if match:
                source_busan_matched += 1
            else:
                source_busan_unmatched += 1
    return {
        "total_count": len(records),
        "valid_business_no_count": len(records),
        "source_busan_count": source_busan_count,
        "matched_count": matched_count,
        "matched_busan_count": matched_busan_count,
        "unmatched_count": len(records) - matched_count,
        "source_busan_matched": source_busan_matched,
        "source_busan_unmatched": source_busan_unmatched,
    }


def backup_sqlite_database(db_path: Path, backup_dir: Path) -> Path:
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{db_path.stem}.before_social_enterprise_{stamp}{db_path.suffix}"
    src = sqlite3.connect(str(db_path))
    try:
        dst = sqlite3.connect(str(backup_path))
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()
    return backup_path


def certification_hash(record: Dict[str, Any]) -> str:
    base = "|".join([
        record.get("canonical_business_no", ""),
        record.get("certification_no", ""),
        record.get("certification_date", ""),
    ])
    return sha256_text(base)


def upsert_source_manifest(
    conn: sqlite3.Connection,
    source_name: str,
    source_file: Path,
    source_refreshed_at: str,
    row_count: int,
    checksum: str,
    status: str,
    error_message: str = "",
) -> None:
    conn.execute(
        """
        INSERT INTO source_manifest (
            source_name, source_type, source_url_or_file, source_refreshed_at,
            row_count, checksum, status, error_message, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(source_name) DO UPDATE SET
            source_type=excluded.source_type,
            source_url_or_file=excluded.source_url_or_file,
            source_refreshed_at=excluded.source_refreshed_at,
            row_count=excluded.row_count,
            checksum=excluded.checksum,
            status=excluded.status,
            error_message=excluded.error_message,
            updated_at=CURRENT_TIMESTAMP
        """,
        (
            source_name,
            "manual_xlsx",
            str(source_file),
            source_refreshed_at,
            row_count,
            checksum,
            status,
            error_message,
        ),
    )


def apply_import(
    conn: sqlite3.Connection,
    records: List[Dict[str, Any]],
    matches: Dict[str, Dict[str, Any]],
    *,
    source_name: str,
    source_file: Path,
    source_file_sha: str,
    source_refreshed_at: str,
    import_batch_id: str,
    uploaded_by: str,
    expire_previous_source: bool,
) -> Dict[str, int]:
    now = now_text()
    policy_upsert_count = 0
    expired_previous_count = 0
    current_policy_keys = set()

    for record in records:
        bizno = record["canonical_business_no"]
        match = matches.get(bizno)
        match_status = "matched" if match else "not_matched"
        matched_company_internal_id = match["company_internal_id"] if match else None
        matched_company_name = match["company_name"] if match else None
        matched_is_busan_company = int(match["is_busan_company"]) if match else 0

        conn.execute(
            """
            INSERT INTO social_enterprise_master (
                source_name, canonical_business_no, region, certification_no,
                company_name, business_description, purpose_type, service_field,
                representative_name, address, homepage, certification_date,
                certification_round, source_file_name, source_file_sha256,
                source_refreshed_at, import_batch_id, matched_company_internal_id,
                matched_company_name, matched_is_busan_company, match_status,
                record_status, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?)
            ON CONFLICT(source_name, canonical_business_no) DO UPDATE SET
                region=excluded.region,
                certification_no=excluded.certification_no,
                company_name=excluded.company_name,
                business_description=excluded.business_description,
                purpose_type=excluded.purpose_type,
                service_field=excluded.service_field,
                representative_name=excluded.representative_name,
                address=excluded.address,
                homepage=excluded.homepage,
                certification_date=excluded.certification_date,
                certification_round=excluded.certification_round,
                source_file_name=excluded.source_file_name,
                source_file_sha256=excluded.source_file_sha256,
                source_refreshed_at=excluded.source_refreshed_at,
                import_batch_id=excluded.import_batch_id,
                matched_company_internal_id=excluded.matched_company_internal_id,
                matched_company_name=excluded.matched_company_name,
                matched_is_busan_company=excluded.matched_is_busan_company,
                match_status=excluded.match_status,
                record_status='active',
                updated_at=excluded.updated_at
            """,
            (
                source_name,
                bizno,
                record.get("region", ""),
                record.get("certification_no", ""),
                record.get("company_name", ""),
                record.get("business_description", ""),
                record.get("purpose_type", ""),
                record.get("service_field", ""),
                record.get("representative_name", ""),
                record.get("address", ""),
                record.get("homepage", ""),
                record.get("certification_date", ""),
                record.get("certification_round", ""),
                source_file.name,
                source_file_sha,
                source_refreshed_at,
                import_batch_id,
                matched_company_internal_id,
                matched_company_name,
                matched_is_busan_company,
                match_status,
                now,
            ),
        )

        if matched_company_internal_id is None:
            continue

        cert_hash = certification_hash(record)
        current_policy_keys.add((matched_company_internal_id, cert_hash))
        conn.execute(
            """
            INSERT INTO policy_company_certification (
                company_internal_id, policy_type, policy_subtype,
                certification_no_hash, certification_valid_from,
                certification_valid_to, validity_status, issuer, source_name,
                source_refreshed_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, NULL, 'valid', ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(company_internal_id, policy_subtype, source_name, certification_no_hash)
            DO UPDATE SET
                certification_valid_from=excluded.certification_valid_from,
                certification_valid_to=excluded.certification_valid_to,
                validity_status='valid',
                issuer=excluded.issuer,
                source_refreshed_at=excluded.source_refreshed_at,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                matched_company_internal_id,
                POLICY_TYPE,
                POLICY_SUBTYPE,
                cert_hash,
                record.get("certification_date") or None,
                ISSUER,
                source_name,
                source_refreshed_at,
            ),
        )
        policy_upsert_count += 1

    if expire_previous_source:
        if current_policy_keys:
            # Expire old records from this source that are absent from the new full-list import.
            rows = conn.execute(
                """
                SELECT policy_cert_id, company_internal_id, certification_no_hash
                FROM policy_company_certification
                WHERE policy_subtype = ?
                  AND source_name = ?
                  AND validity_status = 'valid'
                """,
                (POLICY_SUBTYPE, source_name),
            ).fetchall()
            expired_ids = [
                row["policy_cert_id"]
                for row in rows
                if (row["company_internal_id"], row["certification_no_hash"]) not in current_policy_keys
            ]
            if expired_ids:
                placeholders = ",".join("?" for _ in expired_ids)
                conn.execute(
                    f"""
                    UPDATE policy_company_certification
                    SET validity_status='expired',
                        certification_valid_to=COALESCE(certification_valid_to, DATE('now')),
                        updated_at=CURRENT_TIMESTAMP
                    WHERE policy_cert_id IN ({placeholders})
                    """,
                    expired_ids,
                )
                expired_previous_count = len(expired_ids)

        imported_biznos = [r["canonical_business_no"] for r in records]
        if imported_biznos:
            placeholders = ",".join("?" for _ in imported_biznos)
            conn.execute(
                f"""
                UPDATE social_enterprise_master
                SET record_status='expired',
                    updated_at=CURRENT_TIMESTAMP
                WHERE source_name = ?
                  AND canonical_business_no NOT IN ({placeholders})
                """,
                [source_name] + imported_biznos,
            )

    return {
        "policy_upsert_count": policy_upsert_count,
        "expired_previous_count": expired_previous_count,
    }


def write_import_log(
    conn: sqlite3.Connection,
    *,
    import_batch_id: str,
    source_name: str,
    source_file: Path,
    source_file_sha: str,
    source_refreshed_at: str,
    uploaded_by: str,
    started_at: str,
    finished_at: str,
    status: str,
    summary: Dict[str, Any],
    policy_upsert_count: int,
    expired_previous_count: int,
    error_message: str = "",
) -> None:
    conn.execute(
        """
        INSERT INTO social_enterprise_import_log (
            import_batch_id, source_name, source_file_name, source_file_sha256,
            source_refreshed_at, uploaded_by, started_at, finished_at, status,
            total_count, valid_business_no_count, invalid_count, duplicate_count,
            source_busan_count, matched_count, matched_busan_count, unmatched_count,
            policy_upsert_count, expired_previous_count, error_message
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            import_batch_id,
            source_name,
            source_file.name,
            source_file_sha,
            source_refreshed_at,
            uploaded_by,
            started_at,
            finished_at,
            status,
            int(summary.get("total_count", 0)),
            int(summary.get("valid_business_no_count", 0)),
            int(summary.get("invalid_count", 0)),
            int(summary.get("duplicate_count", 0)),
            int(summary.get("source_busan_count", 0)),
            int(summary.get("matched_count", 0)),
            int(summary.get("matched_busan_count", 0)),
            int(summary.get("unmatched_count", 0)),
            policy_upsert_count,
            expired_previous_count,
            error_message,
        ),
    )


def run(args: argparse.Namespace) -> Dict[str, Any]:
    source_file = Path(args.source_file).expanduser().resolve()
    db_path = Path(args.db).expanduser().resolve()
    if not source_file.exists():
        raise ImportErrorWithContext(f"Source file does not exist: {source_file}")
    if not db_path.exists():
        raise ImportErrorWithContext(f"Database file does not exist: {db_path}")

    source_refreshed_at = args.source_refreshed_at or infer_source_refreshed_at(source_file)
    source_file_sha = file_sha256(source_file)
    import_batch_id = args.import_batch_id or f"social_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    started_at = now_text()

    records, meta = read_social_enterprise_xlsx(source_file)
    if meta["invalid_count"] or meta["duplicate_count"]:
        # Do not silently import a dirty official source. The admin page can show this as a failed preview.
        raise ImportErrorWithContext(
            f"Source validation failed: invalid={meta['invalid_count']}, duplicate={meta['duplicate_count']}"
        )

    if args.apply:
        conn = sqlite3.connect(str(db_path), timeout=30.0)
    else:
        conn = sqlite3.connect(f"file:{db_path.as_posix()}?mode=ro", timeout=30.0, uri=True)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA busy_timeout=30000")
        if args.apply:
            ensure_schema(conn)
        matches = load_company_matches(conn, [r["canonical_business_no"] for r in records])
        summary = summarize_records(records, matches)
        summary.update({
            "invalid_count": meta["invalid_count"],
            "duplicate_count": meta["duplicate_count"],
        })

        backup_path = ""
        apply_counts = {"policy_upsert_count": 0, "expired_previous_count": 0}

        if args.apply:
            backup_dir = Path(args.backup_dir).expanduser().resolve() if args.backup_dir else db_path.parent / "backups"
            backup_path = str(backup_sqlite_database(db_path, backup_dir))
            try:
                conn.execute("BEGIN")
                apply_counts = apply_import(
                    conn,
                    records,
                    matches,
                    source_name=args.source_name,
                    source_file=source_file,
                    source_file_sha=source_file_sha,
                    source_refreshed_at=source_refreshed_at,
                    import_batch_id=import_batch_id,
                    uploaded_by=args.uploaded_by,
                    expire_previous_source=args.expire_previous_source,
                )
                upsert_source_manifest(
                    conn,
                    args.source_name,
                    source_file,
                    source_refreshed_at,
                    len(records),
                    source_file_sha,
                    "success",
                )
                write_import_log(
                    conn,
                    import_batch_id=import_batch_id,
                    source_name=args.source_name,
                    source_file=source_file,
                    source_file_sha=source_file_sha,
                    source_refreshed_at=source_refreshed_at,
                    uploaded_by=args.uploaded_by,
                    started_at=started_at,
                    finished_at=now_text(),
                    status="success",
                    summary=summary,
                    policy_upsert_count=apply_counts["policy_upsert_count"],
                    expired_previous_count=apply_counts["expired_previous_count"],
                )
                conn.execute(
                    """
                    INSERT INTO etl_job_log (
                        job_name, source_name, started_at, finished_at, status,
                        input_row_count, inserted_count, updated_count, skipped_count,
                        error_count, error_message
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "import_social_enterprise_excel",
                        args.source_name,
                        started_at,
                        now_text(),
                        "success",
                        len(records),
                        apply_counts["policy_upsert_count"],
                        0,
                        summary["unmatched_count"],
                        0,
                        "",
                    ),
                )
                conn.commit()
            except Exception:
                conn.rollback()
                # Keep the backup even if import fails.
                raise

        result = {
            "mode": "apply" if args.apply else "dry_run",
            "db_path": str(db_path),
            "source_file": str(source_file),
            "source_name": args.source_name,
            "source_refreshed_at": source_refreshed_at,
            "source_file_sha256": source_file_sha,
            "import_batch_id": import_batch_id,
            "backup_path": backup_path,
            "sheet": meta["sheet_name"],
            "header_row": meta["header_row"],
            **summary,
            **apply_counts,
        }
        return result
    finally:
        conn.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import MOEL social enterprise XLSX into chatbot company DB.")
    parser.add_argument("--source-file", required=True, help="MOEL social enterprise XLSX file.")
    parser.add_argument(
        "--db",
        default=os.environ.get("CHATBOT_DB") or str(Path(__file__).resolve().with_name("chatbot_company.db")),
        help="Path to chatbot_company.db. Default: CHATBOT_DB or script directory chatbot_company.db.",
    )
    parser.add_argument("--source-name", default=DEFAULT_SOURCE_NAME)
    parser.add_argument("--source-refreshed-at", default="", help="YYYY-MM-DD HH:MM:SS. Default: inferred from file name or mtime.")
    parser.add_argument("--uploaded-by", default=os.environ.get("USER") or os.environ.get("USERNAME") or "manual")
    parser.add_argument("--import-batch-id", default="")
    parser.add_argument("--backup-dir", default="")
    parser.add_argument("--apply", action="store_true", help="Mutate DB. Without this flag only validates and matches.")
    parser.add_argument(
        "--expire-previous-source",
        action="store_true",
        help="Expire previous valid rows from the same source_name that are absent from this full-list import.",
    )
    parser.add_argument("--json", action="store_true", help="Print JSON only.")
    return parser


def print_human(result: Dict[str, Any]) -> None:
    print(f"mode: {result['mode']}")
    print(f"source_name: {result['source_name']}")
    print(f"source_refreshed_at: {result['source_refreshed_at']}")
    print(f"total_count: {result['total_count']}")
    print(f"source_busan_count: {result['source_busan_count']}")
    print(f"matched_count: {result['matched_count']}")
    print(f"matched_busan_count: {result['matched_busan_count']}")
    print(f"unmatched_count: {result['unmatched_count']}")
    print(f"source_busan_matched: {result['source_busan_matched']}")
    print(f"source_busan_unmatched: {result['source_busan_unmatched']}")
    print(f"policy_upsert_count: {result['policy_upsert_count']}")
    if result.get("backup_path"):
        print(f"backup_path: {result['backup_path']}")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        result = run(args)
    except Exception as exc:
        if args.json:
            print(json.dumps({"status": "error", "error_message": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print_human(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
