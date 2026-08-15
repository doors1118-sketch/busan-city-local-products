"""
부산 조달 모니터링 REST API 서버
=================================
캐시 파일(api_cache.json)을 읽어서 즉시 응답

실행: python api_server.py
문서: http://localhost:8000/docs
"""
from fastapi import Body, FastAPI, Header, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
import argparse
import base64
import datetime as dt
import json, sys, math, os, sqlite3, subprocess
import re
import secrets
from pathlib import Path
from typing import Optional

sys.stdout.reconfigure(encoding='utf-8')

# NaN-safe JSON encoder: NaN/Inf → null (FastAPI 500 방지)
class NaNSafeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
            return None
        return super().default(obj)

class NaNSafeResponse(JSONResponse):
    def render(self, content):
        return json.dumps(content, ensure_ascii=False, cls=NaNSafeEncoder).encode('utf-8')

app = FastAPI(title="부산 조달 모니터링 API", version="1.0", default_response_class=NaNSafeResponse)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.get("/health", include_in_schema=False)
def health():
    """Liveness probe used by the guarded production deployer."""
    return {"status": "ok", "service": "busan-procurement-api"}

CACHE_FILE = 'api_cache.json'
DB_COMPANIES = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'busan_companies_master.db')
SOCIAL_PURCHASE_CACHE_JSON = Path(os.environ.get(
    "SOCIAL_PURCHASE_CACHE_JSON",
    str(Path(__file__).resolve().parent / "social_purchase_cache.json"),
))
SOCIAL_PURCHASE_CACHE_XLSX = Path(os.environ.get(
    "SOCIAL_PURCHASE_CACHE_XLSX",
    str(Path(__file__).resolve().parent / "social_purchase_cache.xlsx"),
))
SOCIAL_PURCHASE_BUILD_LOG = Path(os.environ.get(
    "SOCIAL_PURCHASE_BUILD_LOG",
    str(Path(__file__).resolve().parent / "sync_log" / "social_purchase_cache_manual.log"),
))


def _trigger_social_purchase_cache_refresh(reason="manual"):
    script = Path(__file__).resolve().parent / "build_social_purchase_validation.py"
    if not script.exists():
        return {
            "automatic": True,
            "started": False,
            "reason": f"cache build script not found: {script}",
        }
    try:
        SOCIAL_PURCHASE_BUILD_LOG.parent.mkdir(parents=True, exist_ok=True)
        log = open(SOCIAL_PURCHASE_BUILD_LOG, "ab", buffering=0)
        log.write(
            f"\n[{dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] refresh requested: {reason}\n".encode("utf-8")
        )
        proc = subprocess.Popen(
            [sys.executable, str(script)],
            cwd=str(script.parent),
            stdout=log,
            stderr=subprocess.STDOUT,
            start_new_session=True,
        )
        log.close()
        return {
            "automatic": True,
            "started": True,
            "mode": "background",
            "pid": proc.pid,
            "script": str(script),
            "log": str(SOCIAL_PURCHASE_BUILD_LOG),
        }
    except Exception as e:
        return {
            "automatic": True,
            "started": False,
            "error": str(e),
            "script": str(script),
            "log": str(SOCIAL_PURCHASE_BUILD_LOG),
        }

def _get_company_db():
    """부산 업체 마스터 DB 연결 (읽기 전용)"""
    conn = sqlite3.connect(f"file:{DB_COMPANIES}?mode=ro", uri=True)
    conn.row_factory = sqlite3.Row
    return conn

def load_cache():
    try:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] 캐시 로드 실패: {e}")
        return {"error": f"캐시 파일 로드 실패: {str(e)}"}


def _load_social_purchase_cache():
    try:
        with SOCIAL_PURCHASE_CACHE_JSON.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        raise HTTPException(
            status_code=503,
            detail=f"Social enterprise purchase cache not found: {SOCIAL_PURCHASE_CACHE_JSON}",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Social enterprise purchase cache load failed: {e}")
    summary = data.get("summary") or {}
    summary.setdefault("default_rate_basis", "shopping_included")
    summary.setdefault(
        "formula_default",
        "분모=대상기관 부산 현장/소비처 물품+용역+쇼핑몰, 분자=대상기관 부산 현장/소비처 공사+용역+물품+쇼핑몰 사회적기업 수주액",
    )
    return data


def _with_default_social_rate(row):
    item = dict(row or {})
    item["기본모수"] = item.get("모수_물품용역쇼핑몰", 0)
    item["기본수주율"] = item.get("수주율_쇼핑몰포함", 0)
    item["기본수주율_기준"] = "쇼핑몰 포함"
    return item


def _filter_social_rows(rows, agency=None, month=None):
    out = []
    agency_q = str(agency or "").strip()
    month_q = str(month or "").strip()
    for row in rows or []:
        if agency_q and agency_q not in str(row.get("기관", "")):
            continue
        if month_q and row.get("월") != month_q:
            continue
        out.append(_with_default_social_rate(row))
    return out

@app.get("/", include_in_schema=False)
def root():
    """루트 접속 시 API 문서로 이동"""
    return RedirectResponse(url="/docs")

@app.get("/api/summary", tags=["대시보드"])
def get_summary():
    """종합 수주율 (전체/분야별/그룹별/그룹별×분야별)"""
    cache = load_cache()
    return {k: v for k, v in cache.items() if not k.startswith("5_")}

@app.get("/api/ranking", tags=["대시보드"])
def get_ranking():
    """비교단위 수주율 랭킹 (전체+분야별, 그룹별 상/하위 10)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "전체": cache.get("5_기관랭킹_전체", {}),
        "분야별": cache.get("5_기관랭킹_분야별", {}),
        "소그룹": cache.get("5_기관랭킹_소그룹", {}),
    }

@app.get("/api/ranking/{sector}", tags=["대시보드"])
def get_ranking_by_sector(sector: str):
    """특정 분야 수주율 랭킹 (공사/용역/물품/쇼핑몰)"""
    cache = load_cache()
    data = cache.get("5_기관랭킹_분야별", {}).get(sector)
    if not data:
        return {"error": f"'{sector}' 분야를 찾을 수 없습니다. (공사/용역/물품/쇼핑몰)"}
    return {"generated_at": cache.get("generated_at"), "분야": sector, "랭킹": data}

@app.get("/api/leakage", tags=["유출 분석"])
def get_leakage():
    """유출 분석 전체: 쇼핑몰 유출품목 + 공사/용역/물품 유출계약"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "쇼핑몰_유출품목": cache.get("6_유출품목_쇼핑몰", []),
        "유출계약": cache.get("7_유출계약_주요", []),
    }

@app.get("/api/leakage/shopping", tags=["유출 분석"])
def get_leakage_shopping():
    """쇼핑몰 유출품목 Top 10 (비부산 업체 유출액 기준)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "유출품목": cache.get("6_유출품목_쇼핑몰", []),
    }

@app.get("/api/leakage/contracts", tags=["유출 분석"])
def get_leakage_contracts():
    """공사/용역/물품 주요 유출계약 Top 10 (유출액 기준, 필터 적용)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "유출계약": cache.get("7_유출계약_주요", []),
    }

@app.get("/api/protection", tags=["보호제도"])
def get_protection():
    """지역업체 보호제도 적용 현황 + 미적용 Top 10"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "현황": cache.get("8_보호제도_현황", {}),
        "미적용_건": cache.get("8_보호제도_미적용", []),
        "기관별_미적용": cache.get("8_보호제도_기관별", []),
    }

@app.get("/api/private-contract", tags=["수의계약"])
def get_private_contract():
    """수의계약 지역업체 수주율 (공사/물품/용역, 국가/부산시 2그룹)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "수의계약": cache.get("9_수의계약", {}),
        "유출_수의계약": cache.get("9_수의계약_유출", []),
        "유출_기관별": cache.get("9_수의계약_유출_기관별", []),
    }

@app.get("/api/local-companies", tags=["지역업체"])
def get_local_companies():
    """지역업체 현황표 (전체/분야별 업체수, 물품 대분류, 공사/용역 업종)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "현황": cache.get("10_지역업체현황", {}),
    }

@app.get("/api/economic-impact", tags=["경제효과"])
def get_economic_impact():
    """지역상품 구매에 따른 지역생산부가가치 및 지역고용기여도 (한국은행 2020 지역산업연관표 부산 계수)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "경제효과": cache.get("11_경제효과", {}),
    }


@app.get("/api/social-enterprise/purchase", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase():
    """사회적기업 구매율 대시보드용 전체 페이로드. 기본 수주율은 쇼핑몰 포함 기준."""
    data = _load_social_purchase_cache()
    return {
        "summary": data.get("summary", {}),
        "overall": _with_default_social_rate(data.get("overall", {})),
        "monthly": [_with_default_social_rate(r) for r in data.get("monthly", [])],
        "agency_rates": [_with_default_social_rate(r) for r in data.get("agency_rates", [])],
        "agency_monthly": [_with_default_social_rate(r) for r in data.get("agency_monthly", [])],
        "social_contracts": data.get("social_contracts", []),
    }


@app.get("/api/social-enterprise/purchase/summary", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase_summary():
    """사회적기업 구매율 전체 요약. 기본 수주율은 쇼핑몰 포함 기준."""
    data = _load_social_purchase_cache()
    return {
        "summary": data.get("summary", {}),
        "overall": _with_default_social_rate(data.get("overall", {})),
    }


@app.get("/api/social-enterprise/purchase/monthly", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase_monthly(
    month: Optional[str] = Query(None, description="YYYY-MM 형식. 미지정 시 전체 월 반환"),
):
    """사회적기업 구매율 월별 추이. 기본 수주율은 쇼핑몰 포함 기준."""
    data = _load_social_purchase_cache()
    return {
        "summary": data.get("summary", {}),
        "rows": _filter_social_rows(data.get("monthly", []), month=month),
    }


@app.get("/api/social-enterprise/purchase/agencies", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase_agencies(
    q: Optional[str] = Query(None, description="기관명 포함 검색"),
    sort: str = Query("social_amount", description="social_amount/rate/denominator/name"),
    limit: int = Query(100, ge=1, le=500),
):
    """대상기관별 사회적기업 구매율. 기본 수주율은 쇼핑몰 포함 기준."""
    data = _load_social_purchase_cache()
    rows = _filter_social_rows(data.get("agency_rates", []), agency=q)
    sort_map = {
        "social_amount": ("사회적기업수주액", True),
        "rate": ("수주율_쇼핑몰포함", True),
        "denominator": ("모수_물품용역쇼핑몰", True),
        "name": ("기관", False),
    }
    key, reverse = sort_map.get(sort, sort_map["social_amount"])
    rows.sort(key=lambda r: r.get(key, 0) or 0, reverse=reverse)
    return {
        "summary": data.get("summary", {}),
        "rows": rows[:limit],
        "count": len(rows),
    }


@app.get("/api/social-enterprise/purchase/agency-monthly", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase_agency_monthly(
    agency: Optional[str] = Query(None, description="기관명 포함 검색"),
    month: Optional[str] = Query(None, description="YYYY-MM 형식"),
    limit: int = Query(500, ge=1, le=5000),
):
    """기관별 월별 사회적기업 구매율. 기본 수주율은 쇼핑몰 포함 기준."""
    data = _load_social_purchase_cache()
    rows = _filter_social_rows(data.get("agency_monthly", []), agency=agency, month=month)
    rows.sort(key=lambda r: (str(r.get("기관", "")), str(r.get("월", ""))))
    return {
        "summary": data.get("summary", {}),
        "rows": rows[:limit],
        "count": len(rows),
    }


@app.get("/api/social-enterprise/purchase/contracts", tags=["사회적기업 구매율"])
def get_social_enterprise_purchase_contracts(
    agency: Optional[str] = Query(None, description="기관명 포함 검색"),
    sector: Optional[str] = Query(None, description="공사/용역/물품/쇼핑몰"),
    limit: int = Query(500, ge=1, le=5000),
):
    """사회적기업 구매실적 원자료. 분자는 공사+용역+물품+쇼핑몰 사회적기업 수주액."""
    data = _load_social_purchase_cache()
    agency_q = str(agency or "").strip()
    sector_q = str(sector or "").strip()
    rows = []
    for row in data.get("social_contracts", []):
        if agency_q and agency_q not in str(row.get("기관", "")):
            continue
        if sector_q and sector_q != str(row.get("분야", "")):
            continue
        rows.append(dict(row))
    rows.sort(key=lambda r: (str(r.get("계약일", "")), str(r.get("기관", ""))), reverse=True)
    return {
        "summary": data.get("summary", {}),
        "rows": rows[:limit],
        "count": len(rows),
    }


@app.get("/api/social-enterprise/purchase/download", tags=["사회적기업 구매율"])
def download_social_enterprise_purchase(
    agency: Optional[str] = Query(None, description="기관명 포함 검색. 미지정 시 전체 검증 XLSX 다운로드"),
    sector: Optional[str] = Query(None, description="공사/용역/물품/쇼핑몰"),
):
    """사회적기업 구매율 검증 XLSX 다운로드."""
    agency_q = str(agency or "").strip()
    sector_q = str(sector or "").strip()
    if not agency_q and not sector_q:
        if not SOCIAL_PURCHASE_CACHE_XLSX.exists():
            raise HTTPException(
                status_code=503,
                detail=f"Social enterprise purchase XLSX not found: {SOCIAL_PURCHASE_CACHE_XLSX}",
            )
        filename = "social_enterprise_purchase_rate_2026.xlsx"
        return FileResponse(
            SOCIAL_PURCHASE_CACHE_XLSX,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    data = _load_social_purchase_cache()
    rows = []
    for row in data.get("social_contracts", []):
        if agency_q and agency_q not in str(row.get("기관", "")):
            continue
        if sector_q and sector_q != str(row.get("분야", "")):
            continue
        rows.append(dict(row))
    if not rows:
        raise HTTPException(
            status_code=404,
            detail="조건에 맞는 사회적기업 수주실적이 없습니다.",
        )
    rows.sort(key=lambda r: (str(r.get("계약일", "")), str(r.get("기관", ""))), reverse=True)
    columns = [
        "기관", "월", "분야", "계약일", "계약명", "계약액",
        "사회적기업수주액", "사회적기업업체", "사회적기업사업자번호", "사회적기업지분율",
    ]
    headers = [
        "기관", "월", "분야", "계약일", "계약명", "계약액",
        "사회적기업 수주액", "사회적기업 업체", "사업자등록번호", "지분율",
    ]
    from urllib.parse import quote
    suffix = agency_q or sector_q or "filtered"
    filename = quote(f"사회적기업_수주실적_{suffix}.xlsx")
    return _make_excel_response(rows, columns, headers, filename)


@app.get("/api/social-enterprise/dashboard", include_in_schema=False)
def social_enterprise_dashboard():
    """주요 정부 및 국가공공기관 부산 사회적기업 구매 실적 모니터링."""
    return HTMLResponse(
        content=r"""
<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>부산 사회적기업 구매 실적 모니터링</title>
  <style>
    :root {
      --bg: #eef6fb;
      --panel: rgba(255,255,255,0.96);
      --panel-strong: #ffffff;
      --line: #cfe0ef;
      --ink: #09224a;
      --muted: #5c6f8e;
      --blue: #1d6ed0;
      --blue-dark: #114a91;
      --cyan: #19c7d5;
      --green: #22b66d;
      --orange: #e38113;
      --shadow: 0 14px 34px rgba(22, 70, 120, 0.13);
      --shadow-soft: 0 8px 22px rgba(22, 70, 120, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Pretendard", "Noto Sans KR", "Segoe UI", Arial, sans-serif;
      color: var(--ink);
      overflow-x: hidden;
      background:
        linear-gradient(180deg, #f6fbff 0%, #eaf6f8 48%, #f5f9fc 100%);
    }
    body::before {
      content: "";
      position: fixed;
      inset: 0;
      pointer-events: none;
      background-image:
        linear-gradient(rgba(23, 93, 152, 0.05) 1px, transparent 1px),
        linear-gradient(90deg, rgba(23, 93, 152, 0.05) 1px, transparent 1px);
      background-size: 44px 44px;
      opacity: 0.32;
    }
    .shell { position: relative; width: min(1720px, calc(100vw - 56px)); margin: 0 auto; padding: 24px 0 52px; }
    .hero {
      position: relative;
      overflow: hidden;
      min-height: 188px;
      padding: 30px 38px;
      color: #fff;
      background:
        linear-gradient(112deg, rgba(11,134,183,0.98) 0%, rgba(35,111,223,0.98) 58%, rgba(7,29,210,0.98) 100%);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .hero::after {
      content: "";
      position: absolute;
      right: -120px;
      top: 0;
      width: 560px;
      height: 100%;
      border: 0;
      border-radius: 0;
      background: linear-gradient(90deg, rgba(255,255,255,0.00), rgba(255,255,255,0.12));
      transform: skewX(-17deg);
    }
    .hero-inner { position: relative; z-index: 1; display: grid; grid-template-columns: minmax(0, 1fr) 370px; gap: 34px; align-items: stretch; }
    .hero-inner > * { min-width: 0; }
    .eyebrow { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; font-size: 14px; font-weight: 800; }
    .badge {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      min-height: 30px;
      padding: 6px 12px;
      border-radius: 999px;
      border: 1px solid rgba(255,255,255,0.28);
      background: rgba(8, 32, 68, 0.26);
      color: #fff;
    }
    br.mobile-only { display: none; }
    h1 { margin: 0; font-size: clamp(34px, 3.55vw, 58px); line-height: 1.08; letter-spacing: 0; }
    .hero p { max-width: 900px; margin: 16px 0 0; font-size: 17px; line-height: 1.58; color: rgba(255,255,255,0.88); }
    .hero-note {
      margin-top: 18px;
      display: inline-flex;
      max-width: 980px;
      padding: 11px 15px;
      border: 1px solid rgba(255,255,255,0.24);
      border-radius: 8px;
      background: rgba(255,255,255,0.12);
      color: rgba(255,255,255,0.94);
      font-weight: 800;
      line-height: 1.55;
    }
    .hero-caution {
      margin-top: 10px;
      max-width: 980px;
      padding: 10px 14px;
      border: 1px solid rgba(255, 217, 128, 0.55);
      border-radius: 8px;
      background: rgba(255, 244, 205, 0.14);
      color: rgba(255,255,255,0.94);
      font-size: 14px;
      font-weight: 800;
      line-height: 1.55;
    }
    .hero-stat {
      align-self: stretch;
      border-radius: 8px;
      background: rgba(7, 25, 73, 0.80);
      border: 1px solid rgba(255,255,255,0.18);
      padding: 22px;
      display: grid;
      gap: 12px;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.08);
    }
    .hero-stat h2 { margin: 0; font-size: 18px; color: #5ef6d5; }
    .hero-number { font-size: 46px; line-height: 1; font-weight: 900; }
    .split-metric { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .mini-box { padding: 12px; border-radius: 8px; background: rgba(255,255,255,0.08); }
    .mini-box span { display: block; color: rgba(255,255,255,0.72); font-size: 13px; font-weight: 700; }
    .mini-box strong { display: block; margin-top: 6px; font-size: 24px; }
    .grid { display: grid; gap: 16px; }
    .kpi-grid { grid-template-columns: repeat(6, 1fr); margin-top: 16px; align-items: stretch; }
    .card {
      border: 1px solid var(--line);
      border-radius: 8px;
      background: linear-gradient(180deg, rgba(255,255,255,0.98), rgba(251,253,255,0.98));
      box-shadow: var(--shadow);
      padding: 18px 20px;
      min-width: 0;
    }
    .kpi.card { box-shadow: var(--shadow-soft); }
    .kpi { min-height: 146px; }
    .kpi.wide { grid-column: span 2; }
    .kpi.mid { grid-column: span 2; }
    .kpi h3, .section h2 { margin: 0; font-size: 18px; color: var(--blue-dark); }
    .kpi .value { margin-top: 10px; font-size: 38px; font-weight: 900; letter-spacing: 0; }
    .subline { margin-top: 8px; color: var(--muted); font-size: 15px; line-height: 1.45; }
    .delta.up { color: #168752; font-weight: 900; }
    .delta.down { color: #d84335; font-weight: 900; }
    .donut-wrap { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; align-items: stretch; margin-top: 8px; }
    .donut-metric { min-width: 0; padding: 10px; border: 1px solid #dbe9f6; border-radius: 8px; background: #f7fbff; display: grid; grid-template-columns: 92px minmax(90px, 1fr); gap: 10px; align-items: center; }
    .donut-metric.target { background: #f2f8ff; }
    .donut { width: 92px; height: 92px; border-radius: 50%; display: grid; place-items: center; background: conic-gradient(var(--accent, var(--blue)) var(--p), #dfeaf7 0); }
    .donut::before { content: ""; position: absolute; width: 60px; height: 60px; border-radius: 50%; background: #fff; }
    .donut strong { position: relative; z-index: 1; font-size: 22px; }
    .donut { position: relative; }
    .donut-copy span { display: block; color: var(--muted); font-size: 12px; font-weight: 900; }
    .donut-copy strong { display: block; margin-top: 4px; color: var(--blue-dark); font-size: 19px; line-height: 1.12; word-break: keep-all; }
    .donut-copy em { display: block; margin-top: 4px; color: var(--muted); font-size: 13px; font-style: normal; font-weight: 900; white-space: nowrap; }
    .achievement-pill { margin-top: 10px; padding: 9px 11px; border-radius: 8px; background: #eaf7f4; color: #087b73; font-size: 14px; font-weight: 900; display: flex; justify-content: space-between; gap: 12px; }
    .sector-bars { display: grid; gap: 9px; margin-top: 14px; }
    .bar-row { display: grid; grid-template-columns: 54px 1fr 82px; gap: 10px; align-items: center; font-size: 14px; font-weight: 800; }
    .bar { height: 9px; border-radius: 999px; background: #dfeaf7; overflow: hidden; }
    .bar span { display: block; height: 100%; border-radius: inherit; background: linear-gradient(90deg, var(--cyan), var(--blue)); }
    .major-contracts { margin-top: 14px; padding-top: 12px; border-top: 1px solid #e2edf7; display: grid; gap: 7px; }
    .major-contracts h4 { margin: 0 0 2px; color: var(--blue-dark); font-size: 15px; }
    .major-contract { display: grid; grid-template-columns: auto minmax(0, 1fr) auto; gap: 8px; align-items: center; font-size: 13px; }
    .major-contract .sector-tag { border-radius: 999px; background: #e9f2ff; color: var(--blue-dark); padding: 3px 7px; font-size: 11px; font-weight: 900; }
    .major-contract .contract-name { overflow: hidden; text-overflow: ellipsis; white-space: nowrap; font-weight: 800; color: #203550; }
    .major-contract .contract-amount { color: var(--blue-dark); font-weight: 900; white-space: nowrap; }
    .linebox { height: 146px; margin-top: 4px; }
    .top-list { display: grid; gap: 7px; margin-top: 10px; }
    .top-item { display: grid; grid-template-columns: auto 1fr auto; gap: 10px; align-items: center; padding: 9px 0; border-bottom: 1px solid #e6eff8; }
    .top-item:last-child { border-bottom: 0; }
    .top-rank { width: 26px; height: 26px; border-radius: 999px; display: inline-grid; place-items: center; background: #e9f2ff; color: var(--blue-dark); font-size: 12px; font-weight: 900; }
    .name { font-weight: 900; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .amount { font-weight: 900; color: var(--blue-dark); }
    .card-head { display: flex; align-items: center; justify-content: space-between; gap: 12px; }
    .mini-download {
      flex: 0 0 auto;
      border: 1px solid #bdd4ee;
      border-radius: 8px;
      background: #ffffff;
      color: var(--blue-dark);
      font-size: 12px;
      font-weight: 900;
      padding: 8px 10px;
      text-decoration: none;
      white-space: nowrap;
    }
    .kpi .value.kpi-pair { font-size: 34px; }
    .monthly-kpi {
      min-height: 198px;
      display: grid;
      align-content: center;
      gap: 12px;
    }
    .monthly-kpi h3 { font-size: 20px; }
    .monthly-kpi-main {
      color: var(--ink);
      font-size: 46px;
      line-height: 1;
      font-weight: 900;
      letter-spacing: 0;
    }
    .monthly-kpi-sub {
      display: flex;
      flex-wrap: wrap;
      gap: 8px 12px;
      align-items: center;
      color: var(--muted);
      font-size: 16px;
      font-weight: 900;
    }
    .monthly-kpi-sub .delta { font-size: 16px; }
    .kpi.top-card { grid-column: span 4; min-height: 244px; }
    .top-card .top-list { grid-template-columns: repeat(2, minmax(0, 1fr)); column-gap: 18px; }
    .top-card .top-item { padding: 7px 0; }
    .toolbar { display: flex; justify-content: space-between; gap: 12px; align-items: center; margin: 22px 0 10px; }
    .toolbar h2 { margin: 0; font-size: 24px; }
    .btn {
      appearance: none;
      border: 0;
      border-radius: 8px;
      background: #087b73;
      color: white;
      font-weight: 900;
      padding: 12px 16px;
      text-decoration: none;
      cursor: pointer;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 42px;
    }
    .btn.secondary { background: #1d6ed0; }
    .table-wrap { overflow: auto; border-radius: 8px; border: 1px solid var(--line); background: #fff; }
    table { width: 100%; border-collapse: collapse; min-width: 900px; }
    th, td { padding: 12px 14px; border-bottom: 1px solid #e6eff8; text-align: right; white-space: nowrap; }
    th:first-child, td:first-child { text-align: left; }
    th { background: #f3f8fd; color: var(--blue-dark); font-size: 14px; position: sticky; top: 0; z-index: 1; }
    td { font-size: 15px; }
    tbody tr:hover { background: #f7fbff; }
    .agency-grid { grid-template-columns: 1fr; }
    .agency-card {
      min-height: 0;
      display: grid;
      grid-template-columns: minmax(420px, 1.36fr) minmax(400px, 0.92fr);
      gap: 14px;
      align-items: stretch;
      padding: 14px;
      box-shadow: var(--shadow-soft);
    }
    .agency-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 10px; }
    .agency-main {
      border-radius: 8px;
      padding: 24px 28px;
      background: linear-gradient(135deg, #283998 0%, #2447ad 56%, #1197aa 100%);
      color: #fff;
      min-width: 0;
      display: grid;
      gap: 11px;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.14);
    }
    .agency-main h3 { margin: 0; font-size: 30px; line-height: 1.18; color: #fff; }
    .rank { flex: 0 0 auto; padding: 7px 11px; border-radius: 999px; background: rgba(255,255,255,0.16); color: #fff; font-weight: 900; font-size: 13px; }
    .agency-main-label { color: rgba(255,255,255,0.70); font-size: 14px; font-weight: 900; }
    .agency-main-total { font-size: 42px; line-height: 1; font-weight: 900; }
    .agency-breakdown { color: rgba(255,255,255,0.72); font-size: 14px; line-height: 1.45; font-weight: 800; }
    .agency-main-rate { margin-top: 4px; }
    .agency-main-rate span { display: block; color: rgba(255,255,255,0.76); font-size: 15px; font-weight: 900; }
    .agency-main-rate strong { display: inline-block; margin-top: 5px; color: #5ef6d5; font-size: 36px; line-height: 1; font-weight: 900; }
    .agency-main-rate em { color: rgba(255,255,255,0.88); font-style: normal; }
    .agency-spark { height: 136px; margin-top: 2px; border-radius: 8px; background: rgba(6,20,56,0.18); border: 1px solid rgba(255,255,255,0.20); overflow: hidden; }
    .agency-spark .spark { height: 136px; border: 0; background: transparent; }
    .agency-side { display: grid; grid-template-rows: auto auto auto; gap: 10px; align-content: start; }
    .agency-sector-panel { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px; }
    .agency-metrics { display: grid; grid-template-columns: 1.35fr 0.85fr; gap: 8px; }
    .agency-metric { padding: 10px; border: 1px solid #dce9f5; border-radius: 8px; background: #f7fbff; }
    .agency-metric span { display: block; color: var(--muted); font-size: 12px; font-weight: 800; }
    .agency-metric strong { display: block; margin-top: 4px; font-size: 20px; }
    .agency-metric em { display: block; margin-top: 3px; color: var(--muted); font-size: 12px; font-style: normal; font-weight: 800; }
    .spark { width: 100%; height: 74px; display: block; border-radius: 8px; background: linear-gradient(180deg, #f4f9ff 0%, #eef7ff 100%); border: 1px solid #e2eef9; }
    .sector-mini {
      min-width: 0;
      padding: 14px 16px;
      border: 1px solid #dce9f5;
      border-radius: 8px;
      background: #f7fbff;
      display: grid;
      align-content: space-between;
      min-height: 112px;
    }
    .sector-mini span { display: block; color: var(--muted); font-size: 14px; font-weight: 900; }
    .sector-mini strong { display: block; margin-top: 5px; color: var(--ink); font-size: 22px; line-height: 1.1; }
    .sector-mini em { display: block; margin-top: 6px; color: var(--muted); font-size: 12px; line-height: 1.3; font-style: normal; font-weight: 800; }
    .sector-progress { height: 8px; margin-top: 10px; border-radius: 999px; background: #dfeaf7; overflow: hidden; }
    .sector-progress i { display: block; height: 100%; border-radius: inherit; background: linear-gradient(90deg, var(--cyan), var(--blue)); }
    .agency-actions { display: flex; justify-content: flex-end; align-items: center; min-height: 38px; }
    .small-link {
      border: 1px solid #bdd4ee;
      border-radius: 8px;
      background: #fff;
      color: var(--blue-dark);
      font-weight: 900;
      padding: 8px 11px;
      text-decoration: none;
      font-size: 13px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 38px;
    }
    .monthly-bar-panel {
      padding: 12px 14px;
      border: 1px solid #dce9f5;
      border-radius: 8px;
      background: #f7fbff;
    }
    .monthly-bar-panel h4 {
      margin: 0 0 10px;
      color: var(--blue-dark);
      font-size: 15px;
      font-weight: 900;
    }
    .monthly-bars {
      height: 104px;
      display: grid;
      grid-auto-flow: column;
      grid-auto-columns: minmax(36px, 1fr);
      gap: 9px;
      align-items: end;
    }
    .monthly-bar {
      min-width: 0;
      display: grid;
      grid-template-rows: 20px 1fr 18px;
      gap: 4px;
      align-items: end;
      text-align: center;
    }
    .monthly-bar strong {
      color: var(--blue-dark);
      font-size: 12px;
      line-height: 1;
      font-weight: 900;
    }
    .monthly-bar i {
      display: block;
      width: 100%;
      min-height: 4px;
      border-radius: 999px 999px 4px 4px;
      background: linear-gradient(180deg, #3fe4dd, #1d6ed0);
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.36);
    }
    .monthly-bar span {
      color: var(--muted);
      font-size: 12px;
      line-height: 1;
      font-weight: 900;
    }
    .status { padding: 18px; background: #fff7e7; border: 1px solid #ffd899; border-radius: 8px; color: #744600; font-weight: 800; }
    @media (max-width: 1320px) {
      .hero-inner { grid-template-columns: 1fr; }
      .kpi-grid { grid-template-columns: repeat(2, 1fr); }
      .kpi.wide, .kpi.mid { grid-column: span 1; }
      .kpi.top-card { grid-column: span 2; }
      .agency-card { grid-template-columns: 1fr; }
    }
    @media (max-width: 840px) {
      .shell { width: min(100vw - 24px, 760px); padding-top: 12px; }
      .hero { padding: 22px; }
      .hero::after { display: none; }
      h1 { font-size: clamp(30px, 8.4vw, 40px); line-height: 1.12; overflow-wrap: break-word; }
      .hero p { font-size: 15px; line-height: 1.56; }
      .hero-note, .hero-caution { display: block; max-width: 100%; font-size: 13px; overflow-wrap: break-word; }
      .hero-stat { padding: 16px; }
      .split-metric, .kpi-grid, .agency-grid { grid-template-columns: 1fr; }
      .agency-card { padding: 14px; }
      .agency-main { padding: 22px 20px; }
      .agency-main h3 { font-size: 27px; }
      .agency-main-total { font-size: 40px; }
      .agency-sector-panel { grid-template-columns: 1fr; }
      .kpi.top-card { grid-column: span 1; }
      .top-card .top-list { grid-template-columns: 1fr; }
      .donut-wrap { grid-template-columns: 1fr; }
      .donut-metric { grid-template-columns: 80px 1fr; }
      .donut { width: 80px; height: 80px; }
      .donut::before { width: 52px; height: 52px; }
      .donut strong { font-size: 18px; }
      .toolbar { align-items: flex-start; flex-direction: column; }
      .toolbar .subline { margin-top: 0; }
    }
    @media (max-width: 520px) {
      .shell { width: 100%; padding: 12px 10px 36px; overflow-x: hidden; }
      .hero { padding: 20px 16px; }
      .eyebrow { gap: 6px; font-size: 12px; }
      .badge { min-height: 28px; padding: 6px 10px; }
      br.mobile-only { display: initial; }
      h1 { font-size: 29px; line-height: 1.12; }
      .hero-note, .hero-caution { overflow-wrap: anywhere; word-break: break-all; line-break: anywhere; }
      .hero-note, .hero-caution { font-size: 0; line-height: 0; }
      .hero-note::before, .hero-caution::before { display: block; font-size: 13px; line-height: 1.55; }
      .hero-note::before { content: "계산식: 사회적기업 수주액 누계 ÷ 총 발주액 누계"; }
      .hero-caution::before { content: "주의: 계약데이터 기준 추정치"; }
      .hero-number { font-size: 42px; }
      .card { padding: 17px 16px; }
      .kpi .value { font-size: 34px; }
      .mini-box strong { font-size: 22px; }
      .table-wrap { max-width: 100%; }
    }
  </style>
</head>
<body>
  <main class="shell">
    <section class="hero">
      <div class="hero-inner">
        <div>
          <div class="eyebrow">
            <span class="badge" id="apiBadge">API 연결 확인 중</span>
            <span class="badge" id="baseBadge">DB 기준 확인 중</span>
          </div>
          <h1>주요 정부 및 국가공공기관<br>부산 사회적기업<br class="mobile-only"> 구매 실적 모니터링</h1>
          <p>22개 정부 및 국가공공기관의 조달청 계약 데이터 기반 사회적기업 구매실적과 수주율을 집계합니다.</p>
          <div class="hero-note">계산식: (공사 + 용역 + 물품 + 조달청 나라장터 쇼핑몰 사회적기업 수주액 누계) ÷ (물품 + 용역 + 조달청 나라장터 쇼핑몰 발주액 누계)</div>
          <div class="hero-caution">주의: 조달 계약데이터 기준 추정치이며, 최종 공식 실적과 차이가 날 수 있습니다.</div>
        </div>
        <aside class="hero-stat">
          <h2>사회적기업 수주율 누계</h2>
          <div class="hero-number" id="heroRate">-</div>
          <div class="split-metric">
            <div class="mini-box"><span>사회적기업 수주액 누계</span><strong id="heroAward">-</strong></div>
            <div class="mini-box"><span>총 발주액 누계</span><strong id="heroBase">-</strong></div>
          </div>
          <div class="mini-box"><span>대상기관</span><strong id="heroAgencies">22개</strong></div>
        </aside>
      </div>
    </section>

    <section class="grid kpi-grid" id="kpiGrid"></section>

    <div class="toolbar">
      <h2>주요 기관별 누계 수주율</h2>
    </div>
    <section class="card section">
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>순위</th>
              <th>기관명</th>
              <th>발주액</th>
              <th>사회적기업 수주액</th>
              <th>수주율</th>
              <th>용역 발주액</th>
              <th>물품 발주액</th>
              <th>쇼핑몰 발주액</th>
            </tr>
          </thead>
          <tbody id="agencyTable"></tbody>
        </table>
      </div>
    </section>

    <div class="toolbar">
      <h2>주요 기관별 누계 수주율 카드</h2>
      <span class="subline">수주율 상위 순서로 기관별 월별 추이와 주요 계약을 확인합니다.</span>
    </div>
    <section class="grid agency-grid" id="agencyCards"></section>
  </main>
  <script>
    const targetRate = 5;
    const payloadUrl = "/api/social-enterprise/purchase";

    const num = (v) => Number(v || 0);
    const pct = (v, digits = 2) => `${num(v).toFixed(digits)}%`;
    function won(v) {
      const n = num(v);
      if (n >= 1000000000000) return `${(n / 1000000000000).toFixed(2)}조`;
      if (n >= 100000000) return `${(n / 100000000).toLocaleString("ko-KR", {maximumFractionDigits: 1})}억`;
      if (n >= 10000) return `${(n / 10000).toLocaleString("ko-KR", {maximumFractionDigits: 0})}만원`;
      return `${n.toLocaleString("ko-KR")}원`;
    }
    function htmlEscape(s) {
      return String(s ?? "").replace(/[&<>"']/g, (ch) => ({ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }[ch]));
    }
    function displayMonth(value) {
      const raw = String(value || "");
      const m = raw.match(/^(\d{4})-(\d{1,2})$/);
      if (m) return `${m[1]}.${Number(m[2])}월`;
      return raw || "이번달";
    }
    function lineSvg(rows, key = "수주율_쇼핑몰포함", labelKey = "월") {
      if (!rows.length) return "";
      const w = 440, h = 154, left = 26, right = 18, top = 28, bottom = 34;
      const vals = rows.map(r => num(r[key]));
      const max = Math.max(...vals, 0.01);
      const min = Math.min(...vals, 0);
      const span = Math.max(max - min, 0.01);
      const pts = rows.map((r, i) => {
        const x = left + (rows.length === 1 ? 0 : i * ((w - left - right) / (rows.length - 1)));
        const y = h - bottom - ((num(r[key]) - min) / span) * (h - top - bottom);
        return [x, y, r];
      });
      const path = pts.map((pt, i) => `${i ? "L" : "M"}${pt[0].toFixed(1)},${pt[1].toFixed(1)}`).join(" ");
      const baseLine = `<line x1="${left}" y1="${h - bottom}" x2="${w - right}" y2="${h - bottom}" stroke="#d6e5f4" stroke-width="1"/>`;
      const valueLabels = pts.map(pt => `<text x="${pt[0].toFixed(1)}" y="${Math.max(13, pt[1] - 9).toFixed(1)}" text-anchor="middle" fill="#0f4b8f" font-size="12" font-weight="900">${pct(pt[2][key], 2)}</text>`).join("");
      const monthLabels = pts.map(pt => `<text x="${pt[0].toFixed(1)}" y="${h - 9}" text-anchor="middle" fill="#5c6f8e" font-size="12" font-weight="900">${htmlEscape(String(pt[2][labelKey] || "").slice(5))}월</text>`).join("");
      const circles = pts.map(pt => `<circle cx="${pt[0].toFixed(1)}" cy="${pt[1].toFixed(1)}" r="4" fill="#fff" stroke="#1d6ed0" stroke-width="2"><title>${htmlEscape(pt[2][labelKey])}: ${pct(pt[2][key])}</title></circle>`).join("");
      return `<svg viewBox="0 0 ${w} ${h}" role="img" aria-label="월별 수주율 변동" preserveAspectRatio="none">${baseLine}<path d="${path}" fill="none" stroke="#1d6ed0" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/><path d="${path}" fill="none" stroke="#45e0e8" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/>${circles}${valueLabels}${monthLabels}</svg>`;
    }
    function cumulativeMonthly(rows) {
      const sorted = [...rows].sort((a, b) => String(a.월).localeCompare(String(b.월)));
      let social = 0;
      let base = 0;
      return sorted.map((row) => {
        social += num(row.사회적기업수주액);
        base += num(row.모수_물품용역쇼핑몰);
        return {
          ...row,
          사회적기업수주액_누계: social,
          모수_물품용역쇼핑몰_누계: base,
          수주율_쇼핑몰포함_누계: base ? (social / base) * 100 : 0,
        };
      });
    }
    function spark(rows, key = "수주율_쇼핑몰포함", title = "월별 수주율 변동") {
      if (!rows.length) {
        return `<svg class="spark" viewBox="0 0 900 136" preserveAspectRatio="none"><text x="450" y="74" text-anchor="middle" fill="#d9faff" font-size="14" font-weight="900">월별 실적 없음</text></svg>`;
      }
      const w = 900, h = 136, left = 54, right = 42, top = 34, bottom = 30;
      const vals = rows.map(r => num(r[key]));
      const max = Math.max(...vals, 0.01);
      const min = Math.min(...vals, 0);
      const span = Math.max(max - min, 0.01);
      const pts = rows.map((r, i) => {
        const x = left + (rows.length === 1 ? (w - left - right) / 2 : i * ((w - left - right) / (rows.length - 1)));
        const y = h - bottom - ((num(r[key]) - min) / span) * (h - top - bottom);
        return [x, y, r];
      });
      const path = pts.map((pt, i) => `${i ? "L" : "M"}${pt[0].toFixed(1)},${pt[1].toFixed(1)}`).join(" ");
      const area = `${path} L${pts.at(-1)[0].toFixed(1)},${h - bottom} L${pts[0][0].toFixed(1)},${h - bottom} Z`;
      const grid = `<line x1="${left}" y1="${h - bottom}" x2="${w - right}" y2="${h - bottom}" stroke="rgba(255,255,255,0.28)" stroke-width="1"/>`;
      const circles = pts.map(pt => `<circle cx="${pt[0].toFixed(1)}" cy="${pt[1].toFixed(1)}" r="5" fill="#fff" stroke="#45e0e8" stroke-width="2.4"><title>${htmlEscape(pt[2].월)}: ${pct(pt[2][key])}</title></circle>`).join("");
      const valueLabels = pts.map(pt => `<text x="${pt[0].toFixed(1)}" y="${Math.max(18, pt[1] - 11).toFixed(1)}" text-anchor="middle" fill="#ffffff" font-size="15" font-weight="900">${pct(pt[2][key])}</text>`).join("");
      const monthLabels = pts.map(pt => `<text x="${pt[0].toFixed(1)}" y="${h - 8}" text-anchor="middle" fill="#d9faff" font-size="14" font-weight="900">${htmlEscape(String(pt[2].월 || "").slice(5))}월</text>`).join("");
      return `<svg class="spark" viewBox="0 0 ${w} ${h}" preserveAspectRatio="none"><text x="${left}" y="19" fill="#d9faff" font-size="14" font-weight="900">${htmlEscape(title)}</text>${grid}<path d="${area}" fill="rgba(94,246,213,0.12)"/><path d="${path}" fill="none" stroke="#45e0e8" stroke-width="4" stroke-linecap="round" stroke-linejoin="round"/><path d="${path}" fill="none" stroke="#ffffff" stroke-width="1.4" stroke-linecap="round" stroke-linejoin="round"/>${circles}${valueLabels}${monthLabels}</svg>`;
    }
    function monthlyBars(rows) {
      if (!rows.length) return `<div class="monthly-bars empty">월별 실적 없음</div>`;
      const vals = rows.map(r => num(r.수주율_쇼핑몰포함));
      const max = Math.max(...vals, 0.01);
      return `<div class="monthly-bars">${rows.map((r) => {
        const rate = num(r.수주율_쇼핑몰포함);
        const height = Math.max(4, Math.min(100, (rate / max) * 100));
        return `<div class="monthly-bar"><strong>${pct(rate, 2)}</strong><i style="height:${height}%"></i><span>${htmlEscape(String(r.월 || "").slice(5))}월</span></div>`;
      }).join("")}</div>`;
    }
    function topCompanies(contracts) {
      const map = new Map();
      for (const c of contracts) {
        const name = String(c.사회적기업업체 || "업체명 미상");
        map.set(name, (map.get(name) || 0) + num(c.사회적기업수주액));
      }
      return [...map.entries()].sort((a, b) => b[1] - a[1]).slice(0, 10);
    }
    function topContracts(contracts) {
      return [...contracts]
        .sort((a, b) => num(b.사회적기업수주액) - num(a.사회적기업수주액))
        .slice(0, 5);
    }
    function sectorTile(label, social, base) {
      const socialAmount = num(social);
      const baseAmount = num(base);
      const hasBase = baseAmount > 0;
      const rate = hasBase ? (socialAmount / baseAmount) * 100 : 0;
      const main = hasBase ? `${won(socialAmount)} / ${pct(rate)}` : won(socialAmount);
      const sub = hasBase ? `발주액 ${won(baseAmount)}` : "사회적기업 수주액";
      const progress = hasBase ? Math.min(100, Math.max(2, rate)) : (socialAmount > 0 ? 100 : 0);
      return `<div class="sector-mini"><div><span>${label}</span><strong>${main}</strong><em>${sub}</em></div><div class="sector-progress"><i style="width:${progress}%"></i></div></div>`;
    }
    function latestMonthly(rows) {
      const sorted = [...rows].sort((a, b) => String(a.월).localeCompare(String(b.월)));
      return [sorted.at(-1) || {}, sorted.at(-2) || {}];
    }
    function render(payload) {
      const summary = payload.summary || {};
      const overall = payload.overall || {};
      const monthly = payload.monthly || [];
      const agencyRates = [...(payload.agency_rates || [])].sort((a, b) => num(b.수주율_쇼핑몰포함) - num(a.수주율_쇼핑몰포함));
      const agencyMonthly = payload.agency_monthly || [];
      const contracts = payload.social_contracts || [];
      const monthlyCumulative = cumulativeMonthly(monthly);
      const [latest, previous] = latestMonthly(monthly);
      const latestAward = num(latest.사회적기업수주액);
      const prevAward = num(previous.사회적기업수주액);
      const delta = latestAward - prevAward;
      const deltaRate = prevAward ? (delta / prevAward) * 100 : 0;
      const achieve = Math.min(100, (num(overall.수주율_쇼핑몰포함) / targetRate) * 100);
      const top = topCompanies(contracts);
      const topContractRows = topContracts(contracts);
      const sectorTotal = num(overall.공사_사회적기업수주액) + num(overall.용역_사회적기업수주액) + num(overall.물품_사회적기업수주액) + num(overall.쇼핑몰_사회적기업수주액);

      document.getElementById("apiBadge").textContent = "API 연결됨";
      document.getElementById("baseBadge").textContent = `DB 기준 ${summary.generated_at || "-"}`;
      document.getElementById("heroRate").textContent = pct(overall.수주율_쇼핑몰포함);
      document.getElementById("heroAward").textContent = won(overall.사회적기업수주액);
      document.getElementById("heroBase").textContent = won(overall.모수_물품용역쇼핑몰);
      document.getElementById("heroAgencies").textContent = `${summary.target_agency_count || agencyRates.length}개`;

      document.getElementById("kpiGrid").innerHTML = `
        <article class="card kpi wide monthly-kpi">
          <h3>월간 사회적기업 수주액·수주율 (${htmlEscape(displayMonth(latest.월))})</h3>
          <div class="monthly-kpi-main">${won(latestAward)} / ${pct(latest.수주율_쇼핑몰포함)}</div>
          <div class="monthly-kpi-sub">
            <span class="delta ${delta >= 0 ? "up" : "down"}">전월 대비 ${delta >= 0 ? "+" : ""}${deltaRate.toFixed(1)}%</span>
            <span>${delta >= 0 ? "증가" : "감소"} ${won(Math.abs(delta))}</span>
          </div>
        </article>
        <article class="card kpi mid">
          <h3>목표 달성률</h3>
          <div class="donut-wrap">
            <div class="donut-metric target">
              <div class="donut" style="--p:100%; --accent:#8fb8f2"><strong>${targetRate.toFixed(1)}%</strong></div>
              <div class="donut-copy">
                <span>목표 수주율</span>
                <strong>연간 기준</strong>
              </div>
            </div>
            <div class="donut-metric">
              <div class="donut" style="--p:${achieve}%; --accent:#0eb8c4"><strong>${pct(overall.수주율_쇼핑몰포함)}</strong></div>
              <div class="donut-copy">
                <span>현재 수주율</span>
                <strong>${pct(overall.수주율_쇼핑몰포함)}</strong>
                <em>조달계약 기준</em>
              </div>
            </div>
          </div>
          <div class="achievement-pill"><span>목표 대비 달성률</span><strong>${achieve.toFixed(1)}%</strong></div>
        </article>
        <article class="card kpi wide">
          <h3>월별 누계 수주율 변동</h3>
          <div class="linebox">${lineSvg(monthlyCumulative, "수주율_쇼핑몰포함_누계")}</div>
        </article>
        <article class="card kpi mid">
          <h3>분야별 사회적기업 수주액</h3>
          <div class="sector-bars">
            ${[
              ["공사", overall.공사_사회적기업수주액],
              ["용역", overall.용역_사회적기업수주액],
              ["물품", overall.물품_사회적기업수주액],
              ["쇼핑몰", overall.쇼핑몰_사회적기업수주액],
            ].map(([label, value]) => `<div class="bar-row"><span>${label}</span><div class="bar"><span style="width:${sectorTotal ? Math.max(2, num(value) / sectorTotal * 100) : 0}%"></span></div><strong>${won(value)}</strong></div>`).join("")}
          </div>
          <div class="major-contracts">
            <h4>주요 사회적기업 수주 항목</h4>
            ${topContractRows.map((row) => `<div class="major-contract"><span class="sector-tag">${htmlEscape(row.분야 || "-")}</span><span class="contract-name" title="${htmlEscape(row.계약명 || "")}">${htmlEscape(row.계약명 || "계약명 미상")}</span><span class="contract-amount">${won(row.사회적기업수주액)}</span></div>`).join("") || "<div class=\"subline\">표시할 수주 항목 없음</div>"}
          </div>
        </article>
        <article class="card kpi top-card">
          <div class="card-head">
            <h3>부산 사회적기업 수주액 상위 10개사</h3>
            <a class="mini-download" href="/api/social-enterprise/purchase/download">전체 수주실적 XLSX</a>
          </div>
          <div class="top-list">
            ${top.map(([name, amount], idx) => `<div class="top-item"><span class="top-rank">${idx + 1}</span><span class="name">${htmlEscape(name)}</span><span class="amount">${won(amount)}</span></div>`).join("") || "<div class=\"subline\">수주 실적 없음</div>"}
          </div>
        </article>
      `;

      document.getElementById("agencyTable").innerHTML = agencyRates.map((r, idx) => `
        <tr>
          <td>${idx + 1}</td>
          <td>${htmlEscape(r.기관)}</td>
          <td>${won(r.모수_물품용역쇼핑몰)}</td>
          <td>${won(r.사회적기업수주액)}</td>
          <td>${pct(r.수주율_쇼핑몰포함)}</td>
          <td>${won(r.용역_발주액)}</td>
          <td>${won(r.물품_발주액)}</td>
          <td>${won(r.쇼핑몰_발주액)}</td>
        </tr>
      `).join("");

      document.getElementById("agencyCards").innerHTML = agencyRates.map((r, idx) => {
        const months = agencyMonthly.filter(m => m.기관 === r.기관).sort((a, b) => String(a.월).localeCompare(String(b.월)));
        const cumulativeMonths = cumulativeMonthly(months);
        const encodedAgency = encodeURIComponent(r.기관);
        return `
          <article class="card agency-card">
            <div class="agency-main">
              <div class="agency-head">
                <h3>${htmlEscape(r.기관)}</h3>
                <span class="rank">수주율 ${idx + 1}위</span>
              </div>
              <div>
                <div class="agency-main-label">총 발주액</div>
                <div class="agency-main-total">${won(r.모수_물품용역쇼핑몰)}</div>
              </div>
              <div class="agency-main-rate">
                <span>사회적기업 수주액 (수주율)</span>
                <strong>${won(r.사회적기업수주액)} <em>(${pct(r.수주율_쇼핑몰포함)})</em></strong>
              </div>
              <div class="agency-spark">${spark(cumulativeMonths, "수주율_쇼핑몰포함_누계", "월별 누계 수주율")}</div>
            </div>
            <div class="agency-side">
              <div class="agency-sector-panel">
                ${sectorTile("공사", r.공사_사회적기업수주액, r.공사_발주액)}
                ${sectorTile("용역", r.용역_사회적기업수주액, r.용역_발주액)}
                ${sectorTile("물품", r.물품_사회적기업수주액, r.물품_발주액)}
                ${sectorTile("쇼핑몰", r.쇼핑몰_사회적기업수주액, r.쇼핑몰_발주액)}
              </div>
              <div class="agency-actions">
                <a class="small-link" href="/api/social-enterprise/purchase/download?agency=${encodedAgency}">주요 계약 XLSX</a>
              </div>
              <div class="monthly-bar-panel">
                <h4>월별 수주율</h4>
                ${monthlyBars(months)}
              </div>
            </div>
          </article>
        `;
      }).join("");
    }

    fetch(payloadUrl, { cache: "no-store" })
      .then(res => {
        if (!res.ok) throw new Error(`HTTP ${res.status}`);
        return res.json();
      })
      .then(render)
      .catch(err => {
        document.querySelector(".shell").insertAdjacentHTML("beforeend", `<div class="status">사회적기업 구매율 API 응답을 불러오지 못했습니다. ${htmlEscape(err.message)}</div>`);
        document.getElementById("apiBadge").textContent = "API 연결 실패";
      });
  </script>
</body>
</html>
""",
    )

@app.get("/api/agency/search", tags=["기관 검색"])
def search_agency(q: str = Query(..., min_length=1, description="검색할 기관명")):
    """특정 수요기관의 총괄 수주율, 금액, 주요 유출계약 정보 검색"""
    cache = load_cache()
    agency_details = cache.get("12_기관별_상세", {})
    
    results = {}
    q_clean = q.strip()
    for unit, details in agency_details.items():
        if q_clean in unit:
            results[unit] = details
            
    return {
        "generated_at": cache.get("generated_at"),
        "검색어": q_clean,
        "검색결과": results
    }

@app.get("/api/agency/suui-search", tags=["기관 검색"])
def search_agency_suui(q: str = Query(..., min_length=1, description="검색할 기관명")):
    """특정 수요기관의 수의계약 유출 현황 검색 (분야별 발주/수주, 유출계약 목록)"""
    cache = load_cache()
    suui_details = cache.get("13_수의계약_기관별_상세", {})
    
    results = {}
    q_clean = q.strip()
    for unit, details in suui_details.items():
        if q_clean in unit:
            results[unit] = details
            
    return {
        "generated_at": cache.get("generated_at"),
        "검색어": q_clean,
        "검색결과": results
    }

@app.get("/api/shopping-contract", tags=["종합쇼핑몰"])
def get_shopping_contract():
    """종합쇼핑몰 유출 현황 (유출 기관별, 유출 계약별)"""
    cache = load_cache()
    return {
        "generated_at": cache.get("generated_at"),
        "유출_쇼핑몰": cache.get("14_쇼핑몰_유출", []),
        "유출_기관별": cache.get("14_쇼핑몰_유출_기관별", []),
        "구군_상세": cache.get("15_쇼핑몰_구군_상세", {}),
        "유형별": cache.get("16_쇼핑몰_유형별", {}),
    }

@app.get("/api/agency/shop-search", tags=["기관 검색"])
def search_agency_shop(q: str = Query(..., min_length=1, description="검색할 기관명")):
    """특정 수요기관의 쇼핑몰 유출 현황 검색"""
    cache = load_cache()
    shop_details = cache.get("14_쇼핑몰_기관별_상세", {})
    
    results = {}
    q_clean = q.strip()
    for unit, details in shop_details.items():
        if q_clean in unit:
            results[unit] = details
            
    return {
        "generated_at": cache.get("generated_at"),
        "검색어": q_clean,
        "검색결과": results
    }


# ════════════════════════════════════════════
#   업체 검색 API (busan_companies_master.db 직접 쿼리)
# ════════════════════════════════════════════

@app.get("/api/company/license-list", tags=["업체 검색"])
def get_license_list(q: Optional[str] = Query(None, description="업종명 검색어 (포함 검색)")):
    """면허업종 목록 + 업체수 (업체수 내림차순). q 파라미터로 필터링 가능"""
    try:
        conn = _get_company_db()
        if q and q.strip():
            rows = conn.execute(
                "SELECT indstrytyNm, COUNT(DISTINCT bizno) cnt FROM company_industry "
                "WHERE indstrytyNm LIKE ? GROUP BY indstrytyNm ORDER BY cnt DESC",
                (f"%{q.strip()}%",)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT indstrytyNm, COUNT(DISTINCT bizno) cnt FROM company_industry "
                "GROUP BY indstrytyNm ORDER BY cnt DESC"
            ).fetchall()
        conn.close()
        return {
            "총업종수": len(rows),
            "업종목록": [{"업종명": r["indstrytyNm"], "업체수": r["cnt"]} for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/company/product-list", tags=["업체 검색"])
def get_product_list(q: Optional[str] = Query(None, description="품명 검색어 (포함 검색)")):
    """대표품명 목록 + 업체수 (업체수 내림차순). q 파라미터로 필터링 가능"""
    try:
        conn = _get_company_db()
        if q and q.strip():
            rows = conn.execute(
                "SELECT rprsntDtlPrdnm, COUNT(*) cnt FROM company_master "
                "WHERE rprsntDtlPrdnm IS NOT NULL AND rprsntDtlPrdnm != '' "
                "AND rprsntDtlPrdnm LIKE ? "
                "GROUP BY rprsntDtlPrdnm ORDER BY cnt DESC",
                (f"%{q.strip()}%",)
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT rprsntDtlPrdnm, COUNT(*) cnt FROM company_master "
                "WHERE rprsntDtlPrdnm IS NOT NULL AND rprsntDtlPrdnm != '' "
                "GROUP BY rprsntDtlPrdnm ORDER BY cnt DESC"
            ).fetchall()
        conn.close()
        return {
            "총품명수": len(rows),
            "품명목록": [{"품명": r["rprsntDtlPrdnm"], "업체수": r["cnt"]} for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/company/license-search", tags=["업체 검색"])
def search_by_license(
    q: str = Query(..., min_length=1, description="면허업종명 (정확 매칭 또는 포함 검색)"),
    exact: bool = Query(False, description="True면 정확 매칭, False면 포함 검색"),
    limit: int = Query(200, ge=1, le=5000, description="최대 반환 건수"),
):
    """면허업종으로 업체 검색 → 업체명/대표자/소재지/주소/본사구분/개업일"""
    try:
        conn = _get_company_db()
        q_clean = q.strip()
        where = "i.indstrytyNm = ?" if exact else "i.indstrytyNm LIKE ?"
        param = q_clean if exact else f"%{q_clean}%"
        rows = conn.execute(f"""
            SELECT DISTINCT c.corpNm, c.bizno, c.ceoNm, c.rgnNm, c.adrs, c.dtlAdrs,
                   c.hdoffceDivNm, c.corpBsnsDivNm, c.opbizDt, c.rgstDt,
                   c.rprsntDtlPrdnm, i.indstrytyNm, i.rprsntIndstrytyYn
            FROM company_industry i
            JOIN company_master c ON i.bizno = c.bizno
            WHERE {where}
            ORDER BY c.corpNm
            LIMIT ?
        """, (param, limit)).fetchall()
        conn.close()
        return {
            "검색어": q_clean,
            "검색결과수": len(rows),
            "업체목록": [{
                "업체명": r["corpNm"], "사업자번호": r["bizno"],
                "대표자": r["ceoNm"], "소재지": r["rgnNm"],
                "주소": r["adrs"], "상세주소": r["dtlAdrs"],
                "본사구분": r["hdoffceDivNm"], "업체구분": r["corpBsnsDivNm"],
                "대표품명": r["rprsntDtlPrdnm"],
                "면허업종": r["indstrytyNm"], "대표업종여부": r["rprsntIndstrytyYn"],
                "개업일": r["opbizDt"], "등록일": r["rgstDt"],
            } for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/company/product-search", tags=["업체 검색"])
def search_by_product(
    q: str = Query(..., min_length=1, description="대표품명 검색어 (포함 검색)"),
    exact: bool = Query(False, description="True면 정확 매칭, False면 포함 검색"),
    limit: int = Query(200, ge=1, le=5000, description="최대 반환 건수"),
):
    """대표품명(세부품명)으로 업체 검색 → 업체명/대표자/소재지/대표품명"""
    try:
        conn = _get_company_db()
        q_clean = q.strip()
        where = "rprsntDtlPrdnm = ?" if exact else "rprsntDtlPrdnm LIKE ?"
        param = q_clean if exact else f"%{q_clean}%"
        rows = conn.execute(f"""
            SELECT corpNm, bizno, ceoNm, rgnNm, adrs, dtlAdrs,
                   hdoffceDivNm, corpBsnsDivNm, rprsntDtlPrdnm, opbizDt, rgstDt
            FROM company_master
            WHERE {where}
            ORDER BY corpNm
            LIMIT ?
        """, (param, limit)).fetchall()
        conn.close()
        return {
            "검색어": q_clean,
            "검색결과수": len(rows),
            "업체목록": [{
                "업체명": r["corpNm"], "사업자번호": r["bizno"],
                "대표자": r["ceoNm"], "소재지": r["rgnNm"],
                "주소": r["adrs"], "상세주소": r["dtlAdrs"],
                "본사구분": r["hdoffceDivNm"], "업체구분": r["corpBsnsDivNm"],
                "대표품명": r["rprsntDtlPrdnm"],
                "개업일": r["opbizDt"], "등록일": r["rgstDt"],
            } for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

# ── 물품 대분류 코드 ↔ 공식 분류명 매핑 (조달청 UNSPSC 기준) ──
UNSPSC_CATEGORIES = {
    "10": "농축수산물", "11": "광물/금속/비금속", "12": "화학약품",
    "14": "종이/고무/플라스틱원재료", "15": "연료/윤활유",
    "20": "광업/유전/가스장비", "21": "농림어업장비",
    "22": "건설/건물유지관리장비", "23": "산업생산/제조장비",
    "24": "산업취급/보관장비", "25": "차량/수송장비",
    "26": "동력/발전장비", "27": "공구/일반기계",
    "30": "구조물/건축자재", "31": "배관/난방자재", "32": "배선/통신자재",
    "39": "전기/조명장비", "40": "냉난방/공조/환기",
    "41": "실험/측정/관측장비", "42": "의료/보건장비",
    "43": "정보통신/방송장비", "44": "사무용기기/용품",
    "46": "안전/방호/소방", "47": "환경/수처리장비",
    "48": "세정/위생장비", "49": "체육/레저/여행",
    "50": "식품/음료/담배", "51": "약품/의약품",
    "52": "가정주방/세탁/가전", "53": "피복/섬유/개인용품",
    "54": "시계/보석/귀금속", "55": "인쇄/출판/광고",
    "56": "가구/인테리어", "60": "악기/게임/완구",
    "70": "서비스(임대/관리)", "72": "건설/유지보수서비스",
    "73": "산업생산/유지보수서비스", "76": "환경/산업청소서비스",
    "77": "교육/훈련서비스", "78": "운송/보관서비스",
    "80": "경영/마케팅서비스", "81": "정보시스템서비스",
    "82": "디자인/엔지니어링서비스", "83": "공공서비스/행정서비스",
    "84": "금융/보험서비스", "85": "보건의료서비스",
    "86": "교육/문화/예술서비스",
    "90": "국방/공공질서", "92": "소방/구조서비스",
    "93": "정치/시민활동",
}
# 역매핑: 분류명 → 코드 (포함검색용)
_CAT_NAME_TO_CODES = {}
for _code, _name in UNSPSC_CATEGORIES.items():
    for _part in _name.replace("/", " ").split():
        _CAT_NAME_TO_CODES.setdefault(_part, []).append(_code)

@app.get("/api/company/category-list", tags=["업체 검색"])
def get_category_list(q: Optional[str] = Query(None, description="분류코드(숫자) 또는 분류명(한글) 검색어")):
    """물품 대분류 목록 + 공식분류명 + 업체수. q로 코드/이름 필터링 가능"""
    try:
        conn = _get_company_db()
        rows = conn.execute("""
            SELECT SUBSTR(rprsntDtlPrdnmNo, 1, 2) AS cat2, COUNT(*) AS cnt
            FROM company_master
            WHERE rprsntDtlPrdnm IS NOT NULL AND rprsntDtlPrdnm != ''
            GROUP BY cat2 ORDER BY cnt DESC
        """).fetchall()
        conn.close()

        result = []
        qc = q.strip() if q else None
        for r in rows:
            code = r["cat2"]
            cat_name = UNSPSC_CATEGORIES.get(code, "기타")
            if qc:
                if qc.isdigit():
                    if not code.startswith(qc):
                        continue
                else:
                    if qc not in cat_name:
                        continue
            result.append({"분류코드": code, "분류명": cat_name, "업체수": r["cnt"]})
        return {"총분류수": len(result), "분류목록": result}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/company/category-search", tags=["업체 검색"])
def search_by_category(
    q: str = Query(..., min_length=1, description="분류코드(예:50) / 분류명(예:식품) / 대표품명(예:우유)"),
    exact: bool = Query(False, description="True면 정확 매칭, False면 포함 검색"),
    limit: int = Query(500, ge=1, le=5000, description="최대 반환 건수"),
):
    """물품 분류코드 / 분류명 / 대표품명으로 업체 검색 (자동 판별)

    - 숫자 → 분류코드 검색 (rprsntDtlPrdnmNo 앞자리 매칭)
    - 분류명 매칭 → 해당 코드의 업체 전체 반환 (예: '식품' → 코드50 전체)
    - 그 외 → 대표품명 키워드 검색 (예: '우유', '컴퓨터')
    """
    try:
        conn = _get_company_db()
        q_clean = q.strip()

        # 1) 숫자 → 분류코드 검색
        if q_clean.isdigit():
            search_mode = "분류코드"
            if exact:
                where = "SUBSTR(rprsntDtlPrdnmNo, 1, 2) = ?"
                param = q_clean
            else:
                where = "rprsntDtlPrdnmNo LIKE ?"
                param = f"{q_clean}%"
        else:
            # 2) 분류명 매칭 시도 (예: '식품' → 코드50, '음료' → 코드50)
            matched_codes = set()
            for part in q_clean.replace("/", " ").split():
                for cat_code, cat_name in UNSPSC_CATEGORIES.items():
                    if part in cat_name:
                        matched_codes.add(cat_code)

            if matched_codes:
                search_mode = "분류명"
                placeholders = ",".join("?" * len(matched_codes))
                where = f"SUBSTR(rprsntDtlPrdnmNo, 1, 2) IN ({placeholders})"
                param = tuple(sorted(matched_codes))
            else:
                # 3) 대표품명 키워드 검색
                search_mode = "대표품명"
                if exact:
                    where = "rprsntDtlPrdnm = ?"
                    param = q_clean
                else:
                    where = "rprsntDtlPrdnm LIKE ?"
                    param = f"%{q_clean}%"

        # 분류명 검색은 IN 절이므로 tuple, 나머지는 단일 param
        if search_mode == "분류명":
            rows = conn.execute(f"""
                SELECT corpNm, bizno, ceoNm, rgnNm, adrs, dtlAdrs,
                       hdoffceDivNm, corpBsnsDivNm, mnfctDivNm,
                       rprsntDtlPrdnmNo, rprsntDtlPrdnm, opbizDt, rgstDt
                FROM company_master
                WHERE rprsntDtlPrdnm IS NOT NULL AND rprsntDtlPrdnm != ''
                  AND {where}
                ORDER BY corpNm
                LIMIT ?
            """, (*param, limit)).fetchall()
            matched_names = [f"{c}({UNSPSC_CATEGORIES[c]})" for c in sorted(matched_codes)]
        else:
            rows = conn.execute(f"""
                SELECT corpNm, bizno, ceoNm, rgnNm, adrs, dtlAdrs,
                       hdoffceDivNm, corpBsnsDivNm, mnfctDivNm,
                       rprsntDtlPrdnmNo, rprsntDtlPrdnm, opbizDt, rgstDt
                FROM company_master
                WHERE rprsntDtlPrdnm IS NOT NULL AND rprsntDtlPrdnm != ''
                  AND {where}
                ORDER BY corpNm
                LIMIT ?
            """, (param, limit)).fetchall()
            matched_names = None
        conn.close()

        resp = {
            "검색어": q_clean,
            "검색방식": search_mode,
            "검색결과수": len(rows),
        }
        if matched_names:
            resp["매칭분류"] = matched_names
        resp["업체목록"] = [{
            "업체명": r["corpNm"], "사업자번호": r["bizno"],
            "대표자": r["ceoNm"], "소재지": r["rgnNm"],
            "주소": r["adrs"], "상세주소": r["dtlAdrs"],
            "본사구분": r["hdoffceDivNm"], "업체구분": r["corpBsnsDivNm"],
            "제조구분": r["mnfctDivNm"],
            "분류코드": r["rprsntDtlPrdnmNo"],
            "분류명": UNSPSC_CATEGORIES.get((r["rprsntDtlPrdnmNo"] or "")[:2], ""),
            "대표품명": r["rprsntDtlPrdnm"],
            "개업일": r["opbizDt"], "등록일": r["rgstDt"],
        } for r in rows]
        return resp
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/company/manufacturers", tags=["업체 검색"])
def get_manufacturers(
    limit: int = Query(5000, ge=1, le=10000, description="최대 반환 건수"),
    format: Optional[str] = Query(None, description="'csv'면 CSV 파일 다운로드, 미지정 시 JSON"),
):
    """제조업체 전체 목록 (mnfctDivNm='제조'). format=csv로 CSV 다운로드 가능"""
    try:
        conn = _get_company_db()
        rows = conn.execute("""
            SELECT c.corpNm, c.bizno, c.ceoNm, c.rgnNm, c.adrs, c.dtlAdrs,
                   c.hdoffceDivNm, c.corpBsnsDivNm, c.mnfctDivNm,
                   c.rprsntDtlPrdnmNo, c.rprsntDtlPrdnm,
                   c.rprsntIndstrytyNm, c.opbizDt, c.rgstDt
            FROM company_master c
            WHERE c.mnfctDivNm = '제조'
            ORDER BY c.corpNm
            LIMIT ?
        """, (limit,)).fetchall()
        conn.close()

        # CSV 다운로드
        if format and format.strip().lower() == "csv":
            import csv, io
            output = io.StringIO()
            writer = csv.writer(output)
            headers = ["업체명","사업자번호","대표자","소재지","주소","상세주소",
                       "본사구분","업체구분","제조구분","분류코드","대표품명","대표업종","개업일","등록일"]
            writer.writerow(headers)
            for r in rows:
                writer.writerow([
                    r["corpNm"], r["bizno"], r["ceoNm"], r["rgnNm"],
                    r["adrs"], r["dtlAdrs"], r["hdoffceDivNm"], r["corpBsnsDivNm"],
                    r["mnfctDivNm"], r["rprsntDtlPrdnmNo"], r["rprsntDtlPrdnm"],
                    r["rprsntIndstrytyNm"], r["opbizDt"], r["rgstDt"],
                ])
            from starlette.responses import StreamingResponse
            csv_bytes = output.getvalue().encode("utf-8-sig")
            return StreamingResponse(
                iter([csv_bytes]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=busan_manufacturers.csv"}
            )

        return {
            "총제조업체수": len(rows),
            "업체목록": [{
                "업체명": r["corpNm"], "사업자번호": r["bizno"],
                "대표자": r["ceoNm"], "소재지": r["rgnNm"],
                "주소": r["adrs"], "상세주소": r["dtlAdrs"],
                "본사구분": r["hdoffceDivNm"], "업체구분": r["corpBsnsDivNm"],
                "제조구분": r["mnfctDivNm"],
                "분류코드": r["rprsntDtlPrdnmNo"], "대표품명": r["rprsntDtlPrdnm"],
                "대표업종": r["rprsntIndstrytyNm"],
                "개업일": r["opbizDt"], "등록일": r["rgstDt"],
            } for r in rows]
        }
    except Exception as e:
        return {"error": str(e)}

# ── 월별 추이 ──
MONTHLY_CACHE_FILE = 'monthly_cache.json'

def load_monthly_cache():
    try:
        with open(MONTHLY_CACHE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"[ERROR] 월별 캐시 로드 실패: {e}")
        return {"error": f"월별 캐시 파일 로드 실패: {str(e)}"}

@app.get("/api/monthly-trend", tags=["종합분석"])
def get_monthly_trend():
    """월별 누계/단월 수주율 추이 (전체, 그룹별, 분야별) + 변동 원인"""
    mc = load_monthly_cache()
    ac = load_cache()
    # 소그룹 누계의 최종 값을 API 캐시 기준으로 보정 (데이터 정합성)
    누계_소그룹 = mc.get("누계_소그룹", {})
    누계_소그룹분야 = mc.get("누계_소그룹분야", {})
    sg_ranking = ac.get("5_기관랭킹_소그룹", {})
    for sg_key, sg_data in sg_ranking.items():
        # 합계 보정
        if sg_key in 누계_소그룹 and 누계_소그룹[sg_key]:
            last = 누계_소그룹[sg_key][-1]
            last['발주액'] = sg_data.get('발주액', last['발주액'])
            last['수주액'] = sg_data.get('수주액', last['수주액'])
            last['수주율'] = sg_data.get('수주율', last['수주율'])
        # 분야별 보정
        api_sec = sg_data.get('분야별', {})
        if sg_key in 누계_소그룹분야 and api_sec:
            for sec_name, sec_vals in api_sec.items():
                sec_list = 누계_소그룹분야[sg_key].get(sec_name, [])
                if sec_list:
                    sec_last = sec_list[-1]
                    sec_last['발주액'] = sec_vals.get('발주액', sec_last['발주액'])
                    sec_last['수주액'] = sec_vals.get('수주액', sec_last['수주액'])
                    sec_last['수주율'] = sec_vals.get('수주율', sec_last['수주율'])
    return {
        "generated_at": mc.get("generated_at"),
        "year": mc.get("year"),
        "months": mc.get("months", []),
        "누계_그룹": mc.get("누계_그룹", {}),
        "누계_분야": mc.get("누계_분야", {}),
        "월간_그룹": mc.get("월간_그룹", {}),
        "월간_분야": mc.get("월간_분야", {}),
        "변동분석": mc.get("변동분석", {}),
        "분야변동": mc.get("분야변동", {}),
        "누계_소그룹": 누계_소그룹,
        "월간_소그룹": mc.get("월간_소그룹", {}),
        "누계_소그룹분야": 누계_소그룹분야,
        "월간_소그룹분야": mc.get("월간_소그룹분야", {}),
        "소그룹_분야변동": mc.get("소그룹_분야변동", {}),
        "소그룹_변동분석": mc.get("소그룹_변동분석", {}),
    }

@app.get("/api/monthly-trend/agency", tags=["종합분석"])
def get_monthly_trend_agency(q: str = Query(..., min_length=1, description="검색할 기관명")):
    """특정 기관의 월별 누계/단월 수주율 추이"""
    mc = load_monthly_cache()
    기관별 = mc.get("기관별", {})
    
    results = {}
    q_clean = q.strip()
    for unit, details in 기관별.items():
        if q_clean in unit:
            results[unit] = details
            
    return {
        "generated_at": mc.get("generated_at"),
        "months": mc.get("months", []),
        "검색어": q_clean,
        "검색결과": results
    }


# ════════════════════════════════════════════
#   챗봇 업체 검색 API (chatbot_company.db)
# ════════════════════════════════════════════

import logging
import datetime
from typing import Literal
from fastapi import Request

logger = logging.getLogger("chatbot_api")

CHATBOT_DB = os.environ.get("CHATBOT_DB", os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chatbot_company.db'))

def _get_chatbot_db():
    conn = sqlite3.connect(CHATBOT_DB)
    conn.row_factory = sqlite3.Row
    return conn

def _get_status_filter_sql(status_filter: str) -> str:
    if status_filter == "all":
        return ""
    elif status_filter == "active_only":
        return " AND IFNULL(cbs.business_status, 'unknown') = 'active' "
    elif status_filter == "needs_check":
        return " AND (IFNULL(cbs.business_status, 'unknown') IN ('unknown', 'api_failed', 'quota_exceeded') OR IFNULL(cbs.business_status_freshness, 'not_checked') != 'fresh') "
    return " AND NOT (IFNULL(cbs.business_status, 'unknown') IN ('closed', 'suspended') AND IFNULL(cbs.business_status_freshness, 'not_checked') = 'fresh') "

def _build_chatbot_response(rows, meta=None, error=None, validity_filter="valid_only"):
    if error:
        return {
            "meta": {},
            "candidates": [],
            "company_source_status": "api_failed",
            "company_cache_mode": "none",
            "company_cache_used": False,
            "company_search_status": "failed",
            "error": error
        }

    candidates = []
    counts_by_type = {}
    latest_refresh = {}

    for r in rows:
        c_raw = dict(r)

        for list_field in ["license_or_business_type", "main_products"]:
            if c_raw.get(list_field):
                c_raw[list_field] = c_raw[list_field].split("|")
            else:
                c_raw[list_field] = []

        for json_field in ["candidate_types", "source_refs"]:
            if c_raw.get(json_field):
                try:
                    c_raw[json_field] = json.loads(c_raw[json_field])
                except Exception:
                    c_raw[json_field] = []
            else:
                c_raw[json_field] = []
        
        c_raw["shopping_mall_flags"] = []
        c_raw["mas_product_summary"] = []
        c_raw["shopping_mall_product_summary"] = []

        # Phase 6-G: whitelist 기반 shopping_mall_flags 필터링
        ALLOWED_SM_FLAGS = {
            "shopping_mall_registered", "mas_registered",
            "third_party_unit_price_registered", "general_unit_price_registered",
            "excellent_procurement_registered"
        }
        ALLOWED_SM_CONTRACT_TYPES = {
            "mas", "third_party_unit_price", "general_unit_price",
            "excellent_procurement", "unknown"
        }

        raw_shopping_flags = c_raw.get("shopping_mall_flags_raw")
        if "shopping_mall_flags_raw" in c_raw:
            del c_raw["shopping_mall_flags_raw"]
            
        if raw_shopping_flags:
            flags = set()
            for group in raw_shopping_flags.split(","):
                for flag in group.split("|"):
                    if flag and flag in ALLOWED_SM_FLAGS:
                        flags.add(flag)
            
            c_raw["shopping_mall_flags"] = list(flags)
            
            if "shopping_mall_registered" in flags and "shopping_mall_supplier" not in c_raw["candidate_types"]:
                c_raw["candidate_types"].append("shopping_mall_supplier")

        raw_policies = c_raw.get("policy_subtypes_raw")
        if "policy_subtypes_raw" in c_raw:
            del c_raw["policy_subtypes_raw"]
            
        c_raw["sme_competition_product"] = bool(c_raw.get("is_sme_competition_product", 0))
        if "is_sme_competition_product" in c_raw:
            del c_raw["is_sme_competition_product"]

        c_raw["policy_subtypes"] = []
        c_raw["policy_validity_summary"] = {}

        if raw_policies:
            for cert in set(raw_policies.split("|")):
                if ":" in cert:
                    subtype, status = cert.split(":")
                    c_raw["policy_validity_summary"][subtype] = status
                    if status == "valid":
                        c_raw["policy_subtypes"].append(subtype)
                else:
                    c_raw["policy_subtypes"].append(cert)

            if c_raw["policy_subtypes"] or c_raw["policy_validity_summary"]:
                if "policy_company_certification" not in c_raw["source_refs"]:
                    c_raw["source_refs"].append("policy_company_certification")

            if c_raw["policy_subtypes"] and "policy_company" not in c_raw["candidate_types"]:
                c_raw["candidate_types"].append("policy_company")

        raw_certified_types = c_raw.get("certified_product_types_raw")
        raw_certified_summary = c_raw.get("certified_product_summary_raw")
        
        if "certified_product_types_raw" in c_raw:
            del c_raw["certified_product_types_raw"]
        if "certified_product_summary_raw" in c_raw:
            del c_raw["certified_product_summary_raw"]

        c_raw["certified_product_types"] = []
        c_raw["certified_product_summary"] = []

        if raw_certified_types:
            for cert in set(raw_certified_types.split("|")):
                if not cert: continue
                parts = cert.split(":")
                if len(parts) == 5:
                    ctype, is_priority, is_innov, is_exc, status = parts
                    if status == "valid":
                        c_raw["certified_product_types"].append(ctype)
                        if is_priority == '1' and "priority_purchase_product" not in c_raw["candidate_types"]:
                            c_raw["candidate_types"].append("priority_purchase_product")
                        if is_innov == '1' and "innovation_product" not in c_raw["candidate_types"]:
                            c_raw["candidate_types"].append("innovation_product")
                        if is_exc == '1' and "excellent_procurement_product" not in c_raw["candidate_types"]:
                            c_raw["candidate_types"].append("excellent_procurement_product")

        if raw_certified_summary:
            items = []
            for cert in raw_certified_summary.split("|||"):
                if not cert: continue
                parts = cert.split("^^")
                if len(parts) == 5:
                    ctype, pname, status, exp_date, src = parts
                    # validity_filter에 따른 필터링 로직 추가
                    if validity_filter == "valid_only" and status != "valid":
                        continue
                    if validity_filter == "include_unknown" and status not in ("valid", "unknown"):
                        continue
                        
                    items.append({
                        "certification_type": ctype,
                        "product_name": pname,
                        "validity_status": status,
                        "expiration_date": exp_date if exp_date else None,
                        "source_name": src
                    })
            # Sort valid first
            items.sort(key=lambda x: (x["validity_status"] != "valid", x["expiration_date"] or ""))
            c_raw["certified_product_summary"] = items[:5] # Max 5
            
            if items and "certified_product" not in c_raw["source_refs"]:
                c_raw["source_refs"].append("certified_product")

        raw_mas_summary = c_raw.get("mas_product_summary_raw")
        raw_sm_summary = c_raw.get("shopping_mall_product_summary_raw")
        
        if "mas_product_summary_raw" in c_raw:
            del c_raw["mas_product_summary_raw"]
        if "shopping_mall_product_summary_raw" in c_raw:
            del c_raw["shopping_mall_product_summary_raw"]

        if raw_sm_summary:
            items = []
            for sm in raw_sm_summary.split("|||"):
                if not sm: continue
                parts = sm.split("^^")
                if len(parts) == 9:
                    pname, dcode, sm_type, status, end_date, price, unit, path_avail, src = parts
                    # whitelist 기반 contract_type 필터링
                    if sm_type not in ALLOWED_SM_CONTRACT_TYPES:
                        sm_type = "unknown"
                    items.append({
                        "product_name": pname,
                        "detail_product_code": dcode if dcode else None,
                        "shopping_mall_contract_type": sm_type,
                        "contract_status": status,
                        "contract_end_date": end_date if end_date else None,
                        "price_amount": float(price) if price else None,
                        "price_unit": unit if unit else None,
                        "order_path_available": True if path_avail == '1' else False,
                        "source_name": src
                    })
            c_raw["shopping_mall_product_summary"] = items
            if items and "shopping_mall_product" not in c_raw.get("source_refs", []):
                c_raw.setdefault("source_refs", []).append("shopping_mall_product")
                
        if raw_mas_summary:
            items = []
            for mas in raw_mas_summary.split("|||"):
                if not mas: continue
                parts = mas.split("^^")
                if len(parts) == 7:
                    pname, dcode, status, end_date, price, unit, src = parts
                    # 뷰에서 이미 active만 제한하여 반환하므로 필터 로직 불필요
                    items.append({
                        "product_name": pname,
                        "detail_product_code": dcode if dcode else None,
                        "contract_status": status,
                        "contract_end_date": end_date if end_date else None,
                        "price_amount": float(price) if price else None,
                        "price_unit": unit if unit else None,
                        "source_name": src
                    })
            c_raw["mas_product_summary"] = items
            if items and "mas_product" not in c_raw["source_refs"]:
                c_raw["source_refs"].append("mas_product")

        # Phase 6-D-2: 업체/정책 속성 파싱
        raw_procurement_attrs = c_raw.get("procurement_attributes_raw")
        if "procurement_attributes_raw" in c_raw:
            del c_raw["procurement_attributes_raw"]
        c_raw["procurement_attributes"] = []
        if raw_procurement_attrs:
            c_raw["procurement_attributes"] = [a for a in raw_procurement_attrs.split("|") if a]

        # Phase 6-D-2: 일반 인증/기타 파싱
        raw_general_certs = c_raw.get("general_certifications_raw")
        if "general_certifications_raw" in c_raw:
            del c_raw["general_certifications_raw"]
        c_raw["general_certifications"] = []
        if raw_general_certs:
            c_raw["general_certifications"] = [g for g in raw_general_certs.split("|") if g]

        # Phase 6-D-2: source_refs 보강
        if c_raw["procurement_attributes"] and "company_procurement_attribute" not in c_raw.get("source_refs", []):
            c_raw.setdefault("source_refs", []).append("company_procurement_attribute")
        if c_raw["general_certifications"] and "product_general_certification" not in c_raw.get("source_refs", []):
            c_raw.setdefault("source_refs", []).append("product_general_certification")

        for ctype in c_raw.get("candidate_types", []):
            counts_by_type[ctype] = counts_by_type.get(ctype, 0) + 1

        for ref in c_raw.get("source_refs", []):
            t = c_raw.get("source_refreshed_at")
            if t:
                if ref not in latest_refresh or t > latest_refresh[ref]:
                    latest_refresh[ref] = t

        ALLOWED_CANDIDATE_FIELDS = [
            "company_id", "company_name", "representative_name", "corporate_phone",
            "location", "detail_address", "is_busan_company", "is_headquarters",
            "license_or_business_type", "main_products", "candidate_types",
            "primary_candidate_type", "manufacturer_type", "business_status",
            "business_status_freshness", "business_status_checked_at", "business_status_source", "display_status",
            "contract_possible_auto_promoted", "source_refs", "source_refreshed_at",
            "policy_subtypes", "policy_validity_summary",
            "certified_product_types", "certified_product_summary",
            "shopping_mall_flags", "shopping_mall_product_summary", "mas_product_summary", "sme_competition_product",
            "procurement_attributes", "general_certifications"
        ]

        filtered = {k: v for k, v in c_raw.items() if k in ALLOWED_CANDIDATE_FIELDS}

        if "actual_business_status" in c_raw and c_raw["actual_business_status"] is not None:
            filtered["business_status"] = c_raw["actual_business_status"]
        if "actual_business_status_freshness" in c_raw and c_raw["actual_business_status_freshness"] is not None:
            filtered["business_status_freshness"] = c_raw["actual_business_status_freshness"]

        candidates.append(filtered)

    final_meta = meta or {}
    final_meta["candidate_counts_by_type"] = counts_by_type
    if "source_refreshed_at" not in final_meta:
        final_meta["source_refreshed_at"] = latest_refresh

    return {
        "meta": final_meta,
        "candidates": candidates,
        "company_source_status": "success",
        "company_search_status": "success",
        "company_cache_used": True,
        "company_cache_mode": "database"
    }

ChatbotStatusFilter = Literal["exclude_closed", "all", "active_only", "needs_check"]
ChatbotValidityFilter = Literal["valid_only", "include_unknown", "all"]
ChatbotPolicySubtype = Literal[
    "women_company", "disabled_company", "social_enterprise",
    "preliminary_social_enterprise", "severe_disabled_production"
]
ChatbotCertificationType = Literal[
    "performance_certification", "excellent_procurement_product", "nep_product",
    "gs_certified_product", "net_certified_product", "innovation_product",
    "excellent_rnd_innovation_product", "innovation_prototype_product",
    "other_innovation_product", "disaster_safety_certified_product",
    "green_technology_product", "industrial_convergence_new_product",
    "excellent_procurement_joint_brand"
]

@app.get("/api/chatbot/health", tags=["챗봇"])
def get_chatbot_health():
    try:
        conn = _get_chatbot_db()
        cm_count = conn.execute("SELECT COUNT(*) FROM company_master").fetchone()[0]
        cp_count = conn.execute("SELECT COUNT(*) FROM company_product").fetchone()[0]
        cl_count = conn.execute("SELECT COUNT(*) FROM company_license").fetchone()[0]
        pol_count = conn.execute("SELECT COUNT(*) FROM policy_company_certification").fetchone()[0]
        mas_count = conn.execute("SELECT COUNT(*) FROM mas_product").fetchone()[0]
        attr_count = conn.execute("SELECT COUNT(*) FROM company_procurement_attribute").fetchone()[0]
        cert_count = conn.execute("SELECT COUNT(*) FROM product_general_certification").fetchone()[0]
        sm_count = conn.execute("SELECT COUNT(*) FROM shopping_mall_product").fetchone()[0]
        sm_active = conn.execute("SELECT COUNT(*) FROM shopping_mall_product WHERE contract_status = 'active'").fetchone()[0]
        conn.close()
        return {
            "status": "ok",
            "db": {
                "chatbot_db_connected": True,
                "company_master_count": cm_count,
                "company_product_count": cp_count,
                "company_license_count": cl_count,
                "policy_company_count": pol_count,
                "mas_product_count": mas_count,
                "procurement_attribute_count": attr_count,
                "general_certification_count": cert_count,
                "shopping_mall_product_count": sm_count,
                "active_shopping_mall_product_count": sm_active
            },
            "production_deployment": "HOLD"
        }
    except Exception as e:
        logger.exception("챗봇 API 오류: health")
        return {"status": "error", "production_deployment": "HOLD"}

@app.get("/api/chatbot/version", tags=["챗봇"])
def get_chatbot_version():
    return {
        "service": "busan-procurement-monitoring-chatbot-api",
        "api_version": "phase-6g-shopping-mall",
        "schema_version": "chatbot_company_v6g",
        "production_deployment": "HOLD",
        "features": [
            "company_search",
            "license_search",
            "product_search",
            "manufacturer_search",
            "policy_company",
            "certified_product",
            "mas_product",
            "sme_competition_product",
            "procurement_attributes",
            "general_certifications",
            "shopping_mall_product",
            "shopping_mall_contract_type",
            "product_policy_summary",
            "product_policy_search"
        ]
    }

@app.get("/api/chatbot/company/license-list", tags=["챗봇"])
def get_chatbot_license_list(limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT cl.license_name, COUNT(DISTINCT m.company_internal_id) as candidate_count
            FROM company_license cl
            JOIN company_master m ON cl.company_internal_id = m.company_internal_id
            LEFT JOIN company_business_status cbs ON m.company_internal_id = cbs.company_internal_id
            WHERE m.is_busan_company = 1
            {_get_status_filter_sql("exclude_closed")}
            GROUP BY cl.license_name
            ORDER BY candidate_count DESC
            LIMIT ? OFFSET ?
        '''
        rows = conn.execute(query, (limit, offset)).fetchall()
        conn.close()
        return {
            "meta": {},
            "candidates": [dict(r) for r in rows],
            "company_source_status": "success",
            "company_search_status": "success",
            "company_cache_used": True,
            "company_cache_mode": "database"
        }
    except Exception:
        logger.exception("챗봇 API 오류: license-list")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/product-list", tags=["챗봇"])
def get_chatbot_product_list(limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT cp.product_name, COUNT(DISTINCT m.company_internal_id) as candidate_count
            FROM company_product cp
            JOIN company_master m ON cp.company_internal_id = m.company_internal_id
            LEFT JOIN company_business_status cbs ON m.company_internal_id = cbs.company_internal_id
            WHERE m.is_busan_company = 1
            {_get_status_filter_sql("exclude_closed")}
            GROUP BY cp.product_name
            ORDER BY candidate_count DESC
            LIMIT ? OFFSET ?
        '''
        rows = conn.execute(query, (limit, offset)).fetchall()
        conn.close()
        return {
            "meta": {},
            "candidates": [dict(r) for r in rows],
            "company_source_status": "success",
            "company_search_status": "success",
            "company_cache_used": True,
            "company_cache_mode": "database"
        }
    except Exception:
        logger.exception("챗봇 API 오류: product-list")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/category-list", tags=["챗봇"])
def get_chatbot_category_list(limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT g.category_code, MAX(g.category_name) as category_name, COUNT(DISTINCT m.company_internal_id) as candidate_count
            FROM company_product cp
            JOIN g2b_product_category g ON cp.g2b_category_code = g.category_code
            JOIN company_master m ON cp.company_internal_id = m.company_internal_id
            LEFT JOIN company_business_status cbs ON m.company_internal_id = cbs.company_internal_id
            WHERE m.is_busan_company = 1
            {_get_status_filter_sql("exclude_closed")}
            GROUP BY g.category_code
            ORDER BY candidate_count DESC
            LIMIT ? OFFSET ?
        '''
        rows = conn.execute(query, (limit, offset)).fetchall()
        conn.close()
        return {
            "meta": {},
            "candidates": [dict(r) for r in rows],
            "company_source_status": "success",
            "company_search_status": "success",
            "company_cache_used": True,
            "company_cache_mode": "database"
        }
    except Exception:
        logger.exception("챗봇 API 오류: category-list")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/manufacturers", tags=["챗봇"])
def get_chatbot_manufacturers(limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE v.manufacturer_type != 'unknown' AND v.is_busan_company = 1
            {_get_status_filter_sql("exclude_closed")}
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        rows = conn.execute(query, (limit, offset)).fetchall()
        conn.close()
        return _build_chatbot_response(rows, meta={"query": {"limit": limit, "offset": offset}})
    except Exception:
        logger.exception("챗봇 API 오류: manufacturers")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/license-search", tags=["챗봇"])
def get_chatbot_license_search(license_name: str, status_filter: ChatbotStatusFilter = "exclude_closed", limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN company_license cl ON i.company_internal_id = cl.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE (cl.license_name LIKE ? OR cl.license_name_normalized LIKE ?) AND v.is_busan_company = 1
            {_get_status_filter_sql(status_filter)}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        p = f"%{license_name}%"
        rows = conn.execute(query, (p, p, limit, offset)).fetchall()
        conn.close()
        return _build_chatbot_response(rows, meta={"query": {"keyword": license_name, "limit": limit, "offset": offset, "status_filter": status_filter}})
    except Exception:
        logger.exception("챗봇 API 오류: license-search")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/product-search", tags=["챗봇"])
def get_chatbot_product_search(product_name: str, status_filter: ChatbotStatusFilter = "exclude_closed", limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN company_product cp ON i.company_internal_id = cp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE (cp.product_name LIKE ? OR cp.product_name_normalized LIKE ?) AND v.is_busan_company = 1
            {_get_status_filter_sql(status_filter)}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        p = f"%{product_name}%"
        rows = conn.execute(query, (p, p, limit, offset)).fetchall()
        conn.close()
        return _build_chatbot_response(rows, meta={"query": {"keyword": product_name, "limit": limit, "offset": offset, "status_filter": status_filter}})
    except Exception:
        logger.exception("챗봇 API 오류: product-search")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/category-search", tags=["챗봇"])
def get_chatbot_category_search(category_name: str, status_filter: ChatbotStatusFilter = "exclude_closed", limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        keyword_normalized = category_name.lower()
        for ch in [" ", "-", "_", "/", "(", ")", "[", "]", ".", ","]:
            keyword_normalized = keyword_normalized.replace(ch, "")
        query = f'''
            WITH matched_category AS (
                SELECT category_code
                FROM g2b_product_category
                WHERE category_name LIKE ?
                   OR category_name_normalized LIKE ?
                   OR category_code = ?
                UNION
                SELECT dtil_prdct_clsfc_no AS category_code
                FROM procurement_product_alias
                WHERE is_active = 1
                  AND (
                      alias LIKE ?
                      OR alias_normalized LIKE ?
                      OR canonical_name LIKE ?
                      OR canonical_name_normalized LIKE ?
                  )
                UNION
                SELECT prdct_clsfc_no AS category_code
                FROM procurement_product_alias
                WHERE is_active = 1
                  AND prdct_clsfc_no <> ''
                  AND (
                      alias LIKE ?
                      OR alias_normalized LIKE ?
                      OR canonical_name LIKE ?
                      OR canonical_name_normalized LIKE ?
                  )
            )
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN company_product cp ON i.company_internal_id = cp.company_internal_id
            JOIN matched_category mc ON (cp.g2b_category_code = mc.category_code OR cp.product_code = mc.category_code)
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE v.is_busan_company = 1
            {_get_status_filter_sql(status_filter)}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        p = f"%{category_name}%"
        pn = f"%{keyword_normalized}%"
        rows = conn.execute(
            query,
            (p, pn, category_name, p, pn, p, pn, p, pn, p, pn, limit, offset),
        ).fetchall()
        conn.close()
        return _build_chatbot_response(rows, meta={"query": {"keyword": category_name, "limit": limit, "offset": offset, "status_filter": status_filter}})
    except Exception:
        logger.exception("챗봇 API 오류: category-search")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/detail", tags=["챗봇"])
def get_chatbot_company_detail(company_id: str, request: Request):
    try:
        conn = _get_chatbot_db()
        row = conn.execute("SELECT company_internal_id FROM company_identity WHERE company_id = ?", (company_id,)).fetchone()
        if not row:
            conn.close()
            return _build_chatbot_response([], error="유효하지 않거나 만료된 업체 식별자입니다.")

        internal_id = row["company_internal_id"]

        cache_row = conn.execute("SELECT business_status, checked_at, business_status_source, business_status_freshness FROM company_business_status WHERE company_internal_id = ?", (internal_id,)).fetchone()

        now = datetime.datetime.now()
        should_fetch = False
        if not cache_row:
            should_fetch = True
        else:
            checked_at_str = cache_row["checked_at"]
            if not checked_at_str or cache_row["business_status"] in ("unknown", "api_failed"):
                should_fetch = True
            else:
                try:
                    checked_at = datetime.datetime.strptime(checked_at_str, "%Y-%m-%d %H:%M:%S")
                    if (now - checked_at).days >= 7:
                        should_fetch = True
                except Exception:
                    should_fetch = True

        if should_fetch:
            import nts_business_status_client
            b_row = conn.execute("SELECT canonical_business_no FROM company_identity WHERE company_internal_id = ?", (internal_id,)).fetchone()
            if b_row and b_row["canonical_business_no"]:
                b_no = b_row["canonical_business_no"]
                res = nts_business_status_client.check_business_status([b_no])
                now_str = now.strftime("%Y-%m-%d %H:%M:%S")
                if res.get("success") and res["results"].get(b_no):
                    r = res["results"][b_no]
                    conn.execute('''
                        INSERT INTO company_business_status
                        (company_internal_id, business_status, business_status_freshness, tax_type, closed_at, api_result_code, checked_at, business_status_source)
                        VALUES (?, ?, 'fresh', ?, ?, ?, ?, 'nts_api')
                        ON CONFLICT(company_internal_id) DO UPDATE SET
                            business_status=excluded.business_status,
                            business_status_freshness='fresh',
                            tax_type=excluded.tax_type,
                            closed_at=excluded.closed_at,
                            api_result_code=excluded.api_result_code,
                            checked_at=excluded.checked_at,
                            business_status_source='nts_api',
                            updated_at=CURRENT_TIMESTAMP
                    ''', (internal_id, r["business_status"], r.get("tax_type"), r.get("closed_at"), r.get("api_result_code"), now_str))
                    conn.commit()
                else:
                    conn.execute('''
                        INSERT INTO company_business_status
                        (company_internal_id, business_status, business_status_freshness, checked_at, business_status_source)
                        VALUES (?, 'unknown', 'api_failed', ?, 'nts_api')
                        ON CONFLICT(company_internal_id) DO UPDATE SET
                            business_status_freshness='api_failed',
                            checked_at=excluded.checked_at,
                            updated_at=CURRENT_TIMESTAMP
                    ''', (internal_id, now_str))
                    conn.commit()

        query = '''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE v.company_id = ?
        '''
        rows = conn.execute(query, (company_id,)).fetchall()
        conn.close()

        resp = _build_chatbot_response(rows, meta={"query": {"company_id": company_id}})
        if resp["candidates"]:
            resp["candidates"][0]["representative_name"] = None
            resp["candidates"][0]["corporate_phone"] = None

        return resp
    except Exception:
        logger.exception("챗봇 API 오류: company-detail")
        return _build_chatbot_response([], error="업체 목록 조회 실패")

@app.get("/api/chatbot/company/policy-search", tags=["챗봇"])
def get_chatbot_policy_search(policy_subtype: ChatbotPolicySubtype = None, status_filter: ChatbotStatusFilter = "exclude_closed", validity_filter: ChatbotValidityFilter = "valid_only", limit: int = Query(50, ge=1, le=50), offset: int = Query(0, ge=0)):
    try:
        conn = _get_chatbot_db()
        if policy_subtype:
            query = f'''
                SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
                FROM chatbot_company_candidate_view v
                JOIN company_identity i ON v.company_id = i.company_id
                JOIN policy_company_certification pcc ON i.company_internal_id = pcc.company_internal_id
                LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
                WHERE pcc.policy_subtype = ? AND pcc.validity_status = 'valid' AND v.is_busan_company = 1
                {_get_status_filter_sql(status_filter)}
                GROUP BY v.company_id
                ORDER BY v.company_id
                LIMIT ? OFFSET ?
            '''
            rows = conn.execute(query, (policy_subtype, limit, offset)).fetchall()
        else:
            query = f'''
                SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
                FROM chatbot_company_candidate_view v
                JOIN company_identity i ON v.company_id = i.company_id
                JOIN policy_company_certification pcc ON i.company_internal_id = pcc.company_internal_id
                LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
                WHERE pcc.validity_status = 'valid' AND v.is_busan_company = 1
                {_get_status_filter_sql(status_filter)}
                GROUP BY v.company_id
                ORDER BY v.company_id
                LIMIT ? OFFSET ?
            '''
            rows = conn.execute(query, (limit, offset)).fetchall()

        conn.close()
        resp = _build_chatbot_response(rows, meta={"query": {"policy_subtype": policy_subtype, "limit": limit, "offset": offset}}, validity_filter=validity_filter)
        for c in resp["candidates"]:
            c["primary_candidate_type"] = "policy_company"
        return resp
    except Exception:
        logger.exception("챗봇 API 오류: policy-search")
        return _build_chatbot_response([], error="정책기업 조회 실패")

@app.get("/api/chatbot/company/policy-list", tags=["챗봇"])
def get_chatbot_policy_list():
    try:
        conn = _get_chatbot_db()
        query = '''
            SELECT
                pcc.policy_subtype,
                COUNT(DISTINCT pcc.company_internal_id) as candidate_count,
                SUM(CASE WHEN pcc.validity_status = 'valid' THEN 1 ELSE 0 END) as valid_count,
                SUM(CASE WHEN pcc.validity_status = 'expired' THEN 1 ELSE 0 END) as expired_count,
                MAX(pcc.source_refreshed_at) as refreshed_at
            FROM policy_company_certification pcc
            JOIN company_master m ON pcc.company_internal_id = m.company_internal_id
            WHERE m.is_busan_company = 1
            GROUP BY pcc.policy_subtype
        '''
        rows = conn.execute(query).fetchall()
        conn.close()

        candidates = []
        latest = None
        for r in rows:
            candidates.append({
                "policy_subtype": r["policy_subtype"],
                "candidate_count": r["candidate_count"],
                "valid_count": r["valid_count"],
                "expired_count": r["expired_count"]
            })
            if r["refreshed_at"]:
                if latest is None or r["refreshed_at"] > latest:
                    latest = r["refreshed_at"]

        return {
            "meta": {
                "source_refreshed_at": {"policy_company_certification": latest} if latest else {}
            },
            "candidates": candidates,
            "company_source_status": "success",
            "company_search_status": "success",
            "company_cache_used": False,
            "company_cache_mode": "none"
        }
    except Exception:
        logger.exception("챗봇 API 오류: policy-list")
        return _build_chatbot_response([], error="정책기업 조회 실패")

# ==========================================
# Phase 5: 인증제품 연동 API
# ==========================================

@app.get("/api/chatbot/product/certified-search", tags=["챗봇"])
def get_chatbot_certified_search(
    certification_type: Optional[ChatbotCertificationType] = None,
    product_name: Optional[str] = None,
    company_keyword: Optional[str] = None,
    status_filter: ChatbotStatusFilter = "exclude_closed",
    validity_filter: ChatbotValidityFilter = "valid_only",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    try:
        conn = _get_chatbot_db()
        where_clauses = ["v.is_busan_company = 1"]
        params = []
        
        if certification_type:
            where_clauses.append("cp.certification_type = ?")
            params.append(certification_type)
        if product_name:
            where_clauses.append("(cp.product_name LIKE ? OR cp.product_name_normalized LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%"])
        if company_keyword:
            where_clauses.append("(v.company_name LIKE ? OR v.company_id = ?)")
            params.extend([f"%{company_keyword}%", company_keyword])
            
        if validity_filter == "valid_only":
            where_clauses.append("cp.validity_status = 'valid'")
        elif validity_filter == "include_unknown":
            where_clauses.append("cp.validity_status IN ('valid', 'unknown')")
            
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join(where_clauses)
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN certified_product cp ON i.company_internal_id = cp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        return _build_chatbot_response(rows, meta={
            "query": {
                "certification_type": certification_type,
                "product_name": product_name,
                "company_keyword": company_keyword,
                "limit": limit,
                "offset": offset
            }
        }, validity_filter=validity_filter)
    except Exception:
        logger.exception("챗봇 API 오류: certified-search")
        return _build_chatbot_response([], error="인증제품 조회 실패")

@app.get("/api/chatbot/product/innovation-search", tags=["챗봇"])
def get_chatbot_innovation_search(
    product_name: Optional[str] = None,
    innovation_type: Literal["all", "excellent_rnd", "prototype", "other"] = "all",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    validity_filter: ChatbotValidityFilter = "valid_only",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    try:
        conn = _get_chatbot_db()
        where_clauses = ["v.is_busan_company = 1"]
        params = []
        
        type_in = []
        if innovation_type == "all":
            type_in = ["innovation_product", "excellent_rnd_innovation_product", "innovation_prototype_product", "other_innovation_product"]
        elif innovation_type == "excellent_rnd":
            type_in = ["excellent_rnd_innovation_product"]
        elif innovation_type == "prototype":
            type_in = ["innovation_prototype_product"]
        else:
            type_in = ["other_innovation_product"]
            
        placeholders = ",".join(["?"] * len(type_in))
        where_clauses.append(f"cp.certification_type IN ({placeholders})")
        params.extend(type_in)
        
        if product_name:
            where_clauses.append("(cp.product_name LIKE ? OR cp.product_name_normalized LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%"])
            
        if validity_filter == "valid_only":
            where_clauses.append("cp.validity_status = 'valid'")
        elif validity_filter == "include_unknown":
            where_clauses.append("cp.validity_status IN ('valid', 'unknown')")
            
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join(where_clauses)
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN certified_product cp ON i.company_internal_id = cp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        return _build_chatbot_response(rows, meta={
            "query": {
                "product_name": product_name,
                "innovation_type": innovation_type,
                "limit": limit,
                "offset": offset
            }
        }, validity_filter=validity_filter)
    except Exception:
        logger.exception("챗봇 API 오류: innovation-search")
        return _build_chatbot_response([], error="혁신제품 조회 실패")

@app.get("/api/chatbot/product/priority-purchase-search", tags=["챗봇"])
def get_chatbot_priority_purchase_search(
    product_name: Optional[str] = None,
    certification_type: Optional[ChatbotCertificationType] = None,
    status_filter: ChatbotStatusFilter = "exclude_closed",
    validity_filter: ChatbotValidityFilter = "valid_only",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    try:
        conn = _get_chatbot_db()
        where_clauses = ["v.is_busan_company = 1", "map.is_priority_purchase_product = 1"]
        params = []
        
        if certification_type:
            where_clauses.append("cp.certification_type = ?")
            params.append(certification_type)
        if product_name:
            where_clauses.append("(cp.product_name LIKE ? OR cp.product_name_normalized LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%"])
            
        if validity_filter == "valid_only":
            where_clauses.append("cp.validity_status = 'valid'")
        elif validity_filter == "include_unknown":
            where_clauses.append("cp.validity_status IN ('valid', 'unknown')")
            
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join(where_clauses)
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN certified_product cp ON i.company_internal_id = cp.company_internal_id
            JOIN certified_product_type_map map ON cp.certification_type = map.normalized_certification_type
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        return _build_chatbot_response(rows, meta={
            "query": {
                "product_name": product_name,
                "certification_type": certification_type,
                "limit": limit,
                "offset": offset
            }
        }, validity_filter=validity_filter)
    except Exception:
        logger.exception("챗봇 API 오류: priority-purchase-search")
        return _build_chatbot_response([], error="기술개발제품(우선구매) 조회 실패")

@app.get("/api/chatbot/product/excellent-procurement-search", tags=["챗봇"])
def get_chatbot_excellent_procurement_search(
    product_name: Optional[str] = None,
    status_filter: ChatbotStatusFilter = "exclude_closed",
    validity_filter: ChatbotValidityFilter = "valid_only",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    try:
        conn = _get_chatbot_db()
        where_clauses = ["v.is_busan_company = 1", "cp.certification_type = 'excellent_procurement_product'"]
        params = []
        
        if product_name:
            where_clauses.append("(cp.product_name LIKE ? OR cp.product_name_normalized LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%"])
            
        if validity_filter == "valid_only":
            where_clauses.append("cp.validity_status = 'valid'")
        elif validity_filter == "include_unknown":
            where_clauses.append("cp.validity_status IN ('valid', 'unknown')")
            
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join(where_clauses)
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN certified_product cp ON i.company_internal_id = cp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        return _build_chatbot_response(rows, meta={
            "query": {
                "product_name": product_name,
                "limit": limit,
                "offset": offset
            }
        }, validity_filter=validity_filter)
    except Exception:
        logger.exception("챗봇 API 오류: excellent-procurement-search")
        return _build_chatbot_response([], error="우수조달물품 조회 실패")

@app.get("/api/chatbot/product/certified-list", tags=["챗봇"])
def get_chatbot_certified_list():
    try:
        conn = _get_chatbot_db()
        query = '''
            SELECT
                cp.certification_type,
                COUNT(DISTINCT cp.company_internal_id) as candidate_count,
                SUM(CASE WHEN cp.validity_status = 'valid' THEN 1 ELSE 0 END) as valid_count,
                SUM(CASE WHEN cp.validity_status = 'expired' THEN 1 ELSE 0 END) as expired_count,
                MAX(cp.source_refreshed_at) as refreshed_at
            FROM certified_product cp
            JOIN company_master m ON cp.company_internal_id = m.company_internal_id
            WHERE m.is_busan_company = 1
            GROUP BY cp.certification_type
        '''
        rows = conn.execute(query).fetchall()
        conn.close()

        candidates = []
        latest = None
        for r in rows:
            candidates.append({
                "certification_type": r["certification_type"],
                "candidate_count": r["candidate_count"],
                "valid_count": r["valid_count"],
                "expired_count": r["expired_count"]
            })
            if r["refreshed_at"]:
                if latest is None or r["refreshed_at"] > latest:
                    latest = r["refreshed_at"]

        return {
            "meta": {
                "source_refreshed_at": {"certified_product": latest} if latest else {}
            },
            "candidates": candidates,
            "company_source_status": "success",
            "company_search_status": "success",
            "company_cache_used": False,
            "company_cache_mode": "none",
            "error": None
        }
    except Exception:
        logger.exception("챗봇 API 오류: certified-list")
        return _build_chatbot_response([], error="인증제품 목록 조회 실패")

# ==========================================
# Phase 6-C: MAS 쇼핑몰 연동 API
# ==========================================

MasContractStatusFilter = Literal["active_only", "include_unknown", "all"]
ShoppingMallContractTypeFilter = Literal["all", "mas", "third_party_unit_price", "general_unit_price", "excellent_procurement", "unknown"]

def _get_mas_status_filter_sql(status_filter: str) -> str:
    if status_filter == "all":
        return ""
    elif status_filter == "include_unknown":
        return " AND mp.contract_status IN ('active', 'unknown') "
    return " AND mp.contract_status = 'active' "

@app.get("/api/chatbot/mas/search", tags=["챗봇"])
def get_chatbot_mas_search(
    product_name: Optional[str] = None,
    detail_product_code: Optional[str] = None,
    company_keyword: Optional[str] = None,
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    MAS 종합 검색. 
    주의: active_only는 종합쇼핑몰에서의 당장 구매 가능을 의미하지 않습니다.
    """
    try:
        conn = _get_chatbot_db()
        where_clauses = ["v.is_busan_company = 1"]
        params = []
        
        if product_name:
            where_clauses.append("(mp.product_name LIKE ? OR mp.product_name_normalized LIKE ? OR mp.detail_product_name LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%", f"%{product_name}%"])
        if detail_product_code:
            where_clauses.append("mp.detail_product_code = ?")
            params.append(detail_product_code)
        if company_keyword:
            where_clauses.append("(v.company_name LIKE ? OR v.company_id = ?)")
            params.extend([f"%{company_keyword}%", company_keyword])
            
        where_clauses.append(_get_mas_status_filter_sql(contract_status_filter).strip(" AND"))
        
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join([w for w in where_clauses if w])
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN mas_product mp ON i.company_internal_id = mp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        # Note: 뷰가 반환하는 mas_product_summary는 active만 포함하지만,
        # 이 MAS 전용 API에서는 뷰 결과를 그대로 제공하되, 향후 필요시 직접 조인 결과로 오버라이드할 수 있습니다.
        # 현재 요구사항인 "include_unknown/all에서는 mas_product_summary에만 표시할 수 있다"에 따라,
        # mas_product_summary를 다시 쿼리하여 오버라이드합니다.
        
        resp = _build_chatbot_response(rows, meta={
            "query": {
                "product_name": product_name,
                "detail_product_code": detail_product_code,
                "company_keyword": company_keyword,
                "contract_status_filter": contract_status_filter,
                "limit": limit,
                "offset": offset
            }
        })
        
        if contract_status_filter != "active_only" and resp["candidates"]:
            # Re-fetch mas products for the candidates to include non-active ones
            cids = [c["company_id"] for c in resp["candidates"]]
            conn = _get_chatbot_db()
            placeholders = ",".join("?" * len(cids))
            
            mas_q = f'''
                SELECT i.company_id, mp.product_name, mp.detail_product_code, mp.contract_status, mp.contract_end_date, mp.price_amount, mp.price_unit, mp.source_name
                FROM mas_product mp
                JOIN company_identity i ON mp.company_internal_id = i.company_internal_id
                WHERE i.company_id IN ({placeholders}) {_get_mas_status_filter_sql(contract_status_filter)}
                ORDER BY mp.contract_end_date DESC
            '''
            mas_rows = conn.execute(mas_q, cids).fetchall()
            conn.close()
            
            mas_map = {}
            for r in mas_rows:
                cid = r["company_id"]
                if cid not in mas_map:
                    mas_map[cid] = []
                mas_map[cid].append({
                    "product_name": r["product_name"],
                    "detail_product_code": r["detail_product_code"],
                    "contract_status": r["contract_status"],
                    "contract_end_date": r["contract_end_date"],
                    "price_amount": float(r["price_amount"]) if r["price_amount"] is not None else None,
                    "price_unit": r["price_unit"],
                    "source_name": r["source_name"]
                })
            
            for c in resp["candidates"]:
                c["mas_product_summary"] = mas_map.get(c["company_id"], [])[:5] # Limit to 5
                
        return resp
    except Exception:
        logger.exception("챗봇 API 오류: mas-search")
        return _build_chatbot_response([], error="MAS 제품 조회 실패")


@app.get("/api/chatbot/mas/product-search", tags=["챗봇"])
def get_chatbot_mas_product_search(
    product_name: str,
    detail_product_code: Optional[str] = None,
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    MAS 제품 검색.
    """
    return get_chatbot_mas_search(
        product_name=product_name,
        detail_product_code=detail_product_code,
        contract_status_filter=contract_status_filter,
        status_filter=status_filter,
        limit=limit,
        offset=offset
    )


@app.get("/api/chatbot/mas/supplier-search", tags=["챗봇"])
def get_chatbot_mas_supplier_search(
    company_keyword: Optional[str] = None,
    is_busan_company: bool = True,
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    MAS 공급업체 검색.
    """
    try:
        conn = _get_chatbot_db()
        where_clauses = []
        if is_busan_company:
            where_clauses.append("v.is_busan_company = 1")
            
        params = []
        if company_keyword:
            where_clauses.append("(v.company_name LIKE ? OR v.company_id = ?)")
            params.extend([f"%{company_keyword}%", company_keyword])
            
        where_clauses.append(_get_mas_status_filter_sql(contract_status_filter).strip(" AND"))
        
        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join([w for w in where_clauses if w])
        
        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN mas_product mp ON i.company_internal_id = mp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            WHERE {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()
        
        resp = _build_chatbot_response(rows, meta={
            "query": {
                "company_keyword": company_keyword,
                "is_busan_company": is_busan_company,
                "contract_status_filter": contract_status_filter,
                "limit": limit,
                "offset": offset
            }
        })
        return resp
    except Exception:
        logger.exception("챗봇 API 오류: mas-supplier-search")
        return _build_chatbot_response([], error="MAS 공급업체 조회 실패")


@app.get("/api/chatbot/mas/list", tags=["챗봇"])
def get_chatbot_mas_list():
    """
    MAS 통계/집계 목록.
    """
    try:
        conn = _get_chatbot_db()
        query = '''
            SELECT
                mp.detail_product_code,
                MAX(mp.product_name) as product_name,
                COUNT(DISTINCT mp.company_internal_id) as supplier_count,
                SUM(CASE WHEN mp.contract_status = 'active' THEN 1 ELSE 0 END) as active_contract_count,
                SUM(CASE WHEN mp.contract_status = 'expired' THEN 1 ELSE 0 END) as expired_contract_count,
                MAX(mp.source_refreshed_at) as refreshed_at
            FROM mas_product mp
            JOIN company_master m ON mp.company_internal_id = m.company_internal_id
            WHERE m.is_busan_company = 1
            GROUP BY mp.detail_product_code
        '''
        rows = conn.execute(query).fetchall()
        conn.close()

        candidates = []
        latest = None
        for r in rows:
            candidates.append({
                "detail_product_code": r["detail_product_code"],
                "product_name": r["product_name"],
                "supplier_count": r["supplier_count"],
                "active_contract_count": r["active_contract_count"],
                "expired_contract_count": r["expired_contract_count"]
            })
            if r["refreshed_at"]:
                if latest is None or r["refreshed_at"] > latest:
                    latest = r["refreshed_at"]

        return {
            "meta": {
                "source_refreshed_at": {"mas_product": latest} if latest else {}
            },
            "candidates": candidates,
            "company_source_status": "live_company_lookup",
            "company_search_status": "success",
            "company_cache_used": False,
            "company_cache_mode": "live_only",
            "error": None
        }
    except Exception:
        logger.exception("챗봇 API 오류: mas-list")
        return _build_chatbot_response([], error="MAS 목록 조회 실패")

# ==========================================
# Phase 6-G: 종합쇼핑몰 전용 API
# ==========================================

def _get_sm_type_filter_sql(contract_type_filter: str) -> str:
    if contract_type_filter == "all":
        return ""
    return f" AND smp.shopping_mall_contract_type = '{contract_type_filter}' "

def _get_sm_status_filter_sql(status_filter: str) -> str:
    if status_filter == "all":
        return ""
    elif status_filter == "include_unknown":
        return " AND smp.contract_status IN ('active', 'unknown') "
    return " AND smp.contract_status = 'active' "

@app.get("/api/chatbot/shopping-mall/search", tags=["챗봇 종합쇼핑몰"])
def get_chatbot_sm_search(
    product_name: Optional[str] = None,
    detail_product_code: Optional[str] = None,
    company_keyword: Optional[str] = None,
    contract_type_filter: ShoppingMallContractTypeFilter = "all",
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    is_busan_company: bool = True,
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    종합쇼핑몰 등록상품 종합 검색.
    shopping_mall_product 기준 조회. contract_type_filter로 MAS/3자단가/일반단가/우수제품 구분.
    주의: 계약상태 active는 전자구매 경로 확인 용도이며, 당장 구매 가능을 의미하지 않습니다.
    """
    try:
        conn = _get_chatbot_db()
        where_clauses = []
        if is_busan_company:
            where_clauses.append("v.is_busan_company = 1")
        params = []

        if product_name:
            where_clauses.append("(smp.product_name LIKE ? OR smp.product_name_normalized LIKE ? OR smp.detail_product_name LIKE ?)")
            params.extend([f"%{product_name}%", f"%{product_name}%", f"%{product_name}%"])
        if detail_product_code:
            where_clauses.append("smp.detail_product_code = ?")
            params.append(detail_product_code)
        if company_keyword:
            where_clauses.append("(v.company_name LIKE ? OR v.company_id = ?)")
            params.extend([f"%{company_keyword}%", company_keyword])

        sm_type_sql = _get_sm_type_filter_sql(contract_type_filter)
        sm_status_sql = _get_sm_status_filter_sql(contract_status_filter)
        if sm_type_sql:
            where_clauses.append(sm_type_sql.strip(" AND"))
        if sm_status_sql:
            where_clauses.append(sm_status_sql.strip(" AND"))

        status_sql = _get_status_filter_sql(status_filter)
        where_sql = " AND ".join([w for w in where_clauses if w])
        if where_sql:
            where_sql = "WHERE " + where_sql

        query = f'''
            SELECT v.*, cbs.business_status as actual_business_status, cbs.business_status_freshness as actual_business_status_freshness, cbs.checked_at as business_status_checked_at, cbs.business_status_source
            FROM chatbot_company_candidate_view v
            JOIN company_identity i ON v.company_id = i.company_id
            JOIN shopping_mall_product smp ON i.company_internal_id = smp.company_internal_id
            LEFT JOIN company_business_status cbs ON i.company_internal_id = cbs.company_internal_id
            {where_sql} {status_sql}
            GROUP BY v.company_id
            ORDER BY v.company_id
            LIMIT ? OFFSET ?
        '''
        params.extend([limit, offset])
        rows = conn.execute(query, params).fetchall()
        conn.close()

        resp = _build_chatbot_response(rows, meta={
            "query": {
                "product_name": product_name,
                "detail_product_code": detail_product_code,
                "company_keyword": company_keyword,
                "contract_type_filter": contract_type_filter,
                "contract_status_filter": contract_status_filter,
                "limit": limit,
                "offset": offset
            }
        })

        # contract_status_filter가 active_only가 아닐 때 shopping_mall_product_summary를 오버라이드
        if contract_status_filter != "active_only" and resp["candidates"]:
            cids = [c["company_id"] for c in resp["candidates"]]
            conn = _get_chatbot_db()
            placeholders = ",".join("?" * len(cids))
            sm_q = f'''
                SELECT i.company_id, smp.product_name, smp.detail_product_code, smp.shopping_mall_contract_type,
                       smp.contract_status, smp.contract_end_date, smp.price_amount, smp.price_unit,
                       smp.order_path_available, smp.source_name
                FROM shopping_mall_product smp
                JOIN company_identity i ON smp.company_internal_id = i.company_internal_id
                WHERE i.company_id IN ({placeholders}) {_get_sm_type_filter_sql(contract_type_filter)} {_get_sm_status_filter_sql(contract_status_filter)}
                ORDER BY smp.contract_end_date DESC
            '''
            sm_rows = conn.execute(sm_q, cids).fetchall()
            conn.close()

            sm_map = {}
            for r in sm_rows:
                cid = r["company_id"]
                if cid not in sm_map:
                    sm_map[cid] = []
                sm_map[cid].append({
                    "product_name": r["product_name"],
                    "detail_product_code": r["detail_product_code"],
                    "shopping_mall_contract_type": r["shopping_mall_contract_type"] if r["shopping_mall_contract_type"] in ("mas", "third_party_unit_price", "general_unit_price", "excellent_procurement", "unknown") else "unknown",
                    "contract_status": r["contract_status"],
                    "contract_end_date": r["contract_end_date"],
                    "price_amount": float(r["price_amount"]) if r["price_amount"] is not None else None,
                    "price_unit": r["price_unit"],
                    "order_path_available": True if r["order_path_available"] else False,
                    "source_name": r["source_name"]
                })

            for c in resp["candidates"]:
                c["shopping_mall_product_summary"] = sm_map.get(c["company_id"], [])[:5]

        return resp
    except Exception:
        logger.exception("챗봇 API 오류: shopping-mall-search")
        return _build_chatbot_response([], error="종합쇼핑몰 제품 조회 실패")


@app.get("/api/chatbot/shopping-mall/product-search", tags=["챗봇 종합쇼핑몰"])
def get_chatbot_sm_product_search(
    product_name: str,
    detail_product_code: Optional[str] = None,
    contract_type_filter: ShoppingMallContractTypeFilter = "all",
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    종합쇼핑몰 등록상품 제품명 검색.
    """
    return get_chatbot_sm_search(
        product_name=product_name,
        detail_product_code=detail_product_code,
        contract_type_filter=contract_type_filter,
        contract_status_filter=contract_status_filter,
        status_filter=status_filter,
        limit=limit,
        offset=offset
    )


@app.get("/api/chatbot/shopping-mall/supplier-search", tags=["챗봇 종합쇼핑몰"])
def get_chatbot_sm_supplier_search(
    company_keyword: Optional[str] = None,
    is_busan_company: bool = True,
    contract_type_filter: ShoppingMallContractTypeFilter = "all",
    contract_status_filter: MasContractStatusFilter = "active_only",
    status_filter: ChatbotStatusFilter = "exclude_closed",
    limit: int = Query(20, ge=1, le=50),
    offset: int = Query(0, ge=0)
):
    """
    종합쇼핑몰 공급업체 검색.
    """
    return get_chatbot_sm_search(
        company_keyword=company_keyword,
        is_busan_company=is_busan_company,
        contract_type_filter=contract_type_filter,
        contract_status_filter=contract_status_filter,
        status_filter=status_filter,
        limit=limit,
        offset=offset
    )


@app.get("/api/chatbot/shopping-mall/list", tags=["챗봇 종합쇼핑몰"])
def get_chatbot_sm_list(
    contract_type_filter: ShoppingMallContractTypeFilter = "all"
):
    """
    종합쇼핑몰 통계/집계 목록 (계약유형별 세부품명 단위).
    """
    try:
        conn = _get_chatbot_db()
        type_sql = _get_sm_type_filter_sql(contract_type_filter)
        query = f'''
            SELECT
                smp.detail_product_code,
                MAX(smp.product_name) as product_name,
                smp.shopping_mall_contract_type,
                COUNT(DISTINCT smp.company_internal_id) as supplier_count,
                SUM(CASE WHEN smp.contract_status = 'active' THEN 1 ELSE 0 END) as active_contract_count,
                SUM(CASE WHEN smp.contract_status = 'expired' THEN 1 ELSE 0 END) as expired_contract_count,
                MAX(smp.source_refreshed_at) as refreshed_at
            FROM shopping_mall_product smp
            JOIN company_master m ON smp.company_internal_id = m.company_internal_id
            WHERE m.is_busan_company = 1 {type_sql}
            GROUP BY smp.detail_product_code, smp.shopping_mall_contract_type
        '''
        rows = conn.execute(query).fetchall()
        conn.close()

        candidates = []
        latest = None
        for r in rows:
            sm_type = r["shopping_mall_contract_type"]
            if sm_type not in ("mas", "third_party_unit_price", "general_unit_price", "excellent_procurement", "unknown"):
                sm_type = "unknown"
            candidates.append({
                "detail_product_code": r["detail_product_code"],
                "product_name": r["product_name"],
                "shopping_mall_contract_type": sm_type,
                "supplier_count": r["supplier_count"],
                "active_contract_count": r["active_contract_count"],
                "expired_contract_count": r["expired_contract_count"]
            })
            if r["refreshed_at"]:
                if latest is None or r["refreshed_at"] > latest:
                    latest = r["refreshed_at"]

        return {
            "meta": {
                "source_refreshed_at": {"shopping_mall_product": latest} if latest else {},
                "contract_type_filter": contract_type_filter
            },
            "candidates": candidates,
            "company_source_status": "live_company_lookup",
            "company_search_status": "success",
            "company_cache_used": False,
            "company_cache_mode": "live_only",
            "error": None
        }
    except Exception:
        logger.exception("챗봇 API 오류: shopping-mall-list")
        return _build_chatbot_response([], error="종합쇼핑몰 목록 조회 실패")





# ============================================================================
# 엑셀 다운로드 API (구군청 담당자용)
# ============================================================================


ProductPolicyFlagFilter = Literal["all", "sme_competition", "construction_material", "busan_coop", "direct_production", "mas"]

def _split_pipe(value):
    if value is None:
        return []
    return [item.strip() for item in str(value).split("|") if item and item.strip()]

def _product_policy_flag_sql(flag_filter: str) -> str:
    if flag_filter == "sme_competition":
        return " AND IFNULL(is_sme_competition_product, 0) = 1"
    if flag_filter == "construction_material":
        return " AND IFNULL(is_construction_material_direct_purchase, 0) = 1"
    if flag_filter == "busan_coop":
        return " AND (IFNULL(busan_eligible_coop_count, 0) > 0 OR IFNULL(busan_coop_joint_product_count, 0) > 0)"
    if flag_filter == "direct_production":
        return " AND IFNULL(direct_production_valid_supplier_count, 0) > 0"
    if flag_filter == "mas":
        return " AND IFNULL(mas_active_supplier_count, 0) > 0"
    return ""

def _build_product_policy_response(rows, meta=None, error=None):
    candidates = []
    for r in rows:
        item = dict(r)
        item["busan_eligible_coops"] = _split_pipe(item.get("busan_eligible_coops"))
        item["busan_coop_joint_product_coops"] = _split_pipe(item.get("busan_coop_joint_product_coops"))
        item["source_refs"] = _split_pipe(item.get("source_refs"))
        for key in (
            "is_sme_competition_product",
            "is_construction_material_direct_purchase",
            "eligible_coop_count",
            "busan_eligible_coop_count",
            "coop_joint_product_count",
            "busan_coop_joint_product_count",
            "mas_active_supplier_count",
            "shopping_mall_active_supplier_count",
            "busan_company_product_count",
            "direct_production_valid_supplier_count",
        ):
            item[key] = int(item.get(key) or 0)
        candidates.append(item)
    return {
        "meta": meta or {},
        "candidates": candidates,
        "company_source_status": "success" if error is None else "error",
        "company_search_status": "success" if error is None else "error",
        "company_cache_used": True,
        "company_cache_mode": "database",
        "error": error,
    }

@app.get("/api/chatbot/product-policy/search", tags=["chatbot product policy"])
def get_chatbot_product_policy_search(
    keyword: Optional[str] = Query(None, description="detail product name or code keyword"),
    detail_product_code: Optional[str] = Query(None, description="detail product code"),
    flag_filter: ProductPolicyFlagFilter = Query("all", description="policy flag filter"),
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    try:
        conn = _get_chatbot_db()
        where = ["1=1"]
        params = []
        if detail_product_code:
            where.append("detail_product_code = ?")
            params.append(detail_product_code.strip())
        if keyword:
            like = f"%{keyword.strip()}%"
            where.append("(detail_product_name LIKE ? OR detail_product_code LIKE ?)")
            params.extend([like, like])
        flag_sql = _product_policy_flag_sql(flag_filter)
        query = f'''
            SELECT
                detail_product_code,
                detail_product_name,
                is_sme_competition_product,
                is_construction_material_direct_purchase,
                required_special_note,
                construction_material_note,
                eligible_coop_count,
                busan_eligible_coop_count,
                busan_eligible_coops,
                coop_joint_product_count,
                busan_coop_joint_product_count,
                busan_coop_joint_product_coops,
                mas_active_supplier_count,
                shopping_mall_active_supplier_count,
                busan_company_product_count,
                direct_production_valid_supplier_count,
                source_refs,
                generated_at
            FROM product_policy_summary
            WHERE {' AND '.join(where)} {flag_sql}
            ORDER BY
                (IFNULL(is_sme_competition_product, 0) + IFNULL(is_construction_material_direct_purchase, 0)) DESC,
                (IFNULL(busan_eligible_coop_count, 0) + IFNULL(busan_coop_joint_product_count, 0)) DESC,
                IFNULL(direct_production_valid_supplier_count, 0) DESC,
                IFNULL(mas_active_supplier_count, 0) DESC,
                detail_product_name
            LIMIT ? OFFSET ?
        '''
        rows = conn.execute(query, params + [limit, offset]).fetchall()
        total_query = f"SELECT COUNT(*) FROM product_policy_summary WHERE {' AND '.join(where)} {flag_sql}"
        total = conn.execute(total_query, params).fetchone()[0]
        latest = conn.execute("SELECT MAX(generated_at) FROM product_policy_summary").fetchone()[0]
        conn.close()
        return _build_product_policy_response(rows, meta={
            "total": total,
            "limit": limit,
            "offset": offset,
            "keyword": keyword,
            "detail_product_code": detail_product_code,
            "flag_filter": flag_filter,
            "source_refreshed_at": {"product_policy_summary": latest} if latest else {},
        })
    except Exception:
        logger.exception("chatbot API error: product-policy-search")
        return _build_product_policy_response([], error="product policy search failed")

@app.get("/api/chatbot/product-policy/list", tags=["chatbot product policy"])
def get_chatbot_product_policy_list(
    flag_filter: ProductPolicyFlagFilter = Query("all", description="policy flag filter"),
    limit: int = Query(50, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    return get_chatbot_product_policy_search(
        keyword=None,
        detail_product_code=None,
        flag_filter=flag_filter,
        limit=limit,
        offset=offset,
    )

DownloadStatusFilter = Literal["active_only", "all", "exclude_closed"]

def _status_join_and_filter(status_filter: str):
    join = " LEFT JOIN company_business_status cbs ON cm.company_internal_id = cbs.company_internal_id"
    if status_filter == "active_only":
        where = " AND IFNULL(cbs.business_status, 'unknown') = 'active'"
    elif status_filter == "exclude_closed":
        where = " AND IFNULL(cbs.business_status, 'unknown') NOT IN ('closed', 'suspended')"
    else:
        where = ""
    col = ", IFNULL(cbs.business_status, 'unknown') as nts_status"
    return join, where, col


def _direct_production_summary_expr(company_alias: str = "cm", product_expr: Optional[str] = None, valid_only: bool = False) -> str:
    """직접생산확인증명 요약 SQL 표현식.

    다운로드 API 전용이며, 사업자번호 등 원천 식별자는 노출하지 않는다.
    product_expr가 있으면 세부품명과 다운로드 행의 품목명을 느슨하게 교차 매칭한다.
    """
    conditions = [f"dpc.company_internal_id = {company_alias}.company_internal_id"]
    if valid_only:
        conditions.append("dpc.validity_status = 'valid'")
    if product_expr:
        conditions.append(
            f"""(
                dpc.detail_product_name LIKE '%' || {product_expr} || '%'
                OR {product_expr} LIKE '%' || dpc.detail_product_name || '%'
            )"""
        )
    where_clause = " AND ".join(conditions)
    return f"""IFNULL((
        SELECT GROUP_CONCAT(
            dpc.detail_product_name || ':' ||
            dpc.validity_status || ':' ||
            IFNULL(dpc.valid_to, ''),
            ', '
        )
        FROM direct_production_certificate dpc
        WHERE {where_clause}
    ), '')"""


def _direct_production_count_expr(company_alias: str = "cm", valid_only: bool = True) -> str:
    where_clause = f"dpc.company_internal_id = {company_alias}.company_internal_id"
    if valid_only:
        where_clause += " AND dpc.validity_status = 'valid'"
    return f"""(
        SELECT COUNT(*)
        FROM direct_production_certificate dpc
        WHERE {where_clause}
    )"""


def _make_excel_response(rows, columns, headers_kr, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    from starlette.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "데이터"
    
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    for col_idx, h in enumerate(headers_kr, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    data_font = Font(name="맑은 고딕", size=10)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row[col_name] if isinstance(row, dict) else row[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
    
    for col_idx, h in enumerate(headers_kr, 1):
        max_len = len(str(h))
        for row_idx in range(2, min(len(rows) + 2, 100)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 40)
    
    ws.auto_filter.ref = ws.dimensions
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/download/license-companies", tags=["다운로드"])
def download_license_companies(
    license_name: str = Query(..., description="면허명 (예: 건축공사업)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터 (active_only/exclude_closed/all)"),
    limit: int = Query(5000, ge=1, le=10000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    rows = conn.execute(f"""
        SELECT cm.company_name, ci.canonical_business_no as bizno,
               cm.display_location, cm.location_detail,
               cm.is_headquarters{s_col},
               cl.license_name, cl.license_code,
               {_direct_production_summary_expr(valid_only=True)} as direct_production_summary
        FROM company_master cm
        JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
        JOIN company_license cl ON cm.company_internal_id = cl.company_internal_id
        {s_join}
        WHERE cl.license_name LIKE ? AND cm.is_busan_company = 1{s_where}
        ORDER BY cm.company_name LIMIT ?
    """, (f"%{license_name}%", limit)).fetchall()
    conn.close()
    
    columns = ["company_name", "bizno", "display_location", "location_detail",
               "is_headquarters", "nts_status", "license_name", "license_code",
               "direct_production_summary"]
    headers = ["업체명", "사업자번호", "소재지", "상세주소", "본사여부", "영업상태", "면허명", "면허코드",
               "직접생산증명"]
    
    from urllib.parse import quote
    fn = quote(f"면허별업체_{license_name}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


@app.get("/api/download/product-companies", tags=["다운로드"])
def download_product_companies(
    product_name: str = Query(..., description="물품명 (예: 컴퓨터)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터"),
    limit: int = Query(5000, ge=1, le=10000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    rows = conn.execute(f"""
        SELECT cm.company_name, ci.canonical_business_no as bizno,
               cm.display_location, cm.location_detail,
               cm.is_headquarters{s_col},
               cp.product_name, cp.product_code,
               {_direct_production_summary_expr(product_expr="cp.product_name")} as direct_production_summary
        FROM company_master cm
        JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
        JOIN company_product cp ON cm.company_internal_id = cp.company_internal_id
        {s_join}
        WHERE cp.product_name LIKE ? AND cm.is_busan_company = 1{s_where}
        ORDER BY cm.company_name LIMIT ?
    """, (f"%{product_name}%", limit)).fetchall()
    conn.close()
    
    columns = ["company_name", "bizno", "display_location", "location_detail",
               "is_headquarters", "nts_status", "product_name", "product_code",
               "direct_production_summary"]
    headers = ["업체명", "사업자번호", "소재지", "상세주소", "본사여부", "영업상태", "물품분류", "물품코드",
               "직접생산증명"]
    
    from urllib.parse import quote
    fn = quote(f"물품별업체_{product_name}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


@app.get("/api/download/policy-companies", tags=["다운로드"])
def download_policy_companies(
    policy_type: Optional[str] = Query(None, description="정책유형 (미지정시 전체)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터"),
    limit: int = Query(5000, ge=1, le=10000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    if policy_type:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location, cm.location_detail,
                   cm.is_headquarters{s_col},
                   pcc.policy_subtype, pcc.validity_status,
                   {_direct_production_summary_expr(valid_only=True)} as direct_production_summary
            FROM company_master cm
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            JOIN policy_company_certification pcc ON cm.company_internal_id = pcc.company_internal_id
            {s_join}
            WHERE pcc.policy_subtype LIKE ? AND cm.is_busan_company = 1{s_where}
            ORDER BY cm.company_name LIMIT ?
        """, (f"%{policy_type}%", limit)).fetchall()
    else:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location, cm.location_detail,
                   cm.is_headquarters{s_col},
                   pcc.policy_subtype, pcc.validity_status,
                   {_direct_production_summary_expr(valid_only=True)} as direct_production_summary
            FROM company_master cm
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            JOIN policy_company_certification pcc ON cm.company_internal_id = pcc.company_internal_id
            {s_join}
            WHERE cm.is_busan_company = 1{s_where}
            ORDER BY pcc.policy_subtype, cm.company_name LIMIT ?
        """, (limit,)).fetchall()
    conn.close()
    
    columns = ["company_name", "bizno", "display_location", "location_detail",
               "is_headquarters", "nts_status", "policy_subtype", "validity_status",
               "direct_production_summary"]
    headers = ["업체명", "사업자번호", "소재지", "상세주소", "본사여부", "영업상태", "정책유형", "인증상태",
               "직접생산증명"]
    
    from urllib.parse import quote
    fn = quote(f"정책업체_{policy_type or '전체'}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


@app.get("/api/download/shopping-mall-products", tags=["다운로드"])
def download_shopping_mall_products(
    contract_type: Optional[str] = Query(None, description="계약유형 (mas, general_unit_price 등)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터"),
    limit: int = Query(10000, ge=1, le=50000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    if contract_type:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location{s_col},
                   smp.product_name, smp.detail_product_name, smp.product_code,
                   smp.shopping_mall_contract_type, smp.contract_status,
                   smp.price_amount, smp.price_unit,
                   smp.contract_start_date, smp.contract_end_date,
                   {_direct_production_summary_expr(product_expr="IFNULL(NULLIF(smp.detail_product_name, ''), smp.product_name)")} as direct_production_summary
            FROM shopping_mall_product smp
            JOIN company_master cm ON smp.company_internal_id = cm.company_internal_id
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            {s_join}
            WHERE cm.is_busan_company = 1 AND smp.shopping_mall_contract_type = ?{s_where}
            ORDER BY cm.company_name LIMIT ?
        """, (contract_type, limit)).fetchall()
    else:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location{s_col},
                   smp.product_name, smp.detail_product_name, smp.product_code,
                   smp.shopping_mall_contract_type, smp.contract_status,
                   smp.price_amount, smp.price_unit,
                   smp.contract_start_date, smp.contract_end_date,
                   {_direct_production_summary_expr(product_expr="IFNULL(NULLIF(smp.detail_product_name, ''), smp.product_name)")} as direct_production_summary
            FROM shopping_mall_product smp
            JOIN company_master cm ON smp.company_internal_id = cm.company_internal_id
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            {s_join}
            WHERE cm.is_busan_company = 1{s_where}
            ORDER BY cm.company_name LIMIT ?
        """, (limit,)).fetchall()
    conn.close()
    
    columns = ["company_name", "bizno", "display_location", "nts_status",
               "product_name", "detail_product_name", "product_code",
               "shopping_mall_contract_type", "contract_status",
               "price_amount", "price_unit", "contract_start_date", "contract_end_date",
               "direct_production_summary"]
    headers = ["업체명", "사업자번호", "소재지", "영업상태",
               "물품분류", "세부물품", "물품코드",
               "계약유형", "계약상태", "단가(원)", "단위", "계약시작일", "계약종료일",
               "직접생산증명"]
    
    from urllib.parse import quote
    fn = quote(f"종합쇼핑몰_{contract_type or '전체'}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


@app.get("/api/download/certified-products", tags=["다운로드"])
def download_certified_products(
    cert_type: Optional[str] = Query(None, description="인증유형 (NEP, NET 등)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터"),
    limit: int = Query(10000, ge=1, le=50000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    if cert_type:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location{s_col},
                   cp.certification_type, cp.certification_type_label,
                   cp.product_name, cp.validity_status, cp.source_name,
                   {_direct_production_summary_expr(product_expr="cp.product_name")} as direct_production_summary
            FROM certified_product cp
            JOIN company_master cm ON cp.company_internal_id = cm.company_internal_id
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            {s_join}
            WHERE cm.is_busan_company = 1 AND cp.certification_type LIKE ?{s_where}
            ORDER BY cm.company_name LIMIT ?
        """, (f"%{cert_type}%", limit)).fetchall()
    else:
        rows = conn.execute(f"""
            SELECT cm.company_name, ci.canonical_business_no as bizno,
                   cm.display_location{s_col},
                   cp.certification_type, cp.certification_type_label,
                   cp.product_name, cp.validity_status, cp.source_name,
                   {_direct_production_summary_expr(product_expr="cp.product_name")} as direct_production_summary
            FROM certified_product cp
            JOIN company_master cm ON cp.company_internal_id = cm.company_internal_id
            JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
            {s_join}
            WHERE cm.is_busan_company = 1{s_where}
            ORDER BY cp.certification_type, cm.company_name LIMIT ?
        """, (limit,)).fetchall()
    conn.close()
    
    columns = ["company_name", "bizno", "display_location", "nts_status",
               "certification_type", "certification_type_label",
               "product_name", "validity_status", "source_name",
               "direct_production_summary"]
    headers = ["업체명", "사업자번호", "소재지", "영업상태",
               "인증유형", "인증유형명", "제품명", "유효상태", "출처",
               "직접생산증명"]
    
    from urllib.parse import quote
    fn = quote(f"인증제품_{cert_type or '전체'}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


@app.get("/api/download/direct-production-certs", tags=["다운로드"])
def download_direct_production_certs(
    product_name: Optional[str] = Query(None, description="세부품명 필터 (예: 데스크톱컴퓨터)"),
    validity_status: Optional[str] = Query("valid", description="유효상태 (valid/expired/unknown/all)"),
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터"),
    limit: int = Query(10000, ge=1, le=50000),
):
    s_join, s_where, s_col = _status_join_and_filter(status)
    params = []
    dpc_where = ""
    if product_name:
        dpc_where += " AND dpc.detail_product_name LIKE ?"
        params.append(f"%{product_name}%")
    if validity_status and validity_status != "all":
        dpc_where += " AND dpc.validity_status = ?"
        params.append(validity_status)
    params.append(limit)

    conn = _get_chatbot_db()
    rows = conn.execute(f"""
        SELECT cm.company_name,
               ci.canonical_business_no as bizno,
               cm.display_location,
               cm.location_detail,
               cm.is_headquarters{s_col},
               dpc.detail_product_name,
               dpc.detail_product_code,
               dpc.valid_from,
               dpc.valid_to,
               dpc.validity_status,
               dpc.source_refreshed_at,
               dpc.source_name
        FROM direct_production_certificate dpc
        JOIN company_master cm ON dpc.company_internal_id = cm.company_internal_id
        JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
        {s_join}
        WHERE cm.is_busan_company = 1{s_where}{dpc_where}
        ORDER BY dpc.detail_product_name, cm.company_name
        LIMIT ?
    """, tuple(params)).fetchall()
    conn.close()

    columns = ["company_name", "bizno", "display_location", "location_detail",
               "is_headquarters", "nts_status", "detail_product_name",
               "detail_product_code", "valid_from", "valid_to", "validity_status",
               "source_refreshed_at", "source_name"]
    headers = ["업체명", "사업자번호", "소재지", "상세주소",
               "본사여부", "영업상태", "세부품명",
               "세부품명번호", "유효기간 시작일", "유효기간 종료일", "유효상태",
               "원천갱신일", "출처"]

    from urllib.parse import quote
    fn = quote(f"직접생산증명_{product_name or '전체'}_{validity_status or 'all'}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)



@app.get("/api/download/all-companies", tags=["다운로드"])
def download_all_companies(
    status: DownloadStatusFilter = Query("active_only", description="영업상태 필터 (active_only/exclude_closed/all)"),
    limit: int = Query(50000, ge=1, le=50000),
):
    """부산 업체 통합 데이터 엑셀 다운로드"""
    s_join, s_where, s_col = _status_join_and_filter(status)
    conn = _get_chatbot_db()
    rows = conn.execute(f"""
        SELECT cm.company_name,
               ci.canonical_business_no as bizno,
               cm.display_location,
               cm.location_detail
               {s_col},
               (SELECT GROUP_CONCAT(cl.license_name, ', ')
                FROM company_license cl WHERE cl.company_internal_id = cm.company_internal_id) as licenses,
               (SELECT GROUP_CONCAT(cp.product_name, ', ')
                FROM company_product cp WHERE cp.company_internal_id = cm.company_internal_id) as products,
               (SELECT GROUP_CONCAT(pcc.policy_subtype, ', ')
                FROM policy_company_certification pcc WHERE pcc.company_internal_id = cm.company_internal_id) as policy_types,
               (SELECT GROUP_CONCAT(cert.certification_type, ', ')
                FROM certified_product cert WHERE cert.company_internal_id = cm.company_internal_id) as cert_types,
               (SELECT GROUP_CONCAT(smp.shopping_mall_contract_type, ', ')
                FROM shopping_mall_product smp
                WHERE smp.company_internal_id = cm.company_internal_id AND smp.contract_status = 'active') as shopping_mall_types,
               (SELECT COUNT(*)
                FROM shopping_mall_product smp
                WHERE smp.company_internal_id = cm.company_internal_id AND smp.contract_status = 'active') as shopping_mall_count,
               {_direct_production_count_expr(valid_only=True)} as direct_production_valid_count,
               {_direct_production_summary_expr(valid_only=True)} as direct_production_summary,
               IFNULL((SELECT manufacturer_type FROM company_manufacturer_status cms
                       WHERE cms.company_internal_id = cm.company_internal_id LIMIT 1), '') as manufacturer_type
        FROM company_master cm
        JOIN company_identity ci ON cm.company_internal_id = ci.company_internal_id
        {s_join}
        WHERE cm.is_busan_company = 1{s_where}
        ORDER BY cm.company_name
        LIMIT ?
    """, (limit,)).fetchall()
    conn.close()

    columns = ["company_name", "bizno", "display_location", "location_detail",
               "nts_status", "licenses", "products", "policy_types",
               "cert_types", "shopping_mall_types", "shopping_mall_count",
               "direct_production_valid_count", "direct_production_summary",
               "manufacturer_type"]
    headers = ["업체명", "사업자번호", "소재지", "상세주소",
               "영업상태", "보유면허", "등록물품", "정책업체유형",
               "인증유형", "쇼핑몰계약유형", "쇼핑몰등록건수",
               "직접생산증명 유효건수", "직접생산증명",
               "제조업체구분"]

    from urllib.parse import quote
    fn = quote(f"부산업체통합_{status}.xlsx")
    return _make_excel_response(rows, columns, headers, fn)


SOCIAL_ADMIN_UPLOAD_DIR = Path(__file__).resolve().parent / "import_sources" / "social_enterprise_admin_uploads"
SOCIAL_ADMIN_MAX_BYTES = int(os.environ.get("SOCIAL_ADMIN_MAX_UPLOAD_BYTES", str(20 * 1024 * 1024)))


def _social_admin_token() -> str:
    return os.environ.get("SOCIAL_ADMIN_TOKEN", "").strip()


def _require_social_admin_token(x_admin_token: Optional[str]) -> None:
    expected = _social_admin_token()
    if not expected:
        raise HTTPException(
            status_code=503,
            detail="SOCIAL_ADMIN_TOKEN is not configured on the server.",
        )
    supplied = (x_admin_token or "").strip()
    if not supplied or not secrets.compare_digest(supplied, expected):
        raise HTTPException(status_code=401, detail="Invalid admin token.")


def _safe_upload_filename(filename: str) -> str:
    name = Path(filename or "social_enterprise_upload.xlsx").name
    name = re.sub(r"[^0-9A-Za-z가-힣._ -]+", "_", name).strip(" .")
    if not name:
        name = "social_enterprise_upload.xlsx"
    suffix = Path(name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm files are supported.")
    return name


def _decode_upload_base64(file_base64: str) -> bytes:
    value = (file_base64 or "").strip()
    if "," in value and value.lower().startswith("data:"):
        value = value.split(",", 1)[1]
    try:
        data = base64.b64decode(value, validate=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid base64 payload: {exc}") from exc
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    if len(data) > SOCIAL_ADMIN_MAX_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Upload exceeds limit: {len(data)} bytes > {SOCIAL_ADMIN_MAX_BYTES} bytes.",
        )
    return data


def _social_table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def _social_admin_status_payload() -> dict:
    db_path = CHATBOT_DB
    payload = {
        "db_path": db_path,
        "db_exists": os.path.exists(db_path),
        "admin_token_configured": bool(_social_admin_token()),
        "source_counts": {},
        "policy_counts": {},
        "latest_imports": [],
        "source_manifest": [],
        "advisor_cache_sync": {
            "automatic": False,
            "required_after_apply": True,
            "reason": "busan-api runs as busan-monitor and cannot safely replace /opt/advisor cache archives.",
            "command": "cd /opt/advisor && python3 scripts/sync_company_view_db.py --source /opt/busan/chatbot_company.db --cache-root /opt/advisor/cache/company --keep-archives 3 --apply",
        },
    }
    if not payload["db_exists"]:
        return payload

    conn = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=30.0)
    conn.row_factory = sqlite3.Row
    try:
        if _social_table_exists(conn, "social_enterprise_master"):
            rows = conn.execute(
                """
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN record_status IN ('valid', 'active') THEN 1 ELSE 0 END) AS valid_total,
                    SUM(CASE WHEN record_status IN ('valid', 'active') AND region LIKE '%부산%' THEN 1 ELSE 0 END) AS valid_busan,
                    SUM(CASE WHEN record_status IN ('valid', 'active') AND matched_company_internal_id IS NOT NULL THEN 1 ELSE 0 END) AS matched_valid,
                    SUM(CASE WHEN record_status IN ('valid', 'active') AND matched_company_internal_id IS NULL THEN 1 ELSE 0 END) AS unmatched_valid,
                    SUM(CASE WHEN record_status IN ('valid', 'active') AND matched_company_internal_id IS NULL AND region LIKE '%부산%' THEN 1 ELSE 0 END) AS busan_unmatched_valid
                FROM social_enterprise_master
                """
            ).fetchone()
            payload["source_counts"] = dict(rows) if rows else {}

        if _social_table_exists(conn, "policy_company_certification"):
            rows = conn.execute(
                """
                SELECT
                    COUNT(*) AS total_social_rows,
                    COUNT(DISTINCT company_internal_id) AS social_company_count,
                    SUM(CASE WHEN validity_status='valid' THEN 1 ELSE 0 END) AS valid_social_rows,
                    COUNT(DISTINCT CASE WHEN validity_status='valid' THEN company_internal_id END) AS valid_social_company_count
                FROM policy_company_certification
                WHERE policy_subtype='social_enterprise'
                """
            ).fetchone()
            payload["policy_counts"] = dict(rows) if rows else {}

        if _social_table_exists(conn, "social_enterprise_import_log"):
            payload["latest_imports"] = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT import_batch_id, source_file_name, source_refreshed_at,
                           uploaded_by, started_at, finished_at, status,
                           total_count, source_busan_count, matched_count,
                           matched_busan_count, unmatched_count, policy_upsert_count,
                           expired_previous_count, error_message
                    FROM social_enterprise_import_log
                    ORDER BY started_at DESC
                    LIMIT 10
                    """
                ).fetchall()
            ]

        if _social_table_exists(conn, "source_manifest"):
            payload["source_manifest"] = [
                dict(row)
                for row in conn.execute(
                    """
                    SELECT source_name, source_type, source_url_or_file,
                           source_refreshed_at, updated_at AS last_imported_at,
                           row_count, checksum AS source_file_sha256, status
                    FROM source_manifest
                    WHERE source_name LIKE '%social%' OR source_url_or_file LIKE '%사회%'
                    ORDER BY updated_at DESC
                    LIMIT 10
                    """
                ).fetchall()
            ]
    finally:
        conn.close()
    return payload


@app.get("/api/admin/social-enterprise/status", tags=["관리자"])
def social_enterprise_admin_status(x_admin_token: Optional[str] = Header(default=None)):
    _require_social_admin_token(x_admin_token)
    return _social_admin_status_payload()


@app.post("/api/admin/social-enterprise/import", tags=["관리자"])
def social_enterprise_admin_import(
    payload: dict = Body(...),
    x_admin_token: Optional[str] = Header(default=None),
):
    _require_social_admin_token(x_admin_token)

    filename = _safe_upload_filename(str(payload.get("filename") or ""))
    file_data = _decode_upload_base64(str(payload.get("file_base64") or ""))
    apply_mode = bool(payload.get("apply"))
    uploaded_by = str(payload.get("uploaded_by") or "social_admin").strip() or "social_admin"
    source_refreshed_at = str(payload.get("source_refreshed_at") or "").strip()
    expire_previous_source = bool(payload.get("expire_previous_source", True))

    SOCIAL_ADMIN_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    upload_path = SOCIAL_ADMIN_UPLOAD_DIR / f"{stamp}_{filename}"
    upload_path.write_bytes(file_data)
    upload_path.chmod(0o640)

    try:
        import import_social_enterprise_excel as social_import

        args = argparse.Namespace(
            source_file=str(upload_path),
            db=CHATBOT_DB,
            source_name=social_import.DEFAULT_SOURCE_NAME,
            source_refreshed_at=source_refreshed_at,
            uploaded_by=uploaded_by,
            import_batch_id="",
            backup_dir=str(Path(CHATBOT_DB).resolve().parent / "backups"),
            apply=apply_mode,
            expire_previous_source=expire_previous_source,
            json=False,
        )
        result = social_import.run(args)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    result["status"] = "success"
    result["uploaded_file_path"] = str(upload_path)
    result["advisor_cache_sync"] = {
        "automatic": False,
        "required": apply_mode,
        "command": "cd /opt/advisor && python3 scripts/sync_company_view_db.py --source /opt/busan/chatbot_company.db --cache-root /opt/advisor/cache/company --keep-archives 3 --apply",
        "note": "업체추천 서비스 반영은 /opt/advisor 캐시 동기화 후 완료됩니다.",
    }
    result["social_purchase_cache_refresh"] = {
        "automatic": True,
        "required": apply_mode,
        "started": False,
        "note": "검증만 실행한 경우 사회적기업 수주율 캐시는 갱신하지 않습니다.",
    }
    if apply_mode:
        result["social_purchase_cache_refresh"] = _trigger_social_purchase_cache_refresh(
            "social_enterprise_admin_import"
        )
    return result


@app.get("/api/admin/social-enterprise", include_in_schema=False)
def social_enterprise_admin_page():
    return HTMLResponse(
        """
<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>사회적기업 관리자</title>
  <style>
    :root { --navy:#0f2f59; --blue:#1e63c7; --teal:#008b83; --line:#d8e2ef; --bg:#f4f8fb; --muted:#66758c; }
    body { margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Noto Sans KR",Arial,sans-serif; background:var(--bg); color:#07172e; }
    .wrap { max-width:1180px; margin:0 auto; padding:36px 24px 64px; }
    .hero { background:linear-gradient(135deg,#0f8aa3,#1757cc 58%,#061a7a); color:#fff; border-radius:20px; padding:30px 34px; box-shadow:0 16px 42px rgba(28,61,108,.22); }
    .hero h1 { margin:8px 0 10px; font-size:36px; letter-spacing:0; }
    .hero p { margin:0; font-size:16px; line-height:1.65; max-width:900px; opacity:.94; }
    .pill { display:inline-flex; align-items:center; gap:8px; padding:8px 14px; border-radius:999px; background:rgba(255,255,255,.14); border:1px solid rgba(255,255,255,.22); font-weight:800; }
    .grid { display:grid; grid-template-columns:1fr 1fr; gap:18px; margin-top:20px; }
    .card { background:#fff; border:1px solid var(--line); border-radius:16px; padding:22px; box-shadow:0 10px 30px rgba(36,72,108,.08); }
    .card h2 { margin:0 0 16px; font-size:22px; }
    .field { margin:14px 0; }
    label { display:block; font-weight:800; margin-bottom:8px; }
    input[type=password], input[type=text], input[type=datetime-local], input[type=file] { width:100%; box-sizing:border-box; border:1px solid #cfdbea; border-radius:10px; padding:12px 13px; font-size:15px; background:#fff; }
    .row { display:flex; gap:10px; flex-wrap:wrap; align-items:center; }
    button { border:0; border-radius:10px; padding:12px 18px; font-size:15px; font-weight:900; cursor:pointer; }
    button.primary { background:linear-gradient(135deg,#098680,#1e63c7); color:#fff; }
    button.secondary { background:#eaf2ff; color:#123d83; border:1px solid #c8ddff; }
    button.danger { background:#fff1e8; color:#b93e00; border:1px solid #ffd3b8; }
    button:disabled { opacity:.55; cursor:not-allowed; }
    .metrics { display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-top:14px; }
    .metric { background:#f8fbff; border:1px solid #dce7f5; border-radius:13px; padding:14px; }
    .metric b { display:block; font-size:28px; margin-top:4px; }
    .metric span { color:var(--muted); font-weight:700; }
    table { width:100%; border-collapse:collapse; font-size:14px; }
    th, td { border-bottom:1px solid #e4edf7; padding:10px 8px; text-align:left; vertical-align:top; }
    th { color:#44546c; background:#f8fbff; }
    .note { color:#52627a; line-height:1.65; }
    .warn { padding:12px 14px; background:#fff8e6; border:1px solid #ffe0a3; border-radius:12px; color:#6e4300; line-height:1.6; }
    .ok { padding:12px 14px; background:#eafaf3; border:1px solid #b8efd5; border-radius:12px; color:#075f3a; line-height:1.6; }
    @media (max-width: 900px) { .grid, .metrics { grid-template-columns:1fr; } .hero h1 { font-size:28px; } }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="hero">
      <div class="pill">사회적기업 원천DB 업로드</div>
      <h1>사회적기업 관리자</h1>
      <p>부산 소재 사회적기업 엑셀을 업로드해 업체추천 DB의 사회적기업 여부를 갱신합니다. 전국 원천 파일도 처리할 수 있지만, 운영 기준은 부산 소재 기업만 업로드하는 방식입니다. 조달업체 DB에 없는 부산 사회적기업은 별도 마스터에 보존해 향후 사회적기업 구매율 산정에 사용할 수 있습니다.</p>
    </section>

    <section class="grid">
      <div class="card">
        <h2>1. 관리자 인증</h2>
        <div class="field">
          <label for="token">관리자 토큰</label>
          <input id="token" type="password" placeholder="SOCIAL_ADMIN_TOKEN">
        </div>
        <div class="row">
          <button class="secondary" onclick="saveToken()">토큰 저장</button>
          <button class="secondary" onclick="loadStatus()">상태 조회</button>
        </div>
        <p class="note">토큰은 브라우저 localStorage에만 저장됩니다. 서버에는 노출되지 않습니다.</p>
      </div>

      <div class="card">
        <h2>2. 엑셀 업로드</h2>
        <div class="field">
          <label for="file">사회적기업 XLSX</label>
          <input id="file" type="file" accept=".xlsx,.xlsm">
        </div>
        <div class="field">
          <label for="refreshedAt">원천 기준일</label>
          <input id="refreshedAt" type="datetime-local">
        </div>
        <div class="field">
          <label for="uploadedBy">업로드 담당자</label>
          <input id="uploadedBy" type="text" value="social_admin">
        </div>
        <div class="row">
          <button class="secondary" onclick="runImport(false)">검증만 실행</button>
          <button class="primary" onclick="runImport(true)">DB 반영</button>
        </div>
      </div>
    </section>

    <section class="card" style="margin-top:20px;">
      <h2>현재 상태</h2>
      <div id="statusMessage" class="warn">상태 조회 전입니다.</div>
      <div class="metrics" id="metrics"></div>
    </section>

    <section class="card" style="margin-top:20px;">
      <h2>최근 임포트 이력</h2>
      <div id="logs"></div>
    </section>

  </main>

  <script>
    const tokenInput = document.getElementById('token');
    tokenInput.value = localStorage.getItem('social_admin_token') || '';

    function saveToken() {
      localStorage.setItem('social_admin_token', tokenInput.value.trim());
      showStatus('ok', '토큰을 저장했습니다. 상태 조회 또는 엑셀 검증을 실행하세요.');
    }

    function tokenHeaders() {
      const token = tokenInput.value.trim() || localStorage.getItem('social_admin_token') || '';
      return { 'X-Admin-Token': token };
    }

    function fmt(v) {
      if (v === null || v === undefined || v === '') return '-';
      if (typeof v === 'number') return v.toLocaleString('ko-KR');
      return String(v);
    }

    function toLocalDatetimeText(value) {
      if (!value) return '';
      return value.replace('T', ' ') + (value.length === 16 ? ':00' : '');
    }

    function showStatus(className, text) {
      const msg = document.getElementById('statusMessage');
      msg.className = className;
      msg.textContent = text;
    }

    async function loadStatus(successMessage = null) {
      showStatus('warn', '상태 조회 중입니다.');
      const res = await fetch('/api/admin/social-enterprise/status', { headers: tokenHeaders() });
      const data = await res.json();
      if (!res.ok) {
        showStatus('warn', data.detail || '상태 조회 실패');
        return;
      }
      renderStatus(data, successMessage);
    }

    function renderStatus(data, successMessage = null) {
      const msg = document.getElementById('statusMessage');
      msg.className = 'ok';
      msg.textContent = successMessage || 'DB 연결 정상. 사회적기업 원천DB 및 정책기업 반영 상태를 조회했습니다.';
      const s = data.source_counts || {};
      const p = data.policy_counts || {};
      document.getElementById('metrics').innerHTML = [
        ['사회적기업 원천 유효건', s.valid_total],
        ['부산 소재 유효건', s.valid_busan],
        ['조달업체 매칭', s.matched_valid],
        ['부산 미매칭 보존', s.busan_unmatched_valid],
        ['정책기업 유효 업체', p.valid_social_company_count],
        ['정책기업 전체 업체', p.social_company_count],
        ['DB 파일', data.db_exists ? '있음' : '없음'],
        ['관리자 토큰', data.admin_token_configured ? '설정됨' : '미설정'],
      ].map(([label, value]) => `<div class="metric"><span>${label}</span><b>${fmt(value)}</b></div>`).join('');

      const rows = data.latest_imports || [];
      if (!rows.length) {
        document.getElementById('logs').innerHTML = '<p class="note">임포트 이력이 없습니다.</p>';
        return;
      }
      document.getElementById('logs').innerHTML = `<table><thead><tr>
        <th>일시</th><th>파일</th><th>상태</th><th>전체</th><th>부산</th><th>매칭</th><th>정책반영</th><th>오류</th>
      </tr></thead><tbody>${rows.map(r => `<tr>
        <td>${fmt(r.started_at)}</td><td>${fmt(r.source_file_name)}</td><td>${fmt(r.status)}</td>
        <td>${fmt(r.total_count)}</td><td>${fmt(r.source_busan_count)}</td>
        <td>${fmt(r.matched_count)}</td><td>${fmt(r.policy_upsert_count)}</td>
        <td>${fmt(r.error_message)}</td>
      </tr>`).join('')}</tbody></table>`;
    }

    async function fileToBase64(file) {
      const buffer = await file.arrayBuffer();
      let binary = '';
      const bytes = new Uint8Array(buffer);
      const chunk = 0x8000;
      for (let i = 0; i < bytes.length; i += chunk) {
        binary += String.fromCharCode.apply(null, bytes.subarray(i, i + chunk));
      }
      return btoa(binary);
    }

    async function runImport(apply) {
      const file = document.getElementById('file').files[0];
      if (!file) {
        showStatus('warn', '업로드할 XLSX 파일을 선택하세요.');
        return;
      }
      if (apply && !confirm('DB에 반영합니다. 계속할까요?')) return;
      showStatus('warn', apply ? 'DB 반영 중입니다.' : '검증 실행 중입니다.');
      const payload = {
        filename: file.name,
        file_base64: await fileToBase64(file),
        source_refreshed_at: toLocalDatetimeText(document.getElementById('refreshedAt').value),
        uploaded_by: document.getElementById('uploadedBy').value || 'social_admin',
        apply,
        expire_previous_source: true,
      };
      const res = await fetch('/api/admin/social-enterprise/import', {
        method: 'POST',
        headers: { ...tokenHeaders(), 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      const data = await res.json();
      if (!res.ok) {
        showStatus('warn', data.detail || (apply ? 'DB 반영 실패' : '검증 실패'));
        return;
      }
      const total = fmt(data.total ?? data.total_count);
      const busan = fmt(data.source_busan ?? data.source_busan_count ?? data.busan);
      const matched = fmt(data.matched ?? data.matched_count);
      const policy = fmt(data.policy_upsert ?? data.policy_upsert_count);
      const summary = `${apply ? 'DB 반영' : '검증'} 완료: 전체 ${total}건, 부산 ${busan}건, 조달업체 매칭 ${matched}건, 정책기업 반영 ${policy}건`;
      await loadStatus(summary);
    }
  </script>
</body>
</html>
        """
    )


if __name__ == '__main__':
    import uvicorn
    print("[API] 부산 조달 모니터링 API 서버 시작")
    print("   http://localhost:8000/docs")
    uvicorn.run(app, host="0.0.0.0", port=8000)
